import csv
import io
import json
import re
import tempfile
import unicodedata
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from dateutil import parser as dtparser
from openpyxl import load_workbook


REQUIRED_COLUMNS = [
    "Timestamp",
    "ISP",
    "SITE / Location",
    "Incident DATE",
    "Incident TIME",
    "Reporter",
    "Incindent Type",
    "Severity",
    "Incident description",
    "Initial Cause",
    "Troubleshooting actions",
    "Root Cause",
    "Incident resolution date",
    "Incident resolution time",
    "Measures to prevent recurrence of incidents (if any)",
]


def _normalize_header(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", s)


HEADER_ALIASES = {
    _normalize_header("timestamp"): "Timestamp",
    _normalize_header("time stamp"): "Timestamp",
    _normalize_header("isp"): "ISP",
    _normalize_header("site/location"): "SITE / Location",
    _normalize_header("site / location"): "SITE / Location",
    _normalize_header("site"): "SITE / Location",
    _normalize_header("incidentdate"): "Incident DATE",
    _normalize_header("incident date"): "Incident DATE",
    _normalize_header("incidenttime"): "Incident TIME",
    _normalize_header("incident time"): "Incident TIME",
    _normalize_header("reporter"): "Reporter",
    _normalize_header("incindenttype"): "Incindent Type",
    _normalize_header("incidenttype"): "Incindent Type",
    _normalize_header("severity"): "Severity",
    _normalize_header("incidentdescription"): "Incident description",
    _normalize_header("initialcause"): "Initial Cause",
    _normalize_header("troubleshootingactions"): "Troubleshooting actions",
    _normalize_header("rootcause"): "Root Cause",
    _normalize_header("incidentresolutiondate"): "Incident resolution date",
    _normalize_header("incidentresolutiontime"): "Incident resolution time",
    _normalize_header(
        "measures to prevent recurrence of incidents (if any)"
    ): "Measures to prevent recurrence of incidents (if any)",
    _normalize_header(
        "measures to prevent recurrence of incidents"
    ): "Measures to prevent recurrence of incidents (if any)",
}


def _clean_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date, time)):
        return str(v)

    def _mojibake_score(txt: str) -> int:
        # Common markers when UTF-8 bytes were decoded as latin1/cp1252.
        markers = ["Ãƒ", "Ã‚", "Ã„", "Ã", "Ã‘", "Ã¢", "Ã¡Âº", "ï¿½"]
        return sum(txt.count(m) for m in markers)

    s = str(v).replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s).strip()

    try:
        if _mojibake_score(s) > 0:
            repaired_candidates: List[str] = []
            for enc in ("latin1", "cp1252"):
                try:
                    repaired = s.encode(enc, errors="strict").decode(
                        "utf-8", errors="strict"
                    )
                    repaired_candidates.append(repaired)
                except Exception:
                    continue
            for repaired in repaired_candidates:
                if (
                    repaired
                    and _mojibake_score(repaired) < _mojibake_score(s)
                    and "ï¿½" not in repaired
                ):
                    s = repaired
                    break
    except Exception:
        pass

    return s


def _fold_text(v: Any) -> str:
    s = _clean_text(v).lower()
    normalized = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _resolve_file_path(file_obj: Any) -> str:
    if isinstance(file_obj, (str, Path)):
        return str(file_obj)
    if isinstance(file_obj, dict):
        for key in (
            "path",
            "file_path",
            "filepath",
            "file",
            "filename",
            "file_name",
            "name",
        ):
            if key in file_obj and file_obj[key]:
                return str(file_obj[key])
        if file_obj.get("id"):
            return str(file_obj["id"])
    return str(file_obj)


def _is_blank_row_any(values: Iterable[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


def _unknown_if_blank(v: Any) -> str:
    s = _clean_text(v)
    return s if s else "UNKNOWN"


RESPONSIBILITY_RULES = [
    (
        "ISP",
        [
            "routing",
            "uplink",
            "transit",
            "fiber",
            "fibre",
            "cable",
            "cap",
            "provider",
            "vnpt",
            "fpt",
            "cmc",
            "netnam",
        ],
    ),
    (
        "INHOUSE",
        [
            "local segment",
            "internal",
            "switch",
            "firewall",
            "configuration",
            "linecard",
            "junos",
            "wifi controller",
            "access point",
        ],
    ),
    (
        "EXTERNAL",
        [
            "electricity",
            "utility",
            "power supply",
            "third-party",
            "third party",
        ],
    ),
]


def _classify_responsibility_domain(root_cause: Any) -> Tuple[str, str]:
    text = _fold_text(root_cause)
    if not text or text == "unknown":
        return "UNKNOWN", "RCA does not support responsibility classification"
    for domain, keywords in RESPONSIBILITY_RULES:
        for keyword in keywords:
            if keyword in text:
                return domain, f"RCA keyword rule: {keyword}"
    return "UNKNOWN", "RCA does not support responsibility classification"


def _duration_minutes(started_at: datetime, resolved_at: Optional[datetime]) -> Optional[int]:
    if resolved_at is None:
        return None
    try:
        return int(max(0, (resolved_at - started_at).total_seconds()) // 60)
    except Exception:
        return None


def _looks_like_html(s: str) -> bool:
    t = s.lstrip().lower()
    if t.startswith("<!doctype html") or t.startswith("<html"):
        return True
    if "<head" in t and "<body" in t:
        return True
    if "accounts.google.com" in t:
        return True
    if "<form" in t and "password" in t:
        return True
    return False


def _make_markdown_file(md_text: Optional[str]) -> Any:
    """Return a file-like markdown payload for Alpha upload nodes.

    In the Alpha runtime, returning a dict or path string can produce a tiny text
    wrapper instead of the real Markdown file. Returning BytesIO keeps the actual
    content as the upload payload while still writing a temp copy for traceability.
    """

    if md_text is None:
        md_text = ""

    md_bytes = md_text.encode("utf-8")
    out_path: Optional[Path] = None
    last_err: Optional[Exception] = None
    for base in (Path(tempfile.gettempdir()), Path.cwd()):
        try:
            base.mkdir(parents=True, exist_ok=True)
            candidate = base / "02_confirmed_incidents.md"
            candidate.write_bytes(md_bytes)
            if not candidate.exists():
                raise ValueError("output file missing after write")
            if candidate.stat().st_size != len(md_bytes):
                raise ValueError(
                    f"output file size mismatch (expected={len(md_bytes)}, got={candidate.stat().st_size})"
                )
            out_path = candidate
            break
        except Exception as e:
            last_err = e

    if out_path is None:
        raise ValueError(f"Failed to write confirmed incidents markdown file: {last_err}")

    bio = io.BytesIO(md_bytes)
    try:
        bio.name = "02_confirmed_incidents.md"  # type: ignore[attr-defined]
    except Exception:
        pass
    bio.seek(0)
    return bio


def _safe_parse_dt_text(v: Any) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = _clean_text(v)
    if not s:
        return None
    try:
        return dtparser.parse(s, dayfirst=False, fuzzy=True)
    except Exception:
        return None


def _build_error(
    reason: str, source_type: str, rejected: Optional[List[Dict[str, Any]]] = None
) -> Dict[str, Any]:
    json_result = {
        "ok": False,
        "source_type": source_type,
        "input_rows": 0,
        "record_count": 0,
        "rejected_row_count": 0,
        "skipped_blank_rows": 0,
        "records": [],
        "rejected_rows": rejected
        if rejected is not None
        else [
            {
                "row_number": None,
                "source_ref": "input",
                "reason": reason,
                "row_data": {},
            }
        ],
    }
    return {
        "json_text": json.dumps(json_result, indent=2, ensure_ascii=False, default=str),
        "markdown": "",
        "confirmed_incidents_file": _make_markdown_file(""),
        "json_result": json_result,
        "summary": {
            "ok": False,
            "input_rows": 0,
            "record_count": 0,
            "rejected_row_count": len(json_result["rejected_rows"]),
            "skipped_blank_rows": 0,
        },
    }


def _parse_datetime_pair(d: Any, t: Any) -> Optional[datetime]:
    if (d is None or _clean_text(d) == "") and (t is None or _clean_text(t) == ""):
        return None

    if isinstance(d, datetime):
        base = d
        if isinstance(t, time):
            return datetime.combine(base.date(), t)
        if isinstance(t, datetime):
            return t
        t_txt = _clean_text(t)
        if t_txt:
            try:
                return datetime.combine(base.date(), dtparser.parse(t_txt).time())
            except Exception:
                return base
        return base

    if isinstance(t, datetime):
        return t

    if isinstance(d, date) and not isinstance(d, datetime):
        d_val = d
    else:
        d_txt = _clean_text(d)
        d_val = dtparser.parse(d_txt, dayfirst=False, fuzzy=True).date() if d_txt else None

    if isinstance(t, time):
        t_val = t
    else:
        t_txt = _clean_text(t)
        if t_txt:
            parsed = dtparser.parse(t_txt, dayfirst=False, fuzzy=True)
            if isinstance(parsed, datetime):
                if d_val is None:
                    return parsed
                return datetime.combine(d_val, parsed.time())
            t_val = parsed.time()
        else:
            t_val = None

    if d_val is None and t_val is None:
        return None
    if d_val is None:
        raise ValueError("Missing date component")
    if t_val is None:
        t_val = time(0, 0, 0)
    return datetime.combine(d_val, t_val)


def _split_site(site_raw: str) -> Tuple[str, str]:
    s = _clean_text(site_raw)
    if not s:
        return "UNKNOWN", "UNKNOWN"
    m = re.match(r"^\s*([A-Za-z0-9]{2,12})\s*[-:/\\|]+\s*(.+)$", s)
    if m:
        return m.group(1).upper(), m.group(2).strip()
    m2 = re.match(r"^\s*([A-Z0-9]{2,12})\s+(.+)$", s)
    if m2:
        return m2.group(1).upper(), m2.group(2).strip()
    return "UNKNOWN", s


def _process_tabular_rows(
    *,
    rows: Iterable[Sequence[Any]],
    source_type: str,
    source_label: str,
) -> Dict[str, Any]:
    records: List[Dict[str, Any]] = []
    rejected_rows: List[Dict[str, Any]] = []
    skipped_blank_rows = 0
    input_rows = 0

    header_row_idx: Optional[int] = None
    header_cells: Optional[Tuple[Any, ...]] = None
    col_index_by_canonical: Dict[str, int] = {}

    rows_iter = iter(rows)
    for i, row in enumerate(rows_iter, start=1):
        if _is_blank_row_any(row):
            continue
        header_row_idx = i
        header_cells = tuple(row)
        for idx, h in enumerate(header_cells, start=1):
            canonical = HEADER_ALIASES.get(_normalize_header(h))
            if canonical:
                col_index_by_canonical[canonical] = idx
        break

    if header_row_idx is None or header_cells is None:
        return _build_error(
            "No header row found",
            source_type,
            rejected=[
                {
                    "row_number": None,
                    "source_ref": source_label,
                    "reason": "No header row found",
                    "row_data": {},
                }
            ],
        )

    missing = [c for c in REQUIRED_COLUMNS if c not in col_index_by_canonical]
    if missing:
        return _build_error(
            "Missing required columns",
            source_type,
            rejected=[
                {
                    "row_number": header_row_idx,
                    "source_ref": f"{source_label}#row:{header_row_idx}",
                    "reason": f"Missing required columns: {', '.join(missing)}",
                    "row_data": {"headers": [str(x) if x is not None else "" for x in header_cells]},
                }
            ],
        )

    def cell(row_tuple: Tuple[Any, ...], canon: str) -> Any:
        idx0 = col_index_by_canonical[canon] - 1
        return row_tuple[idx0] if idx0 < len(row_tuple) else None

    for row_idx, row in enumerate(rows_iter, start=header_row_idx + 1):
        row_tuple = tuple(row)
        if _is_blank_row_any(row_tuple):
            skipped_blank_rows += 1
            continue

        input_rows += 1
        source_ref = (
            f"google_sheet_csv:row-{row_idx}"
            if source_type == "csv"
            else f"{source_label}#row:{row_idx}"
        )

        try:
            started_at = _parse_datetime_pair(
                cell(row_tuple, "Incident DATE"), cell(row_tuple, "Incident TIME")
            )
            if started_at is None:
                raise ValueError("Missing incident date/time")
        except Exception as e:
            rejected_rows.append(
                {
                    "row_number": row_idx,
                    "source_ref": source_ref,
                    "reason": f"Invalid incident datetime: {e}",
                    "row_data": {
                        "Incident DATE": _clean_text(cell(row_tuple, "Incident DATE")),
                        "Incident TIME": _clean_text(cell(row_tuple, "Incident TIME")),
                    },
                }
            )
            continue

        resolved_at = None
        res_date = cell(row_tuple, "Incident resolution date")
        res_time = cell(row_tuple, "Incident resolution time")
        if _clean_text(res_date) != "" or _clean_text(res_time) != "":
            try:
                resolved_at = _parse_datetime_pair(res_date, res_time)
            except Exception as e:
                rejected_rows.append(
                    {
                        "row_number": row_idx,
                        "source_ref": source_ref,
                        "reason": f"Invalid resolution datetime: {e}",
                        "row_data": {
                            "Incident resolution date": _clean_text(res_date),
                            "Incident resolution time": _clean_text(res_time),
                        },
                    }
                )
                continue

        site_code, site_name = _split_site(_clean_text(cell(row_tuple, "SITE / Location")))
        site_code_value = site_code or "UNKNOWN"
        site_name_value = site_name or "UNKNOWN"
        isp_value = _unknown_if_blank(cell(row_tuple, "ISP"))
        incident_type_value = _unknown_if_blank(cell(row_tuple, "Incindent Type"))
        reporter_value = _unknown_if_blank(cell(row_tuple, "Reporter"))
        initial_cause_value = _unknown_if_blank(cell(row_tuple, "Initial Cause"))

        description_parts: List[str] = []
        desc = _clean_text(cell(row_tuple, "Incident description"))
        if desc:
            description_parts.append(desc)
        # Keep initial_cause as its own field instead of merging it into description.
        if reporter_value != "UNKNOWN":
            description_parts.append(f"Reporter: {reporter_value}")

        ts_dt = _safe_parse_dt_text(cell(row_tuple, "Timestamp"))
        ts_txt = _clean_text(cell(row_tuple, "Timestamp"))
        if ts_dt is not None:
            description_parts.append(f"Reported at (form timestamp): {ts_dt.isoformat()}")
        elif ts_txt:
            description_parts.append(f"Reported at (form timestamp): {ts_txt}")

        description_joined = "\n".join(description_parts).strip() or "UNKNOWN"

        root_cause_value = _unknown_if_blank(cell(row_tuple, "Root Cause"))
        responsibility_domain, responsibility_basis = _classify_responsibility_domain(
            root_cause_value
        )

        record = {
            "record_type": "CONFIRMED_INCIDENT",
            "incident_id": f"INC-{row_idx}",
            "source_ref": source_ref,
            "submitted_at": ts_dt.isoformat() if ts_dt else "",
            "started_at": started_at.isoformat(),
            "resolved_at": resolved_at.isoformat() if resolved_at else "",
            "duration_minutes": _duration_minutes(started_at, resolved_at),
            "site_code": site_code_value,
            "site_name": site_name_value,
            "isp": isp_value,
            "reporter": reporter_value,
            "incident_type": incident_type_value,
            "severity": _unknown_if_blank(cell(row_tuple, "Severity")),
            "description": description_joined,
            "initial_cause": initial_cause_value,
            "root_cause": root_cause_value,
            "troubleshooting_actions": _unknown_if_blank(
                cell(row_tuple, "Troubleshooting actions")
            ),
            "preventive_action": _unknown_if_blank(
                cell(row_tuple, "Measures to prevent recurrence of incidents (if any)")
            ),
            "resolution_status": "RESOLVED" if resolved_at else "UNKNOWN",
            "responsibility_domain": responsibility_domain,
            "responsibility_basis": responsibility_basis,
            "recurrence_key": "|".join(
                [site_code_value.lower(), isp_value.lower(), incident_type_value.lower()]
            ),
            "evidence_label": "SOURCE FACT",
            "evidence_note": f"Incident row {row_idx}",
        }

        minimal = [
            record["site_name"],
            record["description"],
            record["incident_type"],
            record["severity"],
        ]
        if all(_clean_text(x) in ("", "UNKNOWN") for x in minimal):
            skipped_blank_rows += 1
            continue

        records.append(record)

    json_result = {
        "ok": True,
        "source_type": source_type,
        "input_rows": input_rows,
        "record_count": len(records),
        "rejected_row_count": len(rejected_rows),
        "skipped_blank_rows": skipped_blank_rows,
        "records": records,
        "rejected_rows": rejected_rows,
    }

    reconciliation_passed = len(records) + len(rejected_rows) == input_rows
    if len(records) > 0 and len(rejected_rows) == 0 and reconciliation_passed:
        validation_status = "PASS"
    elif reconciliation_passed and (len(records) > 0 or len(rejected_rows) > 0):
        validation_status = "WARN"
    else:
        validation_status = "FAIL"

    md_lines: List[str] = [
        "# Incident Records",
        "",
        "```yaml",
        "record_type: CONFIRMED_INCIDENT_SOURCE",
        f"source_type: {source_type}",
        f"source: {source_label}",
        f"record_count: {len(records)}",
        f"rejected_row_count: {len(rejected_rows)}",
        f"reconciliation_passed: {str(reconciliation_passed).lower()}",
        f"validation_status: {validation_status}",
        "```",
        "",
    ]
    for rec in records:
        def _md_scalar(val: Any) -> str:
            s = _clean_text(val)
            return s if s else "UNKNOWN"

        def _md_block(val: Any) -> str:
            s = _clean_text(val)
            if not s:
                s = "UNKNOWN"
            return s.replace("\n", "\n  ")

        md_lines.extend(
            [
                f"## Incident {rec['incident_id']}",
                "",
                "```yaml",
                "record_type: CONFIRMED_INCIDENT",
                f"incident_id: {rec['incident_id']}",
                f"source_ref: {_md_scalar(rec.get('source_ref'))}",
                f"submitted_at: {_md_scalar(rec.get('submitted_at'))}",
                f"started_at: {_md_scalar(rec.get('started_at'))}",
                f"resolved_at: {_md_scalar(rec.get('resolved_at'))}",
                f"duration_minutes: {rec['duration_minutes'] if rec['duration_minutes'] is not None else 'UNKNOWN'}",
                f"site_code: {_md_scalar(rec.get('site_code'))}",
                f"site_name: {_md_scalar(rec.get('site_name'))}",
                f"isp: {_md_scalar(rec.get('isp'))}",
                f"reporter: {_md_scalar(rec.get('reporter'))}",
                f"incident_type: {_md_scalar(rec.get('incident_type'))}",
                f"severity: {_md_scalar(rec.get('severity'))}",
                f"resolution_status: {_md_scalar(rec.get('resolution_status'))}",
                f"responsibility_domain: {_md_scalar(rec.get('responsibility_domain'))}",
                f"responsibility_basis: {_md_scalar(rec.get('responsibility_basis'))}",
                f"recurrence_key: {_md_scalar(rec.get('recurrence_key'))}",
                f"evidence_label: {_md_scalar(rec.get('evidence_label'))}",
                f"evidence_note: {_md_scalar(rec.get('evidence_note'))}",
                "initial_cause: |",
                "  " + _md_block(rec.get("initial_cause")),
                "root_cause: |",
                "  " + _md_block(rec.get("root_cause")),
                "troubleshooting_actions: |",
                "  " + _md_block(rec.get("troubleshooting_actions")),
                "preventive_action: |",
                "  " + _md_block(rec.get("preventive_action")),
                "description: |",
                "  " + _md_block(rec.get("description")),
                "```",
                "",
            ]
        )

    summary = {
        "ok": True,
        "input_rows": input_rows,
        "record_count": len(records),
        "rejected_row_count": len(rejected_rows),
        "skipped_blank_rows": skipped_blank_rows,
        "reconciliation_passed": reconciliation_passed,
        "upload_ready": len(rejected_rows) == 0 and len(records) > 0,
    }
    md_text = "\n".join(md_lines).strip() + "\n"
    md_bytes = md_text.encode("utf-8")

    if len(records) > 0:
        if md_text.strip() == "" or len(md_bytes) < 50:
            raise ValueError(
                "Unexpectedly tiny/missing markdown content for confirmed incidents "
                f"(records={len(records)}, bytes={len(md_bytes)})."
            )

    confirmed_incidents_file = _make_markdown_file(md_text)

    try:
        payload_bytes = (
            confirmed_incidents_file.getvalue()
            if hasattr(confirmed_incidents_file, "getvalue")
            else bytes(confirmed_incidents_file)
        )
    except Exception as e:
        raise ValueError(f"Failed to validate confirmed incidents file payload bytes: {e}")

    if len(records) > 0 and len(payload_bytes) < 50:
        raise ValueError(
            "Confirmed incidents file payload is unexpectedly tiny "
            f"(records={len(records)}, bytes={len(payload_bytes)})."
        )

    if len(payload_bytes) != len(md_bytes):
        raise ValueError(
            "Confirmed incidents file payload size mismatch "
            f"(expected={len(md_bytes)}, got={len(payload_bytes)})."
        )

    return {
        "json_text": json.dumps(json_result, indent=2, ensure_ascii=False),
        "markdown": md_text,
        "confirmed_incidents_file": confirmed_incidents_file,
        "json_result": json_result,
        "summary": summary,
    }


def main(incident_report_file=None, incident_report_body=None):
    if incident_report_body is not None and _clean_text(incident_report_body) != "":
        if isinstance(incident_report_body, (bytes, bytearray)):
            try:
                body_text = bytes(incident_report_body).decode("utf-8-sig", errors="replace")
            except Exception:
                body_text = bytes(incident_report_body).decode("latin1", errors="replace")
        else:
            body_text = str(incident_report_body)
            if body_text.startswith("\ufeff"):
                body_text = body_text.lstrip("\ufeff")

        if _looks_like_html(body_text):
            return _build_error(
                "HTTP body looks like HTML/login page; expected CSV export",
                "csv",
                rejected=[
                    {
                        "row_number": None,
                        "source_ref": "http_body",
                        "reason": "HTML content detected in HTTP body",
                        "row_data": {"body_prefix": body_text[:2000]},
                    }
                ],
            )

        reader = csv.reader(io.StringIO(body_text), delimiter=",", quotechar='"')
        return _process_tabular_rows(
            rows=reader,
            source_type="csv",
            source_label="google_sheet_csv",
        )

    if incident_report_file:
        try:
            path = _resolve_file_path(incident_report_file)
            p = Path(path)
            if not p.exists():
                alt = Path.cwd() / path
                if alt.exists():
                    p = alt
            if p.exists():
                wb = load_workbook(filename=str(p), data_only=True, read_only=True)
                try:
                    sheet_name = "Form Responses 1" if "Form Responses 1" in wb.sheetnames else wb.active.title
                    ws = wb[sheet_name]
                    return _process_tabular_rows(
                        rows=ws.iter_rows(values_only=True),
                        source_type="xlsx",
                        source_label=f"{p.name}:{sheet_name}",
                    )
                finally:
                    try:
                        wb.close()
                    except Exception:
                        pass
        except Exception as e:
            return _build_error(
                f"Failed to read XLSX file: {e}",
                "xlsx",
                rejected=[
                    {
                        "row_number": None,
                        "source_ref": "incident_report_file",
                        "reason": f"Failed to read XLSX file: {e}",
                        "row_data": {"incident_report_file": str(incident_report_file)},
                    }
                ],
            )

    return _build_error(
        "No valid CSV body or XLSX file found",
        "csv",
        rejected=[
            {
                "row_number": None,
                "source_ref": "input",
                "reason": "No valid CSV body or XLSX file found",
                "row_data": {
                    "incident_report_file": str(incident_report_file)
                    if incident_report_file is not None
                    else None,
                    "incident_report_body_type": type(incident_report_body).__name__
                    if incident_report_body is not None
                    else None,
                    "incident_report_body_len": len(incident_report_body)
                    if isinstance(incident_report_body, (str, bytes, bytearray))
                    else None,
                },
            }
        ],
    )
