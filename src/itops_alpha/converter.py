from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook


PACKAGE_VERSION = "1.3.0"
ISSUE_REPORT_FILENAME = "IssueReport.xlsx"
INCIDENT_REPORT_FILENAME = "SEA - Corp IT- ILL- Incident Report   (Responses).xlsx"
ZABBIX_FILENAME = "zbx_problems_export.xlsx"
ZABBIX_EXPORT_DIRNAME = "Export zabbix"
INCIDENT_REPORT_DIRNAME = "ISP Incident Report"
ISSUE_REPORT_DIRNAME = "Ticket"

ISSUE_LEGACY_REQUIRED_COLUMNS = {
    "itcenter.ticket.comment.created_at",
    "itcenter.ticket.id",
    "itcenter.ticket.category_en_name",
    "itcenter.ticket.classification_name",
    "itcenter.ticket.full_title",
    "itcenter.ticket.service_recipient.user.business_unit",
    "itcenter.ticket.service_recipient.user.department_name",
    "itcenter.ticket.service_recipient.user.name",
    "itcenter.ticket.service_recipient.user.email",
    "itcenter.ticket.assignee.user.name",
    "itcenter.ticket.comment.text",
}

ISSUE_MONTHLY_REQUIRED_COLUMNS = {
    "itcenter.ticket.id",
    "itcenter.ticket.created_at",
    "itcenter.ticket.full_title",
    "itcenter.ticket.category_en_name",
    "itcenter.ticket.classification_name",
    "itcenter.ticket.service_recipient.user.business_unit",
    "itcenter.ticket.service_recipient.user.entity_name",
    "itcenter.ticket.location_full_name",
    "itcenter.ticket.comment.created_by.user.email",
    "itcenter.ticket.comment.text",
    "itcenter.ticket.comment.updated_at",
    "itcenter.ticket.comment.created_at",
    "itcenter.ticket.service_recipient.user.office_display",
    "itcenter.ticket.subclassification_name",
}

ISSUE_REQUIRED_COLUMNS = ISSUE_LEGACY_REQUIRED_COLUMNS

INCIDENT_REQUIRED_COLUMNS = {
    "Timestamp",
    "ISP",
    "SITE / Location",
    "Incident DATE",
    "Incident TIME",
    "Reporter",
    "Incindent Type",
    "Severity",
    "Incident description",
    "Initial Cause",
    "Troubleshooting actions",
    "Root Cause",
    "Incident resolution date",
    "Incident resolution time",
    "Measures to prevent recurrence of incidents (if any)",
}

ZABBIX_REQUIRED_COLUMNS = {
    "Severity",
    "Time",
    "Recovery time",
    "Status",
    "Host",
    "Problem",
    "Duration",
    "Ack",
    "Actions",
    "Tags",
}

ZABBIX_NORMALIZED_REQUIRED_COLUMNS = {
    "event_id",
    "started_at_local",
    "recovered_at_local",
    "status",
    "severity",
    "host",
    "site_code",
    "problem_raw",
    "problem_signature",
    "duration_text",
    "acknowledged",
    "action_count",
    "domain",
    "component",
    "scope",
    "tags_json",
}

MARKDOWN_FILENAMES = [
    "00_report_context.md",
    "01_executive_summary.md",
    "02_operational_timeline.md",
    "02_confirmed_incidents.md",
    "03_recurrence_patterns.md",
    "04_alert_patterns.md",
    "05_ticket_impact.md",
    "05_ticket_impact_index.md",
    "06_data_quality.md",
]


@dataclass
class AdapterResult:
    records: list[dict[str, Any]]
    input_rows: int
    skipped_blank_rows: int = 0
    duplicate_rows: int = 0
    rejected_rows: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_multiline_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def normalize_header(value: Any) -> str:
    return clean_text(value)


def stable_hash(*parts: Any, length: int = 12) -> str:
    payload = "\x1f".join(clean_text(part).lower() for part in parts)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:length].upper()


def source_ref(filename: str, sheet_name: str, row_number: int) -> str:
    return f"{filename}#{sheet_name}:row-{row_number}"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = clean_text(value)
    if not text:
        raise ValueError("empty datetime value")
    formats = [
        "%b %d, %Y @ %H:%M:%S.%f",
        "%b %d, %Y @ %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %I:%M:%S %p",
        "%Y-%m-%d",
    ]
    try:
        return datetime.fromisoformat(text).replace(tzinfo=None)
    except ValueError:
        pass
    for date_format in formats:
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            continue
    raise ValueError(f"invalid datetime value: {text}")


def parse_time(value: Any) -> time:
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    text = clean_text(value)
    if not text:
        return time.min
    for time_format in ["%H:%M:%S", "%H:%M"]:
        try:
            return datetime.strptime(text, time_format).time()
        except ValueError:
            continue
    raise ValueError(f"invalid time value: {text}")


def combine_date_time(date_value: Any, time_value: Any) -> datetime:
    parsed_date = parse_datetime(date_value).date()
    parsed_time = parse_time(time_value)
    return datetime.combine(parsed_date, parsed_time)


def iso_datetime(value: datetime | None) -> str:
    return "" if value is None else value.isoformat(timespec="seconds")


def _is_blank_row(values: Iterable[Any]) -> bool:
    return not any(clean_text(value) for value in values)


def _load_sheet(path: Path, sheet_name: str | None = None):
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    raw_rows = list(worksheet.iter_rows(values_only=True))
    if not raw_rows:
        workbook.close()
        raise ValueError(f"No rows found in {path.name}#{worksheet.title}")
    headers = [normalize_header(value) for value in raw_rows[0]]
    loaded_sheet_name = worksheet.title
    workbook.close()
    return loaded_sheet_name, headers, raw_rows[1:]


def _load_csv_rows(path: Path):
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        raw_rows = list(reader)
    if not raw_rows:
        raise ValueError(f"No rows found in {path.name}")
    headers = [normalize_header(value) for value in raw_rows[0]]
    return "csv", headers, raw_rows[1:]


def _validate_columns(path: Path, sheet_name: str, headers: list[str], required: set[str]):
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(
            f"Missing required columns in {path.name}#{sheet_name}: {', '.join(missing)}"
        )


def _row_dict(headers: list[str], values: Iterable[Any]) -> dict[str, Any]:
    values = list(values)
    padded = values + [""] * max(0, len(headers) - len(values))
    return {header: padded[index] if index < len(padded) else "" for index, header in enumerate(headers)}


def _load_issue_sources(path: Path) -> list[tuple[Path, str, list[str], list[Iterable[Any]]]]:
    if path.is_dir():
        sources = []
        for csv_path in sorted(path.glob("*.csv")):
            sheet_name, headers, rows = _load_csv_rows(csv_path)
            sources.append((csv_path, sheet_name, headers, rows))
        if not sources:
            raise FileNotFoundError(f"No CSV files found in input directory: {path}")
        return sources
    if path.suffix.lower() == ".csv":
        sheet_name, headers, rows = _load_csv_rows(path)
        return [(path, sheet_name, headers, rows)]
    sheet_name, headers, rows = _load_sheet(path)
    return [(path, sheet_name, headers, rows)]


def _issue_schema(headers: list[str]) -> tuple[str, set[str]]:
    header_set = set(headers)
    if ISSUE_MONTHLY_REQUIRED_COLUMNS <= header_set:
        return "monthly_csv", ISSUE_MONTHLY_REQUIRED_COLUMNS
    if ISSUE_LEGACY_REQUIRED_COLUMNS <= header_set:
        return "legacy_xlsx", ISSUE_LEGACY_REQUIRED_COLUMNS
    return "monthly_csv", ISSUE_MONTHLY_REQUIRED_COLUMNS


def _row_value(row: dict[str, Any], name: str) -> Any:
    return row.get(name, "")


def _first_row_value(row: dict[str, Any], *names: str) -> Any:
    for name in names:
        value = row.get(name, "")
        if clean_text(value):
            return value
    return ""


def _email_domain(value: Any) -> str:
    text = clean_text(value).lower()
    if "@" not in text:
        return "UNKNOWN"
    return text.rsplit("@", 1)[1] or "UNKNOWN"


def derive_site_code(site_name: Any) -> str:
    text = clean_text(site_name)
    if not text:
        return "UNKNOWN"
    return text.split(" - ", 1)[0].strip()


def derive_site_code_from_host(host: Any) -> str:
    match = re.match(r"^VNM([A-Z0-9]{3})", clean_text(host).upper())
    return match.group(1) if match else "UNKNOWN"


def load_incidents(path: Path) -> AdapterResult:
    sheet_name, headers, rows = _load_sheet(path, "Form Responses 1")
    _validate_columns(path, sheet_name, headers, INCIDENT_REQUIRED_COLUMNS)
    records: list[dict[str, Any]] = []
    rejected_rows: list[dict[str, Any]] = []
    skipped_blank_rows = 0
    input_rows = 0

    for row_number, values in enumerate(rows, start=2):
        if _is_blank_row(values):
            skipped_blank_rows += 1
            continue
        input_rows += 1
        row = _row_dict(headers, values)
        reference = source_ref(path.name, sheet_name, row_number)
        try:
            submitted_at = (
                parse_datetime(row["Timestamp"]) if clean_text(row["Timestamp"]) else None
            )
            started_at = combine_date_time(row["Incident DATE"], row["Incident TIME"])
            resolved_at = None
            if clean_text(row["Incident resolution date"]):
                resolved_at = combine_date_time(
                    row["Incident resolution date"], row["Incident resolution time"]
                )
        except ValueError as error:
            rejected_rows.append({"source_ref": reference, "reason": str(error)})
            continue

        site_code = derive_site_code(row["SITE / Location"])
        isp = clean_text(row["ISP"]) or "UNKNOWN"
        incident_type = clean_text(row["Incindent Type"]) or "UNKNOWN"
        duration_minutes = (
            round((resolved_at - started_at).total_seconds() / 60, 2)
            if resolved_at is not None
            else None
        )
        records.append(
            {
                "incident_id": f"INC-{row_number}",
                "source_ref": reference,
                "submitted_at": iso_datetime(submitted_at),
                "started_at": iso_datetime(started_at),
                "resolved_at": iso_datetime(resolved_at),
                "duration_minutes": duration_minutes,
                "site_code": site_code,
                "site_name": clean_text(row["SITE / Location"]) or "UNKNOWN",
                "isp": isp,
                "reporter": clean_text(row["Reporter"]) or "UNKNOWN",
                "incident_type": incident_type,
                "severity": clean_text(row["Severity"]) or "UNKNOWN",
                "description": clean_multiline_text(row["Incident description"]) or "UNKNOWN",
                "initial_cause": clean_multiline_text(row["Initial Cause"]) or "UNKNOWN",
                "troubleshooting_actions": clean_multiline_text(row["Troubleshooting actions"])
                or "UNKNOWN",
                "root_cause": clean_multiline_text(row["Root Cause"]) or "UNKNOWN",
                "preventive_action": clean_multiline_text(
                    row["Measures to prevent recurrence of incidents (if any)"]
                )
                or "UNKNOWN",
                "resolution_status": "RESOLVED" if resolved_at else "UNKNOWN",
                "recurrence_key": "|".join(
                    [site_code.lower(), isp.lower(), incident_type.lower()]
                ),
                "evidence_label": "SOURCE FACT",
            }
        )
    return AdapterResult(
        records=records,
        input_rows=input_rows,
        skipped_blank_rows=skipped_blank_rows,
        rejected_rows=rejected_rows,
    )


def load_issue_tickets(path: Path) -> AdapterResult:
    grouped: dict[str, dict[str, Any]] = {}
    rejected_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    skipped_blank_rows = 0
    input_rows = 0

    for source_path, sheet_name, headers, rows in _load_issue_sources(path):
        schema_name, required_columns = _issue_schema(headers)
        _validate_columns(source_path, sheet_name, headers, required_columns)
        for row_number, values in enumerate(rows, start=2):
            if _is_blank_row(values):
                skipped_blank_rows += 1
                continue
            input_rows += 1
            row = _row_dict(headers, values)
            reference = source_ref(source_path.name, sheet_name, row_number)
            try:
                comment_at = parse_datetime(row["itcenter.ticket.comment.created_at"])
                ticket_created_at = (
                    parse_datetime(row["itcenter.ticket.created_at"])
                    if schema_name == "monthly_csv"
                    else None
                )
                comment_updated_at = (
                    parse_datetime(row["itcenter.ticket.comment.updated_at"])
                    if clean_text(row.get("itcenter.ticket.comment.updated_at"))
                    else comment_at
                )
            except ValueError as error:
                rejected_rows.append({"source_ref": reference, "reason": str(error)})
                continue
            ticket_id = clean_text(row["itcenter.ticket.id"])
            if not ticket_id:
                rejected_rows.append({"source_ref": reference, "reason": "empty ticket id"})
                continue
            comment_text = clean_multiline_text(row["itcenter.ticket.comment.text"])
            ticket = grouped.setdefault(
                ticket_id,
                {
                    "ticket_id": ticket_id,
                    "source_refs": [],
                    "source_files": set(),
                    "partition_months": set(),
                    "partition_comment_counts": defaultdict(int),
                    "ticket_created_times": [],
                    "comment_times": [],
                    "comment_updated_times": [],
                    "comments": [],
                    "comment_author_domains": set(),
                    "category": clean_text(row["itcenter.ticket.category_en_name"])
                    or "UNKNOWN",
                    "classification": clean_text(row["itcenter.ticket.classification_name"])
                    or "UNKNOWN",
                    "subclassification": clean_text(
                        _row_value(row, "itcenter.ticket.subclassification_name")
                    )
                    or "UNKNOWN",
                    "title": clean_text(row["itcenter.ticket.full_title"]) or "UNKNOWN",
                    "business_unit": clean_text(
                        row["itcenter.ticket.service_recipient.user.business_unit"]
                    )
                    or "UNKNOWN",
                    "entity_name": clean_text(
                        _row_value(row, "itcenter.ticket.service_recipient.user.entity_name")
                    )
                    or "UNKNOWN",
                    "department": clean_text(
                        _row_value(row, "itcenter.ticket.service_recipient.user.department_name")
                    )
                    or "UNKNOWN",
                    "location_full_name": clean_text(
                        _row_value(row, "itcenter.ticket.location_full_name")
                    )
                    or "UNKNOWN",
                    "office_display": clean_text(
                        _row_value(row, "itcenter.ticket.service_recipient.user.office_display")
                    )
                    or "UNKNOWN",
                    "user_name": clean_text(_row_value(row, "itcenter.ticket.service_recipient.user.name"))
                    or "UNKNOWN",
                    "user_email": clean_text(_row_value(row, "itcenter.ticket.service_recipient.user.email"))
                    or "UNKNOWN",
                    "assignee": clean_text(_row_value(row, "itcenter.ticket.assignee.user.name"))
                    or "UNKNOWN",
                },
            )
            ticket["source_refs"].append(reference)
            ticket["source_files"].add(source_path.name)
            partition_month = comment_at.strftime("%Y-%m")
            ticket["partition_months"].add(partition_month)
            ticket["partition_comment_counts"][partition_month] += 1
            if ticket_created_at is not None:
                ticket["ticket_created_times"].append(ticket_created_at)
            ticket["comment_times"].append(comment_at)
            ticket["comment_updated_times"].append(comment_updated_at)
            ticket["comments"].append(comment_text)
            comment_author = _first_row_value(
                row,
                "itcenter.ticket.comment.created_by.user.email",
                "itcenter.ticket.service_recipient.user.email",
            )
            if comment_author:
                ticket["comment_author_domains"].add(_email_domain(comment_author))

    records = []
    for ticket_id in sorted(grouped):
        ticket = grouped[ticket_id]
        comments = [comment for comment in ticket.pop("comments") if comment]
        comment_times = ticket.pop("comment_times")
        comment_updated_times = ticket.pop("comment_updated_times")
        ticket_created_times = ticket.pop("ticket_created_times")
        source_files = sorted(ticket.pop("source_files"))
        partition_months = sorted(ticket.pop("partition_months"))
        partition_comment_counts = dict(sorted(ticket.pop("partition_comment_counts").items()))
        comment_author_domains = sorted(ticket.pop("comment_author_domains"))
        if ticket_created_times:
            unique_created_values = {
                iso_datetime(created_at) for created_at in ticket_created_times
            }
            if len(unique_created_values) > 1:
                warnings.append(
                    f"Ticket {ticket_id} has multiple ticket_created_at values; using earliest"
                )
            ticket_created_at = min(ticket_created_times)
            evidence_time_basis = "ticket_created_at"
        else:
            ticket_created_at = min(comment_times)
            evidence_time_basis = "first_comment_at_fallback"
        first_comment_at = min(comment_times)
        last_comment_at = max(comment_times)
        last_activity_at = max(comment_updated_times or comment_times)
        evidence_last_seen_at = max(last_comment_at, last_activity_at)
        ticket.update(
            {
                "ticket_created_at": iso_datetime(ticket_created_at),
                "first_comment_at": iso_datetime(first_comment_at),
                "last_comment_at": iso_datetime(last_comment_at),
                "last_activity_at": iso_datetime(last_activity_at),
                "evidence_started_at": iso_datetime(ticket_created_at),
                "evidence_last_seen_at": iso_datetime(evidence_last_seen_at),
                "evidence_time_basis": evidence_time_basis,
                "comment_count": len(comment_times),
                "comments_summary": " | ".join(comments[:3]) or "UNKNOWN",
                "impact_evidence": ticket["title"],
                "source_files": source_files,
                "partition_months": partition_months,
                "partition_comment_counts": partition_comment_counts,
                "comment_author_domains": comment_author_domains,
                "evidence_label": "SOURCE FACT",
            }
        )
        records.append(ticket)
    return AdapterResult(
        records=records,
        input_rows=input_rows,
        skipped_blank_rows=skipped_blank_rows,
        rejected_rows=rejected_rows,
        warnings=warnings,
    )


def parse_tags(value: Any) -> dict[str, list[str]]:
    tags: defaultdict[str, list[str]] = defaultdict(list)
    for token in clean_text(value).split(", "):
        if ": " not in token:
            continue
        key, tag_value = token.split(": ", 1)
        if tag_value not in tags[key]:
            tags[key].append(tag_value)
    return dict(tags)


def parse_tags_json(value: Any) -> dict[str, list[str]]:
    text = clean_text(value)
    if not text:
        return {}
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        return parse_tags(text)
    if not isinstance(payload, dict):
        return {}
    tags: dict[str, list[str]] = {}
    for key, raw_values in payload.items():
        if isinstance(raw_values, list):
            values = [clean_text(item) for item in raw_values if clean_text(item)]
        else:
            values = [clean_text(raw_values)] if clean_text(raw_values) else []
        if values:
            tags[clean_text(key)] = values
    return tags


def normalize_problem_signature(problem: Any) -> str:
    text = clean_text(problem).lower()
    text = re.sub(
        r"current:\s*\d+(?:\.\d+)?\s*mbps",
        "current:<value> mbps",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(
        r"\d+(?:\.\d+)?(?:-\d+(?:\.\d+)?)?\s*(?:mbps|gbps|%|hz|m|s|℃|°c)\b",
        "<value>",
        text,
        flags=re.IGNORECASE,
    )
    return text


def _load_zabbix_sources(path: Path) -> list[tuple[Path, str, list[str], list[Iterable[Any]]]]:
    if path.is_dir():
        csv_paths = sorted(path.glob("*.csv"), key=lambda item: item.name.lower())
        if not csv_paths:
            raise FileNotFoundError(f"No CSV files found in Zabbix export directory: {path}")
        normalized_paths = [
            csv_path
            for csv_path in csv_paths
            if csv_path.name.lower().startswith("zbx_problems_normalized_")
        ]
        if normalized_paths:
            csv_paths = normalized_paths
        return [
            (csv_path, *_load_csv_rows(csv_path))
            for csv_path in csv_paths
        ]
    sheet_name, headers, rows = _load_sheet(path)
    return [(path, sheet_name, headers, rows)]


def load_zabbix_alerts(path: Path) -> AdapterResult:
    sources = _load_zabbix_sources(path)
    records: list[dict[str, Any]] = []
    rejected_rows: list[dict[str, Any]] = []
    skipped_blank_rows = 0
    input_rows = 0
    duplicate_rows = 0
    seen_rows: set[tuple[str, ...]] = set()

    for source_path, sheet_name, headers, rows in sources:
        header_set = set(headers)
        is_normalized = ZABBIX_NORMALIZED_REQUIRED_COLUMNS.issubset(header_set)
        if not is_normalized:
            _validate_columns(source_path, sheet_name, headers, ZABBIX_REQUIRED_COLUMNS)
        for row_number, values in enumerate(rows, start=2):
            if _is_blank_row(values):
                skipped_blank_rows += 1
                continue
            input_rows += 1
            row = _row_dict(headers, values)
            reference = source_ref(source_path.name, sheet_name, row_number)
            if is_normalized:
                dedupe_key = ("normalized", clean_text(row.get("event_id")))
            else:
                dedupe_key = tuple(clean_text(row.get(header)) for header in sorted(ZABBIX_REQUIRED_COLUMNS))
            if dedupe_key in seen_rows:
                duplicate_rows += 1
                continue
            seen_rows.add(dedupe_key)
            if is_normalized:
                try:
                    seen_at = parse_datetime(row["started_at_local"])
                    recovered_at = (
                        parse_datetime(row["recovered_at_local"])
                        if clean_text(row["recovered_at_local"])
                        else None
                    )
                except ValueError as error:
                    rejected_rows.append({"source_ref": reference, "reason": str(error)})
                    continue
                host = clean_text(row["host"]) or "UNKNOWN"
                problem = clean_text(row["problem_raw"]) or "UNKNOWN"
                tags = parse_tags_json(row["tags_json"])
                action_count = clean_text(row["action_count"])
                records.append(
                    {
                        "alert_id": f"ALT-ZBX-{clean_text(row['event_id'])}",
                        "source_ref": reference,
                        "seen_at": iso_datetime(seen_at),
                        "recovered_at": iso_datetime(recovered_at),
                        "status": clean_text(row["status"]) or "UNKNOWN",
                        "severity": clean_text(row["severity"]) or "UNKNOWN",
                        "host": host,
                        "site_code": clean_text(row["site_code"]) or derive_site_code_from_host(host),
                        "problem": problem,
                        "problem_signature": clean_text(row["problem_signature"])
                        or normalize_problem_signature(problem),
                        "duration": clean_text(row["duration_text"]) or "UNKNOWN",
                        "ack": "Yes" if clean_text(row["acknowledged"]).lower() == "true" else "No",
                        "actions": f"Actions ({action_count})" if action_count and action_count != "0" else "UNKNOWN",
                        "domain": clean_text(row["domain"]) or ",".join(tags.get("class", [])) or "UNKNOWN",
                        "component": clean_text(row["component"])
                        or ",".join(tags.get("component", []))
                        or "UNKNOWN",
                        "scope": clean_text(row["scope"]) or ",".join(tags.get("scope", [])) or "UNKNOWN",
                        "tags": json.dumps(tags, ensure_ascii=False, sort_keys=True),
                        "evidence_label": clean_text(row.get("evidence_label")) or "SOURCE FACT",
                    }
                )
            else:
                try:
                    seen_at = parse_datetime(row["Time"])
                    recovered_at = (
                        parse_datetime(row["Recovery time"])
                        if clean_text(row["Recovery time"])
                        else None
                    )
                except ValueError as error:
                    rejected_rows.append({"source_ref": reference, "reason": str(error)})
                    continue
                host = clean_text(row["Host"]) or "UNKNOWN"
                problem = clean_text(row["Problem"]) or "UNKNOWN"
                tags = parse_tags(row["Tags"])
                records.append(
                    {
                        "alert_id": f"ALT-{stable_hash(source_path.name, row_number)}",
                        "source_ref": reference,
                        "seen_at": iso_datetime(seen_at),
                        "recovered_at": iso_datetime(recovered_at),
                        "status": clean_text(row["Status"]) or "UNKNOWN",
                        "severity": clean_text(row["Severity"]) or "UNKNOWN",
                        "host": host,
                        "site_code": derive_site_code_from_host(host),
                        "problem": problem,
                        "problem_signature": normalize_problem_signature(problem),
                        "duration": clean_text(row["Duration"]) or "UNKNOWN",
                        "ack": clean_text(row["Ack"]) or "UNKNOWN",
                        "actions": clean_text(row["Actions"]) or "UNKNOWN",
                        "domain": ",".join(tags.get("class", [])) or "UNKNOWN",
                        "component": ",".join(tags.get("component", [])) or "UNKNOWN",
                        "scope": ",".join(tags.get("scope", [])) or "UNKNOWN",
                        "tags": json.dumps(tags, ensure_ascii=False, sort_keys=True),
                        "evidence_label": "SOURCE FACT",
                    }
                )
    return AdapterResult(
        records=records,
        input_rows=input_rows,
        skipped_blank_rows=skipped_blank_rows,
        duplicate_rows=duplicate_rows,
        rejected_rows=rejected_rows,
    )


def _join_unique(values: Iterable[str]) -> str:
    return ",".join(sorted({value for value in values if value and value != "UNKNOWN"})) or "UNKNOWN"


def _classify_alert_pattern_family(pattern: dict[str, Any]) -> str:
    signature = clean_text(pattern.get("problem_signature")).lower()
    host = clean_text(pattern.get("host")).lower()
    component = clean_text(pattern.get("component")).lower()
    domain = clean_text(pattern.get("domain")).lower()
    joined = " ".join([signature, host, component, domain])
    if "http monitoring" in signature:
        return "HTTP monitoring"
    if any(term in joined for term in ["icmp", "unavailable", "unreachable", " is down", "down!"]):
        return "Network reachability"
    if any(term in joined for term in ["active checks", "zabbix agent", "nodata"]):
        return "Zabbix agent/active check"
    if any(term in joined for term in ["restarted", "uptime"]):
        return "Host restart/uptime"
    if any(term in joined for term in ["ups", "battery", "frequency", "power"]):
        return "Power/UPS"
    return "Other"


def _alert_investigation_priority(pattern: dict[str, Any], family: str) -> str:
    signature = clean_text(pattern.get("problem_signature")).lower()
    host = clean_text(pattern.get("host")).lower()
    alert_count = int(pattern.get("alert_count") or 0)
    if family == "Network reachability":
        if alert_count > 1 or "fw" in host or "sw" in host or "down" in signature:
            return "HIGH"
        return "MEDIUM"
    if family == "HTTP monitoring":
        return "MEDIUM" if alert_count >= 10 else "LOW"
    if family in {"Zabbix agent/active check", "Host restart/uptime"}:
        return "LOW"
    if family == "Power/UPS":
        return "HIGH"
    return "LOW"


def _alert_why_it_matters(pattern: dict[str, Any], family: str) -> str:
    host = clean_text(pattern.get("host"))
    signature = clean_text(pattern.get("problem_signature"))
    if family == "Network reachability":
        return (
            f"{host} reported `{signature}`. This can indicate short connectivity loss, "
            "device/path flap, or monitoring-path instability and should be checked before HTTP-only noise."
        )
    if family == "HTTP monitoring":
        return (
            f"{host} reported `{signature}`. This usually means application/service probe thresholds "
            "were crossed; investigate threshold sensitivity, maintenance windows, or repeated endpoint timeouts."
        )
    if family == "Zabbix agent/active check":
        return (
            f"{host} reported `{signature}`. This points to monitoring-agent availability rather than "
            "confirmed user impact unless supported by ticket or incident evidence."
        )
    if family == "Host restart/uptime":
        return (
            f"{host} reported `{signature}`. This is host-health context and should not be treated "
            "as service impact without another source."
        )
    if family == "Power/UPS":
        return (
            f"{host} reported `{signature}`. This may indicate power/UPS instability and deserves "
            "facility or device-side verification."
        )
    return f"{host} reported `{signature}`. Treat as monitoring context until corroborated."


def group_alert_patterns(alerts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for alert in alerts:
        grouped[(alert["host"], alert["problem_signature"])].append(alert)

    patterns = []
    for (host, signature), records in grouped.items():
        records = sorted(records, key=lambda record: (record["seen_at"], record["source_ref"]))
        pattern = {
            "alert_pattern_id": f"ALP-{stable_hash(host, signature)}",
            "source_refs": [record["source_ref"] for record in records],
            "first_seen_at": records[0]["seen_at"],
            "last_seen_at": records[-1]["seen_at"],
            "last_recovered_at": max(
                (record["recovered_at"] for record in records if record["recovered_at"]),
                default="",
            ),
            "alert_count": len(records),
            "resolved_count": sum(record["status"] == "RESOLVED" for record in records),
            "open_count": sum(record["status"] != "RESOLVED" for record in records),
            "host": host,
            "site_code": records[0]["site_code"],
            "severity": _join_unique(record["severity"] for record in records),
            "problem_signature": signature,
            "domain": _join_unique(record["domain"] for record in records),
            "component": _join_unique(record["component"] for record in records),
            "scope": _join_unique(record["scope"] for record in records),
            "observations": [
                {
                    "seen_at": record["seen_at"],
                    "recovered_at": record["recovered_at"],
                    "status": record["status"],
                    "severity": record["severity"],
                    "source_ref": record["source_ref"],
                }
                for record in records
            ],
            "assessment": "RAW ALERT PATTERN - NOT A CONFIRMED INCIDENT",
            "evidence_label": "COMPUTED FACT",
        }
        pattern["pattern_family"] = _classify_alert_pattern_family(pattern)
        pattern["investigation_priority"] = _alert_investigation_priority(
            pattern, pattern["pattern_family"]
        )
        pattern["why_it_matters"] = _alert_why_it_matters(pattern, pattern["pattern_family"])
        patterns.append(pattern)
    return sorted(patterns, key=lambda pattern: (-pattern["alert_count"], pattern["alert_pattern_id"]))


def group_incident_recurrence(incidents: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for incident in incidents:
        grouped[incident["recurrence_key"]].append(incident)
    patterns = []
    for key, records in grouped.items():
        records = sorted(records, key=lambda record: (record["started_at"], record["incident_id"]))
        first = records[0]
        patterns.append(
            {
                "recurrence_id": f"REC-{stable_hash(key)}",
                "recurrence_key": key,
                "site_code": first["site_code"],
                "site_name": first["site_name"],
                "isp": first["isp"],
                "incident_type": first["incident_type"],
                "incident_count": len(records),
                "incident_ids": [record["incident_id"] for record in records],
                "source_refs": [record["source_ref"] for record in records],
                "first_seen_at": records[0]["started_at"],
                "last_seen_at": records[-1]["started_at"],
                "root_causes": _join_unique(record["root_cause"] for record in records),
                "evidence_label": "COMPUTED FACT",
            }
        )
    return sorted(patterns, key=lambda pattern: (-pattern["incident_count"], pattern["recurrence_id"]))


RESPONSIBILITY_RULES = [
    (
        "ISP",
        [
            "routing",
            "uplink",
            "transit",
            "fiber",
            "fibre",
            "cable",
            "cáp",
            "provider",
            "vnpt",
            "fpt",
            "cmc",
            "netnam",
        ],
    ),
    (
        "INHOUSE",
        [
            "local segment",
            "internal",
            "switch",
            "firewall",
            "configuration",
            "linecard",
            "junos",
            "wifi controller",
            "access point",
        ],
    ),
    (
        "EXTERNAL",
        [
            "electricity",
            "utility",
            "power supply",
            "third-party",
            "third party",
        ],
    ),
]


def classify_responsibility_domain(root_cause: Any) -> tuple[str, str]:
    text = clean_text(root_cause).lower()
    if not text or text == "unknown":
        return "UNKNOWN", "RCA does not support responsibility classification"
    for domain, keywords in RESPONSIBILITY_RULES:
        for keyword in keywords:
            if keyword in text:
                return domain, f"RCA keyword rule: {keyword}"
    return "UNKNOWN", "RCA does not support responsibility classification"


def _time_windows_overlap(
    left_start: str,
    left_end: str,
    right_start: str,
    right_end: str,
    buffer_minutes: int = 60,
) -> bool:
    left_start_at = parse_datetime(left_start)
    left_end_at = parse_datetime(left_end or left_start)
    right_start_at = parse_datetime(right_start)
    right_end_at = parse_datetime(right_end or right_start)
    buffer = timedelta(minutes=buffer_minutes)
    return left_start_at - buffer <= right_end_at and right_start_at <= left_end_at + buffer


def _alert_pattern_id_for_record(alert: dict[str, Any]) -> str:
    return f"ALP-{stable_hash(alert['host'], alert['problem_signature'])}"


def _alert_event_overlaps_incident(
    incident: dict[str, Any], alert: dict[str, Any], buffer_minutes: int
) -> bool:
    if alert["site_code"] != incident["site_code"]:
        return False
    alert_end = alert.get("recovered_at") or incident.get("resolved_at") or incident["started_at"]
    return _time_windows_overlap(
        incident["started_at"],
        incident.get("resolved_at") or incident["started_at"],
        alert["seen_at"],
        alert_end,
        buffer_minutes,
    )


def _related_alert_pattern_ids(
    incident: dict[str, Any],
    alert_patterns: list[dict[str, Any]],
    buffer_minutes: int,
    alerts: list[dict[str, Any]] | None = None,
) -> list[str]:
    if alerts is None:
        return sorted(
            pattern["alert_pattern_id"]
            for pattern in alert_patterns
            if pattern["site_code"] == incident["site_code"]
            and _time_windows_overlap(
                incident["started_at"],
                incident.get("resolved_at") or incident["started_at"],
                pattern["first_seen_at"],
                pattern["last_seen_at"],
                buffer_minutes,
            )
        )

    known_pattern_ids = {pattern["alert_pattern_id"] for pattern in alert_patterns}
    return sorted(
        {
            pattern_id
            for alert in alerts
            if _alert_event_overlaps_incident(incident, alert, buffer_minutes)
            for pattern_id in [_alert_pattern_id_for_record(alert)]
            if pattern_id in known_pattern_ids
        }
    )


def _related_alert_events(
    incident: dict[str, Any],
    alerts: list[dict[str, Any]] | None,
    buffer_minutes: int,
) -> list[dict[str, Any]]:
    if alerts is None:
        return []
    events = []
    for alert in alerts:
        if not _alert_event_overlaps_incident(incident, alert, buffer_minutes):
            continue
        events.append(
            {
                **alert,
                "alert_pattern_id": _alert_pattern_id_for_record(alert),
            }
        )
    return sorted(events, key=lambda alert: (alert["seen_at"], alert["alert_id"]))


NETWORK_INCIDENT_TERMS = {
    "bandwidth",
    "cable",
    "cap",
    "cáp",
    "connectivity",
    "fiber",
    "fibre",
    "internet",
    "latency",
    "link",
    "loss",
    "packet",
    "provider",
    "routing",
    "transit",
    "uplink",
    "wan",
}

NETWORK_ALERT_TERMS = {
    "bandwidth",
    "bgp",
    "download",
    "icmp",
    "interface",
    "latency",
    "link",
    "loss",
    "packet",
    "ping",
    "port",
    "unavailable",
    "unreachable",
    "uplink",
    "upload",
}

POWER_INCIDENT_TERMS = {"ats", "battery", "electricity", "frequency", "power", "ups"}
POWER_ALERT_TERMS = {"battery", "frequency", "power", "ups"}

LOCAL_INCIDENT_TERMS = {
    "access point",
    "controller",
    "firewall",
    "junos",
    "linecard",
    "local segment",
    "switch",
    "wifi",
}
LOCAL_ALERT_TERMS = {
    "access point",
    "controller",
    "firewall",
    "junos",
    "linecard",
    "switch",
    "temperature",
    "wifi",
}


def _text_has_any(text: str, terms: set[str]) -> bool:
    return any(term in text for term in terms)


def _alert_supports_incident(
    incident: dict[str, Any],
    alert: dict[str, Any],
) -> tuple[bool, str]:
    incident_text = clean_text(
        " ".join(
            [
                incident.get("incident_type", ""),
                incident.get("description", ""),
                incident.get("root_cause", ""),
            ]
        )
    ).lower()
    alert_text = clean_text(
        " ".join(
            [
                alert.get("problem_signature", ""),
                alert.get("problem", ""),
                alert.get("domain", ""),
                alert.get("component", ""),
                alert.get("scope", ""),
            ]
        )
    ).lower()

    if _text_has_any(incident_text, NETWORK_INCIDENT_TERMS):
        if _text_has_any(alert_text, NETWORK_ALERT_TERMS):
            return True, "network/fiber incident with network-relevant alert signature"
        return (
            False,
            "same-site/time alert exists, but its signature does not directly support a network/fiber RCA",
        )

    if _text_has_any(incident_text, POWER_INCIDENT_TERMS):
        if _text_has_any(alert_text, POWER_ALERT_TERMS):
            return True, "power incident with power-relevant alert signature"
        return (
            False,
            "same-site/time alert exists, but its signature does not directly support a power RCA",
        )

    if _text_has_any(incident_text, LOCAL_INCIDENT_TERMS):
        if _text_has_any(alert_text, LOCAL_ALERT_TERMS):
            return True, "local infrastructure incident with matching infrastructure alert signature"
        return (
            False,
            "same-site/time alert exists, but its signature does not directly support the local RCA",
        )

    return (
        True,
        "no specific semantic relevance rule matched; preserving same-site/time alert as supporting context",
    )


def _text_contains_site_code(record: dict[str, Any], site_code: str) -> bool:
    if not site_code or site_code == "UNKNOWN":
        return False
    text = " ".join(
        clean_text(record.get(field))
        for field in [
            "title",
            "comments_summary",
            "impact_evidence",
            "location_full_name",
            "office_display",
        ]
    )
    return bool(
        re.search(
            rf"(?<![A-Za-z0-9]){re.escape(site_code)}(?![A-Za-z0-9])",
            text,
            flags=re.IGNORECASE,
        )
    )


def _ticket_window_start(ticket: dict[str, Any]) -> str:
    return (
        clean_text(ticket.get("evidence_started_at"))
        or clean_text(ticket.get("ticket_created_at"))
        or clean_text(ticket.get("first_comment_at"))
    )


def _ticket_window_end(ticket: dict[str, Any]) -> str:
    return (
        clean_text(ticket.get("evidence_last_seen_at"))
        or clean_text(ticket.get("last_activity_at"))
        or clean_text(ticket.get("last_comment_at"))
        or _ticket_window_start(ticket)
    )


def _estimated_user_impact(description: Any) -> bool:
    text = clean_text(description).lower()
    return any(
        keyword in text
        for keyword in [
            "user",
            "impact",
            "affected",
            "service interruption",
            "service disruption",
            "service is down",
            "outage",
        ]
    )


def build_operational_timeline(
    incidents: list[dict[str, Any]],
    alert_patterns: list[dict[str, Any]],
    tickets: list[dict[str, Any]],
    buffer_minutes: int = 60,
    alerts: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    related_incidents_by_alert: defaultdict[str, list[str]] = defaultdict(list)
    events: list[dict[str, Any]] = []

    for incident in incidents:
        related_alerts = _related_alert_pattern_ids(
            incident, alert_patterns, buffer_minutes, alerts
        )
        for alert_pattern_id in related_alerts:
            related_incidents_by_alert[alert_pattern_id].append(incident["incident_id"])
        related_tickets = [
            ticket["ticket_id"]
            for ticket in tickets
            if _text_contains_site_code(ticket, incident["site_code"])
            and _time_windows_overlap(
                incident["started_at"],
                incident["resolved_at"],
                _ticket_window_start(ticket),
                _ticket_window_end(ticket),
                buffer_minutes,
            )
        ]
        responsibility_domain, responsibility_basis = classify_responsibility_domain(
            incident["root_cause"]
        )
        user_impact_status = (
            "CONFIRMED"
            if related_tickets
            else "POTENTIAL_IMPACT"
            if _estimated_user_impact(incident["description"])
            else "NO_DIRECT_EVIDENCE"
        )
        correlation_basis = (
            f"Same-site buffered time-window correlation ({buffer_minutes} minutes); "
            "correlation is context, not proof of RCA or user impact."
        )
        events.append(
            {
                "event_id": f"EVT-{incident['incident_id']}",
                "event_type": "CONFIRMED_INCIDENT",
                "started_at": incident["started_at"],
                "ended_at": incident["resolved_at"] or "UNKNOWN",
                "site_code": incident["site_code"],
                "summary": f"{incident['incident_type']}: {incident['description']}",
                "responsibility_domain": responsibility_domain,
                "responsibility_basis": responsibility_basis,
                "user_impact_status": user_impact_status,
                "resolution_status": incident["resolution_status"],
                "confirmed_rca": incident["root_cause"],
                "related_incident_ids": [incident["incident_id"]],
                "related_alert_pattern_ids": sorted(related_alerts),
                "related_ticket_ids": sorted(related_tickets),
                "correlation_basis": correlation_basis,
                "source_refs": [incident["source_ref"]],
                "evidence_label": "COMPUTED FACT",
            }
        )

    for pattern in alert_patterns:
        events.append(
            {
                "event_id": f"EVT-{pattern['alert_pattern_id']}",
                "event_type": "ZABBIX_ALERT_PATTERN",
                "started_at": pattern["first_seen_at"],
                "ended_at": pattern["last_seen_at"],
                "site_code": pattern["site_code"],
                "summary": (
                    f"{pattern['host']}: {pattern['problem_signature']} "
                    f"({pattern['alert_count']} raw alerts)"
                ),
                "responsibility_domain": "UNKNOWN",
                "responsibility_basis": "Raw alert pattern cannot establish responsibility",
                "user_impact_status": "UNKNOWN",
                "resolution_status": "UNKNOWN",
                "confirmed_rca": "UNKNOWN",
                "related_incident_ids": sorted(
                    related_incidents_by_alert[pattern["alert_pattern_id"]]
                ),
                "related_alert_pattern_ids": [pattern["alert_pattern_id"]],
                "related_ticket_ids": [],
                "correlation_basis": (
                    f"Same-site buffered time-window correlation ({buffer_minutes} minutes); "
                    "raw alert pattern remains monitoring evidence, not a confirmed incident."
                ),
                "source_refs": pattern["source_refs"],
                "evidence_label": "COMPUTED FACT",
            }
        )
    return sorted(events, key=lambda event: (event["started_at"], event["event_id"]))


RELEVANCE_STOPWORDS = {
    "and",
    "are",
    "at",
    "for",
    "from",
    "has",
    "high",
    "into",
    "not",
    "of",
    "on",
    "out",
    "over",
    "same",
    "the",
    "this",
    "to",
    "with",
}


def _relevance_tokens(value: Any) -> set[str]:
    return {
        token
        for token in re.findall(r"[a-z0-9]+", clean_text(value).lower())
        if len(token) >= 3 and token not in RELEVANCE_STOPWORDS
    }


def _ticket_matches_relevance(ticket: dict[str, Any], relevance_text: str) -> bool:
    title_text = clean_text(ticket.get("title"))
    ticket_text = " ".join(
        clean_text(ticket.get(field))
        for field in ["title", "comments_summary", "impact_evidence"]
    )
    title_tokens = _relevance_tokens(title_text)
    ticket_tokens = _relevance_tokens(ticket_text)
    evidence_tokens = _relevance_tokens(relevance_text)
    if title_tokens & evidence_tokens:
        return True
    if len(ticket_tokens & evidence_tokens) >= 2:
        return True
    joined = f" {title_text.lower()} "
    categories = {
        "network": {
            "access",
            "bandwidth",
            "connect",
            "connection",
            "connectivity",
            "latency",
            "network",
            "offline",
            "routing",
            "slow",
            "unreachable",
            "uplink",
        },
        "power": {"ups", "power", "battery", "electricity", "frequency"},
        "hardware": {"temperature", "sensor", "hardware", "switch", "firewall"},
    }
    evidence_categories = {
        category
        for category, terms in categories.items()
        if evidence_tokens & terms
    }
    ticket_categories = {
        category
        for category, terms in categories.items()
        if any(term in joined for term in terms)
    }
    return bool(evidence_categories & ticket_categories)


def _related_tickets(
    tickets: list[dict[str, Any]],
    site_code: str,
    started_at: str,
    ended_at: str,
    relevance_text: str,
    buffer_minutes: int,
) -> list[dict[str, Any]]:
    return [
        ticket
        for ticket in tickets
        if _text_contains_site_code(ticket, site_code)
        and _ticket_matches_relevance(ticket, relevance_text)
        and _time_windows_overlap(
            started_at,
            ended_at,
            _ticket_window_start(ticket),
            _ticket_window_end(ticket),
            buffer_minutes,
        )
    ]


def _evidence_coverage(
    has_zabbix: bool,
    has_incident: bool,
    has_ticket: bool,
    has_zabbix_context: bool = False,
) -> str:
    if has_zabbix_context and not has_zabbix and has_incident:
        labels = ["INCIDENT FORM", "ZABBIX CONTEXT"]
        if has_ticket:
            labels.append("TICKET")
        return " + ".join(labels)
    if has_zabbix_context and has_zabbix and not has_incident:
        labels = ["ZABBIX", "INCIDENT CONTEXT"]
        if has_ticket:
            labels.append("TICKET")
        return " + ".join(labels)

    labels = []
    if has_zabbix:
        labels.append("ZABBIX")
    if has_incident:
        labels.append("INCIDENT FORM")
    if has_ticket:
        labels.append("TICKET")
    return " + ".join(labels) or "NONE"


def _ordered_milestones(items: list[dict[str, str]]) -> list[dict[str, str]]:
    return sorted(items, key=lambda item: (item["at"] == "UNKNOWN", item["at"], item["type"]))


def build_operational_stories(
    incidents: list[dict[str, Any]],
    alert_patterns: list[dict[str, Any]],
    tickets: list[dict[str, Any]],
    recurrence: list[dict[str, Any]],
    buffer_minutes: int = 60,
    alerts: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    related_incidents_by_alert: defaultdict[str, list[str]] = defaultdict(list)
    contextual_incidents_by_alert: defaultdict[str, list[str]] = defaultdict(list)
    alert_patterns_by_id = {
        pattern["alert_pattern_id"]: pattern for pattern in alert_patterns
    }
    recurrence_by_incident = {
        incident_id: pattern
        for pattern in recurrence
        for incident_id in pattern["incident_ids"]
    }
    stories: list[dict[str, Any]] = []

    for incident in incidents:
        related_alert_events = _related_alert_events(incident, alerts, buffer_minutes)
        supporting_alert_events: list[dict[str, Any]] = []
        contextual_alert_events: list[dict[str, Any]] = []
        relevance_basis_by_pattern: dict[str, str] = {}
        for alert_event in related_alert_events:
            supports_incident, basis = _alert_supports_incident(incident, alert_event)
            relevance_basis_by_pattern.setdefault(alert_event["alert_pattern_id"], basis)
            if supports_incident:
                supporting_alert_events.append(alert_event)
            else:
                contextual_alert_events.append(alert_event)

        if related_alert_events:
            supporting_alert_pattern_ids = sorted(
                {
                    alert_event["alert_pattern_id"]
                    for alert_event in supporting_alert_events
                    if alert_event["alert_pattern_id"] in alert_patterns_by_id
                }
            )
            contextual_alert_pattern_ids = sorted(
                {
                    alert_event["alert_pattern_id"]
                    for alert_event in contextual_alert_events
                    if alert_event["alert_pattern_id"] in alert_patterns_by_id
                }
            )
        else:
            supporting_alert_pattern_ids = _related_alert_pattern_ids(
                incident, alert_patterns, buffer_minutes, alerts
            )
            contextual_alert_pattern_ids = []
            for pattern_id in supporting_alert_pattern_ids:
                pattern = alert_patterns_by_id.get(pattern_id)
                if pattern:
                    relevance_basis_by_pattern[pattern_id] = (
                        "pattern-level same-site/time correlation; no raw alert event was available"
                    )
        related_alerts = [
            alert_patterns_by_id[pattern_id]
            for pattern_id in supporting_alert_pattern_ids
            if pattern_id in alert_patterns_by_id
        ]
        contextual_alerts = [
            alert_patterns_by_id[pattern_id]
            for pattern_id in contextual_alert_pattern_ids
            if pattern_id in alert_patterns_by_id
        ]
        for pattern in related_alerts:
            related_incidents_by_alert[pattern["alert_pattern_id"]].append(
                incident["incident_id"]
            )
        for pattern in contextual_alerts:
            contextual_incidents_by_alert[pattern["alert_pattern_id"]].append(
                incident["incident_id"]
            )
        related_tickets = _related_tickets(
            tickets,
            incident["site_code"],
            incident["started_at"],
            incident["resolved_at"],
            " ".join(
                [
                    incident["incident_type"],
                    incident["description"],
                    incident["root_cause"],
                ]
            ),
            buffer_minutes,
        )
        responsibility_domain, responsibility_basis = classify_responsibility_domain(
            incident["root_cause"]
        )
        user_impact_status = (
            "CONFIRMED"
            if related_tickets
            else "POTENTIAL_IMPACT"
            if _estimated_user_impact(incident["description"])
            else "NO_DIRECT_EVIDENCE"
        )
        milestones = [
            {
                "at": incident["started_at"],
                "type": "INCIDENT_STARTED",
                "message": (
                    f"IT recorded confirmed incident {incident['incident_id']} at "
                    f"{incident['site_code']}: {incident['incident_type']}."
                ),
            },
            {
                "at": incident.get("submitted_at") or incident["started_at"],
                "type": "RCA_RECORDED",
                "message": f"Confirmed RCA recorded: {incident['root_cause']}",
            },
        ]
        if supporting_alert_events:
            first_event_by_pattern: dict[str, dict[str, Any]] = {}
            for alert_event in supporting_alert_events:
                first_event_by_pattern.setdefault(alert_event["alert_pattern_id"], alert_event)
            for alert_event in first_event_by_pattern.values():
                milestones.append(
                    {
                        "at": alert_event["seen_at"],
                        "type": "ZABBIX_DETECTED",
                        "message": (
                            f"Zabbix detected {alert_event['problem_signature']} at "
                            f"{alert_event['host']} ({alert_event['alert_pattern_id']}) "
                            "within the incident correlation window."
                        ),
                    }
                )
        if contextual_alert_events:
            first_context_event_by_pattern: dict[str, dict[str, Any]] = {}
            for alert_event in contextual_alert_events:
                first_context_event_by_pattern.setdefault(
                    alert_event["alert_pattern_id"], alert_event
                )
            for alert_event in first_context_event_by_pattern.values():
                milestones.append(
                    {
                        "at": alert_event["seen_at"],
                        "type": "ZABBIX_CONTEXT",
                        "message": (
                            f"Zabbix context only: {alert_event['problem_signature']} at "
                            f"{alert_event['host']} ({alert_event['alert_pattern_id']}) "
                            "overlapped the incident window but does not directly support the incident RCA/type."
                        ),
                    }
                )
        elif not related_alert_events:
            for pattern in related_alerts:
                milestones.append(
                    {
                        "at": pattern["first_seen_at"],
                        "type": "ZABBIX_DETECTED",
                        "message": (
                            f"Zabbix detected {pattern['problem_signature']} at "
                            f"{pattern['host']} ({pattern['alert_pattern_id']})."
                        ),
                    }
                )
        for ticket in related_tickets:
            milestones.append(
                {
                    "at": _ticket_window_start(ticket),
                    "type": "USER_TICKET",
                    "message": (
                        f"Ticket {ticket['ticket_id']} provided direct site-explicit "
                        f"user evidence: {ticket['title']}"
                    ),
                }
            )
        if incident["resolved_at"]:
            milestones.append(
                {
                    "at": incident["resolved_at"],
                    "type": "SERVICE_RECOVERED",
                    "message": "Service recovered according to the confirmed incident record.",
                }
            )
        recurrence_pattern = recurrence_by_incident.get(incident["incident_id"])
        recurrence_summary = (
            f"{recurrence_pattern['incident_count']} confirmed incidents "
            f"({recurrence_pattern['recurrence_id']})"
            if recurrence_pattern
            else "1 confirmed incident"
        )
        if related_alerts:
            zabbix_correlation_status = "RELATED_TO_CONFIRMED_INCIDENT"
        elif contextual_alerts:
            zabbix_correlation_status = "TIME_ALIGNED_CONTEXT_ONLY"
        else:
            zabbix_correlation_status = "NO_MATCHING_ZABBIX_SIGNAL"
        zabbix_relevance_basis = (
            "; ".join(
                f"{pattern_id}: {basis}"
                for pattern_id, basis in sorted(relevance_basis_by_pattern.items())
            )
            if relevance_basis_by_pattern
            else "No same-site Zabbix signal matched this incident window."
        )
        gaps = []
        if contextual_alerts and not related_alerts:
            gaps.append(
                "Only time-aligned Zabbix context was found; no alert signature directly supported the incident RCA/type."
            )
        elif not related_alerts:
            gaps.append("No same-site Zabbix signal matched this incident window.")
        if not related_tickets:
            gaps.append("No direct site-explicit ticket matched this incident window.")
        if incident["root_cause"] == "UNKNOWN":
            gaps.append("Confirmed RCA is missing.")
        stories.append(
            {
                "episode_id": f"OPS-{incident['incident_id']}",
                "episode_type": "CONFIRMED_INCIDENT",
                "started_at": incident["started_at"],
                "ended_at": incident["resolved_at"] or "UNKNOWN",
                "site_code": incident["site_code"],
                "headline": f"{incident['site_code']} | {incident['incident_type']}",
                "milestones": _ordered_milestones(milestones),
                "signal_assessment": zabbix_correlation_status,
                "zabbix_correlation_status": zabbix_correlation_status,
                "zabbix_relevance_basis": zabbix_relevance_basis,
                "user_impact_status": user_impact_status,
                "responsibility_domain": responsibility_domain,
                "responsibility_basis": responsibility_basis,
                "confirmed_rca": incident["root_cause"],
                "resolution_status": incident["resolution_status"],
                "recurrence_summary": recurrence_summary,
                "evidence_coverage": _evidence_coverage(
                    bool(related_alerts),
                    True,
                    bool(related_tickets),
                    has_zabbix_context=bool(contextual_alerts),
                ),
                "investigation_gaps": gaps or ["None."],
                "related_incident_ids": [incident["incident_id"]],
                "supporting_alert_pattern_ids": sorted(
                    pattern["alert_pattern_id"] for pattern in related_alerts
                ),
                "contextual_alert_pattern_ids": sorted(
                    pattern["alert_pattern_id"] for pattern in contextual_alerts
                ),
                "related_alert_pattern_ids": sorted(
                    pattern["alert_pattern_id"] for pattern in related_alerts
                ),
                "related_ticket_ids": sorted(ticket["ticket_id"] for ticket in related_tickets),
                "source_refs": [incident["source_ref"]],
                "evidence_label": "COMPUTED FACT",
            }
        )

    for pattern in alert_patterns:
        pattern_family = clean_text(pattern.get("pattern_family")) or _classify_alert_pattern_family(pattern)
        investigation_priority = clean_text(
            pattern.get("investigation_priority")
        ) or _alert_investigation_priority(pattern, pattern_family)
        why_it_matters = clean_text(pattern.get("why_it_matters")) or _alert_why_it_matters(
            pattern, pattern_family
        )
        related_incident_ids = sorted(
            related_incidents_by_alert[pattern["alert_pattern_id"]]
        )
        contextual_incident_ids = sorted(
            contextual_incidents_by_alert[pattern["alert_pattern_id"]]
        )
        related_tickets = _related_tickets(
            tickets,
            pattern["site_code"],
            pattern["first_seen_at"],
            pattern["last_seen_at"],
            " ".join(
                [
                    pattern["problem_signature"],
                    pattern.get("domain", ""),
                    pattern.get("component", ""),
                ]
            ),
            buffer_minutes,
        )
        if related_incident_ids:
            assessment = "RELATED_TO_CONFIRMED_INCIDENT"
        elif contextual_incident_ids:
            assessment = "TIME_ALIGNED_CONTEXT_ONLY"
        elif related_tickets:
            assessment = "USER_IMPACT_SIGNAL"
        elif pattern["alert_count"] > 1:
            assessment = "LIKELY_NOISE_OR_THRESHOLD_REVIEW"
        else:
            assessment = "UNCONFIRMED_MONITORING_SIGNAL"
        milestones = [
            {
                "at": pattern["first_seen_at"],
                "type": "ZABBIX_DETECTED",
                "message": (
                    f"Zabbix detected {pattern['problem_signature']} at "
                    f"{pattern['host']}."
                ),
            }
        ]
        if pattern["alert_count"] > 1:
            milestones.append(
                {
                    "at": pattern["last_seen_at"],
                    "type": "ALERTS_GROUPED",
                    "message": (
                        f"Grouped {pattern['alert_count']} raw alerts with the same "
                        f"host and normalized signature into {pattern['alert_pattern_id']}."
                    ),
                }
            )
        for ticket in related_tickets:
            milestones.append(
                {
                    "at": _ticket_window_start(ticket),
                    "type": "USER_TICKET",
                    "message": (
                        f"Ticket {ticket['ticket_id']} provided direct site-explicit "
                        f"user evidence: {ticket['title']}"
                    ),
                }
            )
        if pattern.get("last_recovered_at"):
            milestones.append(
                {
                    "at": pattern["last_recovered_at"],
                    "type": "SIGNAL_RECOVERED",
                    "message": (
                        f"Latest recovered Zabbix observation recorded for "
                        f"{pattern['alert_pattern_id']}."
                    ),
                }
            )
        gaps = []
        if contextual_incident_ids and not related_incident_ids:
            gaps.append(
                "Only time-aligned incident context matched this monitoring signal; the alert signature did not directly support the incident RCA/type."
            )
        elif not related_incident_ids:
            gaps.append("No confirmed incident matched this monitoring-signal window.")
        if not related_tickets:
            gaps.append("No direct site-explicit ticket matched this monitoring-signal window.")
        stories.append(
            {
                "episode_id": f"OPS-{pattern['alert_pattern_id']}",
                "episode_type": "MONITORING_SIGNAL",
                "started_at": pattern["first_seen_at"],
                "ended_at": pattern["last_recovered_at"] or pattern["last_seen_at"],
                "site_code": pattern["site_code"],
                "headline": f"{pattern['site_code']} | {pattern['host']} | {pattern['problem_signature']}",
                "host": pattern["host"],
                "problem_signature": pattern["problem_signature"],
                "pattern_family": pattern_family,
                "investigation_priority": investigation_priority,
                "why_it_matters": why_it_matters,
                "milestones": _ordered_milestones(milestones),
                "signal_assessment": assessment,
                "zabbix_correlation_status": assessment,
                "zabbix_relevance_basis": (
                    "Monitoring signal is time-aligned with a confirmed incident, but semantic rules did not treat the alert signature as supporting evidence."
                    if contextual_incident_ids and not related_incident_ids
                    else "Monitoring signal directly supported by same-site/time confirmed incident correlation."
                    if related_incident_ids
                    else "No same-site confirmed incident correlation found."
                ),
                "user_impact_status": "CONFIRMED" if related_tickets else "UNKNOWN",
                "responsibility_domain": "UNKNOWN",
                "responsibility_basis": "Raw alert pattern cannot establish responsibility",
                "confirmed_rca": "UNKNOWN",
                "resolution_status": (
                    "SIGNAL_RECOVERED" if pattern["open_count"] == 0 else "SIGNAL_OPEN"
                ),
                "recurrence_summary": f"{pattern['alert_count']} raw alerts in grouped pattern",
                "evidence_coverage": _evidence_coverage(
                    True,
                    bool(related_incident_ids),
                    bool(related_tickets),
                    has_zabbix_context=bool(contextual_incident_ids),
                ),
                "investigation_gaps": gaps or ["None."],
                "related_incident_ids": related_incident_ids,
                "contextual_incident_ids": contextual_incident_ids,
                "supporting_alert_pattern_ids": (
                    [pattern["alert_pattern_id"]] if related_incident_ids else []
                ),
                "contextual_alert_pattern_ids": (
                    [pattern["alert_pattern_id"]]
                    if contextual_incident_ids and not related_incident_ids
                    else []
                ),
                "related_alert_pattern_ids": [pattern["alert_pattern_id"]],
                "related_ticket_ids": sorted(ticket["ticket_id"] for ticket in related_tickets),
                "source_refs": pattern["source_refs"],
                "observations": pattern.get("observations", []),
                "evidence_label": "COMPUTED FACT",
            }
        )
    return sorted(stories, key=lambda story: (story["started_at"], story["episode_id"]))


def build_operational_timeline_payload(
    stories: list[dict[str, Any]], generated_at: str | None = None
) -> dict[str, Any]:
    return {
        "schema_version": "1.0",
        "record_type": "OPERATIONAL_TIMELINE_COLLECTION",
        "source_type": "operational_timeline",
        "generated_at": generated_at or "",
        "record_count": len(stories),
        "records": stories,
    }


def _query_window_end(story: dict[str, Any]) -> str:
    ended_at = clean_text(story.get("ended_at"))
    return ended_at if ended_at and ended_at != "UNKNOWN" else story["started_at"]


def _timeline_record_type(story: dict[str, Any]) -> str:
    return clean_text(story.get("episode_type") or story.get("event_type")) or "UNKNOWN"


def _point_in_query_window(at: str, from_iso: str, to_iso: str) -> bool:
    if not clean_text(at) or at == "UNKNOWN":
        return False
    return _time_windows_overlap(from_iso, to_iso, at, at, buffer_minutes=0)


def _query_observations(
    story: dict[str, Any],
    from_iso: str,
    to_iso: str,
) -> list[dict[str, Any]]:
    observations = story.get("observations") or []
    return [
        observation
        for observation in observations
        if _point_in_query_window(clean_text(observation.get("seen_at")), from_iso, to_iso)
    ]


def _story_matches_query_window(
    story: dict[str, Any],
    from_iso: str,
    to_iso: str,
) -> bool:
    record_type = _timeline_record_type(story)
    if record_type == "MONITORING_SIGNAL" and story.get("observations"):
        return bool(_query_observations(story, from_iso, to_iso))
    if record_type == "MONITORING_SIGNAL" and story.get("milestones"):
        return any(
            _point_in_query_window(milestone["at"], from_iso, to_iso)
            for milestone in story["milestones"]
            if clean_text(milestone.get("at")) and milestone["at"] != "UNKNOWN"
        )
    return _time_windows_overlap(
        from_iso,
        to_iso,
        story["started_at"],
        _query_window_end(story),
        buffer_minutes=0,
    )


PRIORITY_RANK = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "UNKNOWN": 3}


def _query_milestones(
    story: dict[str, Any],
    from_iso: str,
    to_iso: str,
) -> list[dict[str, Any]]:
    return [
        milestone
        for milestone in story.get("milestones", [])
        if _point_in_query_window(clean_text(milestone.get("at")), from_iso, to_iso)
    ]


def _enrich_query_record(
    story: dict[str, Any],
    from_iso: str,
    to_iso: str,
) -> dict[str, Any]:
    record = dict(story)
    observations = _query_observations(story, from_iso, to_iso)
    record.pop("observations", None)
    if _timeline_record_type(story) != "MONITORING_SIGNAL":
        return record

    query_milestones = _query_milestones(story, from_iso, to_iso)
    if observations:
        seen_values = [observation["seen_at"] for observation in observations]
        recovered_values = [
            observation["recovered_at"]
            for observation in observations
            if clean_text(observation.get("recovered_at"))
        ]
        record.update(
            {
                "query_alert_count": len(observations),
                "query_count_basis": "raw Zabbix observations with seen_at inside the requested window",
                "query_first_seen_at": min(seen_values),
                "query_last_seen_at": max(seen_values),
                "query_last_recovered_at": max(recovered_values) if recovered_values else "",
                "query_source_refs": [
                    observation["source_ref"] for observation in observations[:20]
                ],
                "query_source_ref_count": len(observations),
                "query_observations": observations[:20],
            }
        )
    else:
        record.update(
            {
                "query_alert_count": None,
                "query_count_basis": "raw observations were not available; matched by timeline milestone",
                "query_first_seen_at": query_milestones[0]["at"] if query_milestones else story["started_at"],
                "query_last_seen_at": query_milestones[-1]["at"] if query_milestones else _query_window_end(story),
                "query_last_recovered_at": "",
                "query_source_refs": story.get("source_refs", [])[:20],
                "query_source_ref_count": len(story.get("source_refs", [])),
                "query_observations": [],
            }
        )
    return record


def _monitoring_summary_base(records: list[dict[str, Any]], group_key: str) -> list[dict[str, Any]]:
    grouped: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        if _timeline_record_type(record) != "MONITORING_SIGNAL":
            continue
        key = clean_text(record.get(group_key)) or "UNKNOWN"
        grouped[key].append(record)

    summaries: list[dict[str, Any]] = []
    for key, items in grouped.items():
        counts = [
            int(item["query_alert_count"])
            for item in items
            if isinstance(item.get("query_alert_count"), int)
        ]
        first_seen = [
            item["query_first_seen_at"]
            for item in items
            if clean_text(item.get("query_first_seen_at"))
        ]
        last_seen = [
            item["query_last_seen_at"]
            for item in items
            if clean_text(item.get("query_last_seen_at"))
        ]
        priorities = sorted(
            {clean_text(item.get("investigation_priority")) or "UNKNOWN" for item in items},
            key=lambda priority: PRIORITY_RANK.get(priority, 99),
        )
        summaries.append(
            {
                group_key: key,
                "pattern_count": len(items),
                "query_alert_count": sum(counts) if counts else None,
                "query_count_basis": (
                    "sum of raw Zabbix observations in the requested window"
                    if counts
                    else "raw observation count unavailable"
                ),
                "first_seen_at": min(first_seen) if first_seen else "UNKNOWN",
                "last_seen_at": max(last_seen) if last_seen else "UNKNOWN",
                "assessments": sorted(
                    {clean_text(item.get("signal_assessment")) or "UNKNOWN" for item in items}
                ),
                "investigation_priorities": priorities,
                "sample_hosts": sorted(
                    {
                        clean_text(item.get("host"))
                        for item in items
                        if clean_text(item.get("host"))
                    }
                )[:10],
                "sample_signatures": sorted(
                    {
                        clean_text(item.get("problem_signature"))
                        for item in items
                        if clean_text(item.get("problem_signature"))
                    }
                )[:10],
                "example_episode_ids": [
                    clean_text(item.get("episode_id"))
                    for item in sorted(
                        items,
                        key=lambda item: (
                            PRIORITY_RANK.get(clean_text(item.get("investigation_priority")), 99),
                            -(item.get("query_alert_count") or 0),
                            clean_text(item.get("episode_id")),
                        ),
                    )[:8]
                ],
                "why_it_matters": clean_text(items[0].get("why_it_matters")),
            }
        )
    return sorted(
        summaries,
        key=lambda item: (
            min(
                (PRIORITY_RANK.get(priority, 99) for priority in item["investigation_priorities"]),
                default=99,
            ),
            -(item["query_alert_count"] or 0),
            item[group_key],
        ),
    )


def query_operational_timeline(
    stories: list[dict[str, Any]],
    site_code: str,
    from_at: str,
    to_at: str,
) -> dict[str, Any]:
    requested_site = clean_text(site_code).upper()
    from_dt = parse_datetime(from_at)
    to_dt = parse_datetime(to_at)
    if to_dt < from_dt:
        raise ValueError("to_at must be greater than or equal to from_at")

    from_iso = from_dt.isoformat(timespec="seconds")
    to_iso = to_dt.isoformat(timespec="seconds")
    records = [
        story
        for story in stories
        if clean_text(story.get("site_code")).upper() == requested_site
        and _story_matches_query_window(story, from_iso, to_iso)
    ]
    records = [_enrich_query_record(story, from_iso, to_iso) for story in records]
    counts = Counter(_timeline_record_type(story) for story in records)
    monitoring_family_summary = _monitoring_summary_base(records, "pattern_family")
    monitoring_pattern_summary = _monitoring_summary_base(records, "problem_signature")
    return {
        "ok": True,
        "record_type": "OPERATIONAL_TIMELINE_QUERY_RESULT",
        "evidence_label": "COMPUTED FACT",
        "query": {
            "site_code": requested_site,
            "from_at": from_iso,
            "to_at": to_iso,
        },
        "match_rule": (
            "Same-site confirmed incidents use event-window overlap. Monitoring-signal "
            "stories use milestone timestamps when available, so long-running grouped "
            "patterns do not imply continuous impact. UNKNOWN ended_at is treated as started_at. "
            "This filter is deterministic and does not infer RCA or user impact."
        ),
        "matched_event_count": len(records),
        "counts_by_episode_type": dict(sorted(counts.items())),
        "monitoring_family_summary": monitoring_family_summary,
        "monitoring_pattern_summary": monitoring_pattern_summary,
        "answer_contract": {
            "audience": "middle-management operations reader",
            "required_sections": [
                "Management summary",
                "What happened",
                "Impact and evidence",
                "Operational conclusion",
                "Follow-up / gaps",
            ],
            "style_rules": [
                "Lead with the operational takeaway, not evidence labels.",
                "Keep the answer concise: 2-4 short paragraphs or 5-8 bullets.",
                "Use plain language before technical labels.",
                "Mention record IDs/source references only in a compact evidence line.",
                "Explain CONFIRMED_INCIDENT and MONITORING_SIGNAL in business-readable language.",
                "Explain NO_DIRECT_EVIDENCE as missing direct ticket/user evidence, not as proof of no impact.",
                "Explain POTENTIAL_IMPACT as possible impact without direct ticket/user confirmation.",
                "For Zabbix-alert questions, answer by pattern family and problem signature before listing episode IDs.",
                "Prioritize investigation using pattern family, investigation_priority, query_alert_count, and whether multiple hosts fired together.",
                "When monitoring_family_summary exists, use it as the primary answer source. Treat individual OPS-ALP records as supporting evidence only.",
                "Do not list more than 3 individual Zabbix episode IDs in the main body unless the user explicitly asks for all records.",
                "Do not answer a complete site/month Zabbix summary from individual records alone. Use monitoring_family_summary and monitoring_pattern_summary.",
                "Never say 'at least 1 alert' when summary counts are available.",
                "For incomplete data, prioritize ISP Incident Report for confirmed incident/RCA/status, ITcenter ticket for direct user impact, and Zabbix for monitoring context.",
                "If sources conflict, state the conflict instead of silently merging them.",
                "Do not infer RCA, user impact, or ownership beyond provided fields.",
                "Do not end with a generic follow-up offer.",
            ],
            "preferred_answer_shape": (
                "Trong khoảng [from_at] đến [to_at], site [site_code] ghi nhận [count] "
                "sự kiện vận hành. [Nêu sự kiện chính, thời điểm, đã khôi phục chưa, "
                "RCA/domain nếu có evidence]. [Nêu Zabbix/ticket evidence tìm thấy hoặc "
                "không tìm thấy, giải thích impact]. Kết luận vận hành: [điều đã xác nhận] "
                "và [gap/follow-up cụ thể]."
            ),
            "zabbix_alert_answer_shape": (
                "Trong khoảng [from_at] đến [to_at], site [site_code] ghi nhận [pattern_count] "
                "Zabbix monitoring patterns. Nhóm đáng chú ý nhất là [pattern_family/signature] "
                "vì [why_it_matters], xuất hiện trên [hosts] trong khoảng [first] đến [last]. "
                "Nêu nhóm HTTP/noise nếu nhiều nhưng ít ý nghĩa điều tra; ưu tiên network/firewall/ICMP "
                "nếu có. Nếu không có confirmed incident/ticket khớp trực tiếp, user impact vẫn là UNKNOWN."
            ),
        },
        "records": records,
    }


def load_confirmed_master_data(path: Path | None) -> dict[str, Any]:
    result = {
        "sites": {},
        "hosts": {},
        "network_links": {},
        "services": {},
        "user_groups": {},
        "service_dependencies": {},
        "unconfirmed_rows": 0,
    }
    if path is None or not path.exists():
        return result
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_keys = {
        "sites": "site_code",
        "hosts": "host_name",
        "network_links": "link_code",
        "services": "service_code",
        "user_groups": "group_code",
        "service_dependencies": "dependency_id",
    }
    for sheet_name, key_field in sheet_keys.items():
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [normalize_header(value) for value in rows[0]]
        for values in rows[1:]:
            if _is_blank_row(values):
                continue
            row = _row_dict(headers, values)
            if clean_text(row.get("review_status")).upper() != "CONFIRMED":
                result["unconfirmed_rows"] += 1
                continue
            key = clean_text(row.get(key_field))
            if key:
                result[sheet_name][key] = {
                    header: clean_text(value) for header, value in row.items()
                }
    workbook.close()
    return result


def build_validation_report(
    source_profile: dict[str, int],
    warnings: list[str],
    known_checks: list[dict[str, Any]],
) -> dict[str, Any]:
    invariant_specs = [
        (
            "incident row reconciliation",
            source_profile.get("incident_form_rows", 0),
            source_profile.get("normalized_incidents", 0)
            + source_profile.get("rejected_incident_rows", 0),
        ),
        (
            "zabbix row reconciliation",
            source_profile.get("zabbix_rows", 0),
            source_profile.get("normalized_alert_rows", 0)
            + source_profile.get("rejected_alert_rows", 0)
            + source_profile.get("duplicate_alert_rows", 0),
        ),
        (
            "issue comment reconciliation",
            source_profile.get("issue_comment_rows", 0),
            source_profile.get("aggregated_ticket_comment_count", 0)
            + source_profile.get("rejected_comment_rows", 0),
        ),
    ]
    invariants = [
        {"name": name, "input_rows": input_rows, "accounted_rows": accounted_rows, "passed": input_rows == accounted_rows}
        for name, input_rows, accounted_rows in invariant_specs
    ]
    required_checks_passed = all(
        check.get("passed", False) for check in known_checks if check.get("required", True)
    )
    status = (
        "PASS"
        if all(invariant["passed"] for invariant in invariants) and required_checks_passed
        else "FAIL"
    )
    return {
        "status": status,
        "upload_allowed": status == "PASS",
        "source_profile": source_profile,
        "invariants": invariants,
        "known_checks": known_checks,
        "warnings": sorted(set(warnings)),
    }


def _known_sample_checks(
    source_profile: dict[str, int], recurrence: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    if (
        source_profile.get("normalized_incidents") != 52
        or source_profile.get("normalized_alert_rows") != 1000
        or source_profile.get("normalized_tickets") != 923
    ):
        return []
    by_key = {
        (pattern["site_code"], pattern["isp"], pattern["incident_type"]): pattern[
            "incident_count"
        ]
        for pattern in recurrence
    }
    expected = [
        (("MS2", "VNPT", "High latency"), 16),
        (("XAS", "CMC", "Fiber optic cable failure"), 4),
    ]
    return [
        {
            "name": "sample recurrence " + " | ".join(key),
            "expected": count,
            "actual": by_key.get(key, 0),
            "passed": by_key.get(key, 0) == count,
            "required": True,
        }
        for key, count in expected
    ]


def _write_json(path: Path, payload: Any):
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _escape_markdown(value: Any) -> str:
    return clean_multiline_text(value).replace("\n", "  \n")


def _render_validation_markdown(report: dict[str, Any]) -> str:
    lines = [
        "# Validation Report",
        "",
        f"**Status:** `{report['status']}`",
        "",
        f"**Upload allowed:** `{'YES' if report['upload_allowed'] else 'NO'}`",
        "",
        "## Reconciliation Invariants",
        "",
    ]
    for invariant in report["invariants"]:
        state = "PASS" if invariant["passed"] else "FAIL"
        lines.append(
            f"- `{state}` {invariant['name']}: input={invariant['input_rows']}, accounted={invariant['accounted_rows']}"
        )
    lines.extend(["", "## Known Checks", ""])
    if report["known_checks"]:
        for check in report["known_checks"]:
            state = "PASS" if check["passed"] else "FAIL"
            lines.append(
                f"- `{state}` {check['name']}: expected={check['expected']}, actual={check['actual']}"
            )
    else:
        lines.append("- No supplied-sample checks were applied.")
    lines.extend(["", "## Warnings", ""])
    if report["warnings"]:
        lines.extend(f"- {warning}" for warning in report["warnings"])
    else:
        lines.append("- None.")
    lines.append("")
    return "\n".join(lines)


def _render_context(profile: dict[str, int], warnings: list[str]) -> str:
    warning_lines = "\n".join(f"- {warning}" for warning in warnings) or "- None."
    return f"""# IT Operations Report Context

## Rules for Alpha Intelligence

- Raw alerts are evidence, not confirmed incidents.
- Use statistics exactly as provided in this package. Do not count retrieved fragments.
- Distinguish `SOURCE FACT`, `COMPUTED FACT`, `POTENTIAL_IMPACT`, `AI INFERENCE`, `NO_DIRECT_EVIDENCE`, and `UNKNOWN`.
- Cite record IDs and source references.
- Never invent RCA, affected users, resolution status, or preventive actions.
- Start date-range investigations with `02_operational_timeline.md`.
- Treat timeline correlation as context, not proof of RCA or user impact.
- Same-site/time Zabbix correlation is not enough to prove support for an incident RCA.
  Treat semantically unrelated alerts as `TIME_ALIGNED_CONTEXT_ONLY`.
- Evidence priority for incomplete or conflicting data:
  1. ISP Incident Report for confirmed incident existence, start/end time, RCA,
     resolution status, ISP/provider, and responsibility domain.
  2. ITcenter ticket for direct user evidence, affected scope, business/user
     symptoms, escalation, and impact confirmation.
  3. Zabbix Alert for monitoring signals, telemetry, timing context, recurrence,
     and technical symptoms.
- Do not ignore lower-priority evidence. Use it to add context, confirm timing, or expose gaps.
- If sources conflict, state the conflict and cite record IDs/source references.
- For user impact, direct ITcenter ticket/user evidence outranks incident description.
  Zabbix alone cannot confirm user impact.
- For Zabbix-alert questions, answer by pattern family and problem signature before listing episode IDs.
- For Zabbix-alert questions, prioritize investigation using `pattern_family`,
  `investigation_priority`, query alert count, query first/last seen time, and whether
  multiple hosts fired together.
- For site/date-range questions, do not answer from arbitrary retrieved fragments.
  Prefer the deterministic query workflow/result. If no query result is available,
  retrieve `02_operational_timeline.md` and `04_alert_patterns.md` summary sections
  before answering.
- If only individual `OPS-ALP-*` records are retrieved for a Zabbix-alert question,
  state that the context is incomplete and ask to run or refresh the query. Do not
  claim a complete monthly/site summary from partial records.
- Never say "at least 1 alert" for a complete site/month alert summary unless the
  retrieved context is explicitly partial. Use exact counts only from
  `monitoring_family_summary`, `monitoring_pattern_summary`, or `Site Pattern Family Summary`.
- Distinguish responsibility domains: `INHOUSE`, `ISP`, `EXTERNAL`, and `UNKNOWN`.
- For date-range site questions, write a middle-management brief with:
  management summary, what happened, impact and evidence, operational conclusion,
  and follow-up/gaps.
- Lead with the operational takeaway and use plain language before technical labels.
- Keep date-range site answers concise: 2-4 short paragraphs or 5-8 bullets.
- Do not return only raw fields such as `matched_event_count`, `related_ticket_ids`, or `evidence_coverage`.
- Do not over-expose audit labels such as `SOURCE FACT`, `COMPUTED FACT`, or `NO_MATCHING_ZABBIX_SIGNAL`
  unless they are needed to explain evidence quality.
- Explain `NO_DIRECT_EVIDENCE` as missing direct ticket/user evidence, not proof that no user was affected.
- Explain `POTENTIAL_IMPACT` as possible impact without direct ticket/user confirmation.

## Source Profile

- Confirmed incidents: `{profile['normalized_incidents']}`
- Raw Zabbix alerts: `{profile['normalized_alert_rows']}`
- Operational stories: `{profile['operational_timeline_events']}`
- Unique tickets: `{profile['normalized_tickets']}`
- Issue-report comment rows: `{profile['aggregated_ticket_comment_count']}`
- Ticket monthly partitions: `{profile.get('ticket_partition_count', 0)}`
- Confirmed incident period: `{profile['incident_period_start']}` to `{profile['incident_period_end']}`
- Raw alert period: `{profile['alert_period_start']}` to `{profile['alert_period_end']}`
- Ticket evidence period: `{profile['ticket_period_start']}` to `{profile['ticket_period_end']}`
- Ticket activity period: `{profile.get('ticket_activity_period_start', 'UNKNOWN')}` to `{profile.get('ticket_activity_period_end', 'UNKNOWN')}`

## Data Quality Warnings

{warning_lines}
"""


def _render_executive_summary(
    profile: dict[str, int], recurrence: list[dict[str, Any]], alerts: list[dict[str, Any]]
) -> str:
    lines = [
        "# Executive Summary",
        "",
        "## Deterministic KPIs",
        "",
        f"- Confirmed incidents: `{profile['normalized_incidents']}`",
        f"- Raw Zabbix alerts before grouping: `{profile['normalized_alert_rows']}`",
        f"- Alert patterns after grouping: `{len(alerts)}`",
        f"- Operational stories: `{profile['operational_timeline_events']}`",
        f"- Unique tickets: `{profile['normalized_tickets']}`",
        f"- Issue-report comment rows: `{profile['aggregated_ticket_comment_count']}`",
        f"- Ticket monthly partitions: `{profile.get('ticket_partition_count', 0)}`",
        "",
        "## Operational Story Coverage",
        "",
        f"- Confirmed-incident stories: `{profile['timeline_confirmed_incident_events']}`",
        f"- Monitoring-signal stories: `{profile['timeline_zabbix_pattern_events']}`",
        f"- Signals related to confirmed incidents: `{profile['signal_assessment_related_to_incident']}`",
        f"- Signals time-aligned as context only: `{profile['signal_assessment_context_only']}`",
        f"- Signals with direct user-impact evidence: `{profile['signal_assessment_user_impact']}`",
        f"- Signals needing noise or threshold review: `{profile['signal_assessment_noise_review']}`",
        f"- Unconfirmed monitoring signals: `{profile['signal_assessment_unconfirmed']}`",
        "",
        "## Confirmed Incident Responsibility Classification",
        "",
        f"- `ISP`: `{profile['responsibility_incidents_isp']}`",
        f"- `INHOUSE`: `{profile['responsibility_incidents_inhouse']}`",
        f"- `EXTERNAL`: `{profile['responsibility_incidents_external']}`",
        f"- `UNKNOWN`: `{profile['responsibility_incidents_unknown']}`",
        "",
        "## Confirmed Incident User-Impact Evidence",
        "",
        f"- `CONFIRMED`: `{profile['user_impact_incidents_confirmed']}`",
        f"- `POTENTIAL_IMPACT`: `{profile['user_impact_incidents_potential_impact']}`",
        f"- `NO_DIRECT_EVIDENCE`: `{profile['user_impact_incidents_no_direct_evidence']}`",
        "",
        "## Top Incident Recurrence Patterns",
        "",
    ]
    for pattern in recurrence[:20]:
        lines.append(
            f"- `{pattern['recurrence_id']}` {pattern['site_code']} | {pattern['isp']} | {pattern['incident_type']}: `{pattern['incident_count']}` confirmed incidents"
        )
    lines.extend(["", "## High-Volume Alert Patterns", ""])
    for pattern in alerts[:20]:
        lines.append(
            f"- `{pattern['alert_pattern_id']}` {pattern['host']} | {pattern['problem_signature']}: `{pattern['alert_count']}` raw alerts"
        )
    lines.append("")
    return "\n".join(lines)


def _render_operational_timeline(stories: list[dict[str, Any]]) -> str:
    lines = [
        "# Operational Timeline",
        "",
        "> Primary date-range investigation stories. A monitoring signal remains evidence, not a confirmed incident.",
        "",
        "## How To Use This Timeline",
        "",
        "- Filter by the requested time range first.",
        "- Read `CONFIRMED_INCIDENT` and `MONITORING_SIGNAL` as different episode types.",
        "- Treat responsibility domain as a computed RCA classification.",
        "- Treat correlated evidence as investigation context, not proof of RCA or user impact.",
        "",
    ]
    for story in stories:
        lines.extend(
            [
                f"## {story['episode_id']} | {story['site_code']} | {story['started_at']} | {story['episode_type']}",
                "",
                f"**{_escape_markdown(story['headline'])}**",
                "",
                "```yaml",
                "record_type: OPERATIONAL_TIMELINE_EVENT",
                f"episode_id: {story['episode_id']}",
                f"episode_type: {story['episode_type']}",
                f"site_code: {story['site_code']}",
                f"started_at: {story['started_at']}",
                f"ended_at: {story['ended_at']}",
                f"signal_assessment: {story['signal_assessment']}",
                f"zabbix_correlation_status: {story.get('zabbix_correlation_status', story['signal_assessment'])}",
                f"pattern_family: {story.get('pattern_family', 'UNKNOWN')}",
                f"investigation_priority: {story.get('investigation_priority', 'UNKNOWN')}",
                f"user_impact_status: {story['user_impact_status']}",
                f"responsibility_domain: {story['responsibility_domain']}",
                f"resolution_status: {story['resolution_status']}",
                f"evidence_coverage: {story['evidence_coverage']}",
                f"evidence_label: {story['evidence_label']}",
                "```",
                "",
                "### Timeline",
                "",
            ]
        )
        lines.extend(
            f"- `{milestone['at']}` {milestone['message']}"
            for milestone in story["milestones"]
        )
        lines.extend(
            [
                "",
                "### Conclusion",
                "",
                f"- Episode type: `{story['episode_type']}`",
                f"- Signal assessment: `{story['signal_assessment']}`",
                f"- Zabbix correlation status: `{story.get('zabbix_correlation_status', story['signal_assessment'])}`",
                f"- Zabbix relevance basis: {_escape_markdown(story.get('zabbix_relevance_basis', 'UNKNOWN'))}",
                f"- Pattern family: `{story.get('pattern_family', 'UNKNOWN')}`",
                f"- Investigation priority: `{story.get('investigation_priority', 'UNKNOWN')}`",
                f"- Why it matters: {_escape_markdown(story.get('why_it_matters', 'UNKNOWN'))}",
                f"- User impact: `{story['user_impact_status']}`",
                f"- Responsibility domain: `{story['responsibility_domain']}`",
                f"- Responsibility basis: {_escape_markdown(story['responsibility_basis'])}",
                f"- Confirmed RCA: {_escape_markdown(story['confirmed_rca'])}",
                f"- Resolution status: `{story['resolution_status']}`",
                f"- Recurrence: {_escape_markdown(story['recurrence_summary'])}",
                f"- Evidence coverage: `{story['evidence_coverage']}`",
                "",
                "### Investigation gaps",
                "",
            ]
        )
        lines.extend(f"- {gap}" for gap in story["investigation_gaps"])
        lines.extend(
            [
                "",
                "### Evidence",
                "",
                f"- Evidence label: `{story['evidence_label']}`",
                f"- Related incident IDs: `{', '.join(story['related_incident_ids']) or 'NONE'}`",
                f"- Contextual incident IDs: `{', '.join(story.get('contextual_incident_ids', [])) or 'NONE'}`",
                f"- Supporting alert pattern IDs: `{', '.join(story.get('supporting_alert_pattern_ids', [])) or 'NONE'}`",
                f"- Contextual alert pattern IDs: `{', '.join(story.get('contextual_alert_pattern_ids', [])) or 'NONE'}`",
                f"- Related alert pattern IDs: `{', '.join(story['related_alert_pattern_ids']) or 'NONE'}`",
                f"- Related ticket IDs: `{', '.join(story['related_ticket_ids']) or 'NONE'}`",
                f"- Source references: `{'; '.join(story['source_refs'][:20])}`",
                f"- Source reference count: `{len(story['source_refs'])}`",
                "",
            ]
        )
    return "\n".join(lines)


def _render_operational_query_result(result: dict[str, Any]) -> str:
    query = result["query"]
    lines = [
        "# Operational Timeline Query Result",
        "",
        "```yaml",
        "record_type: OPERATIONAL_TIMELINE_QUERY_RESULT",
        f"site_code: {query['site_code']}",
        f"from_at: {query['from_at']}",
        f"to_at: {query['to_at']}",
        f"matched_event_count: {result['matched_event_count']}",
        f"evidence_label: {result['evidence_label']}",
        "```",
        "",
        f"Match rule: {result['match_rule']}",
        "",
        "## How To Answer",
        "",
        "- Write for a middle-management operations reader, not as a raw JSON/field dump.",
        "- Lead with the operational takeaway: what happened, duration, recovery, and owner/domain if supported.",
        "- Keep it concise: 2-4 short paragraphs or 5-8 bullets.",
        "- Use plain language before technical labels; keep IDs/source references in a compact evidence line.",
        "- Separate confirmed incidents from monitoring signals, but explain both in business-readable language.",
        "- Explain missing Zabbix/ticket evidence as a limitation, not as proof of no impact.",
        "- For Zabbix-alert questions, answer by pattern family and problem signature before listing episode IDs.",
        "- Prioritize investigation using pattern family, investigation priority, query alert count, and whether multiple hosts fired together.",
        "- When Monitoring Pattern Family Summary exists, use it as the primary answer source. Treat individual OPS-ALP records as supporting evidence only.",
        "- Do not list more than 3 individual Zabbix episode IDs in the main body unless the user explicitly asks for all records.",
        "- Do not answer a complete site/month Zabbix summary from individual records alone. Use Monitoring Pattern Family Summary and Monitoring Problem Signature Summary.",
        "- Never say 'at least 1 alert' when summary counts are available.",
        "- For incomplete data, use this evidence priority: ISP Incident Report for confirmed incident/RCA/status, ITcenter ticket for direct user impact, Zabbix for monitoring context.",
        "- If sources conflict, state the conflict instead of silently merging them.",
        "- End with concrete follow-up/gaps, not a generic offer.",
        "",
        "Preferred final-answer shape:",
        "",
        "```text",
        "Trong khoảng [from_at] đến [to_at], site [site_code] ghi nhận [count] sự kiện vận hành.",
        "[Nêu sự kiện chính, thời điểm, đã khôi phục chưa, RCA/domain nếu có evidence.]",
        "[Nêu Zabbix/ticket evidence tìm thấy hoặc không tìm thấy, giải thích impact.]",
        "Kết luận vận hành: [điều đã xác nhận] và [gap/follow-up cụ thể].",
        "```",
        "",
        "Preferred Zabbix-alert answer shape:",
        "",
        "```text",
        "Trong khoảng [from_at] đến [to_at], site [site_code] ghi nhận [pattern_count] Zabbix monitoring patterns.",
        "Nhóm đáng chú ý nhất là [pattern_family/signature] vì [why_it_matters], xuất hiện trên [hosts] trong khoảng [first] đến [last].",
        "Nêu nhóm HTTP/noise nếu nhiều nhưng ít ý nghĩa điều tra; ưu tiên network/firewall/ICMP nếu có.",
        "Nếu không có confirmed incident/ticket khớp trực tiếp, user impact vẫn là UNKNOWN.",
        "```",
        "",
    ]
    if result.get("monitoring_family_summary"):
        lines.extend(["## Monitoring Pattern Family Summary", ""])
        for item in result["monitoring_family_summary"]:
            count = (
                str(item["query_alert_count"])
                if item.get("query_alert_count") is not None
                else "UNKNOWN"
            )
            lines.extend(
                [
                    f"- `{item['pattern_family']}`: `{item['pattern_count']}` patterns, `{count}` query alerts, priority `{', '.join(item['investigation_priorities'])}`.",
                    f"  Signatures: `{', '.join(item['sample_signatures']) or 'UNKNOWN'}`",
                    f"  Hosts: `{', '.join(item['sample_hosts']) or 'UNKNOWN'}`",
                    f"  Window: `{item['first_seen_at']}` to `{item['last_seen_at']}`",
                    f"  Why it matters: {_escape_markdown(item.get('why_it_matters') or 'UNKNOWN')}",
                ]
            )
        lines.append("")

    if result.get("monitoring_pattern_summary"):
        lines.extend(["## Monitoring Problem Signature Summary", ""])
        for item in result["monitoring_pattern_summary"][:15]:
            count = (
                str(item["query_alert_count"])
                if item.get("query_alert_count") is not None
                else "UNKNOWN"
            )
            lines.extend(
                [
                    f"- `{item['problem_signature']}`: `{item['pattern_count']}` patterns, `{count}` query alerts, priority `{', '.join(item['investigation_priorities'])}`.",
                    f"  Hosts: `{', '.join(item['sample_hosts']) or 'UNKNOWN'}`",
                    f"  Episodes: `{', '.join(item['example_episode_ids']) or 'UNKNOWN'}`",
                    f"  Window: `{item['first_seen_at']}` to `{item['last_seen_at']}`",
                ]
            )
        lines.append("")

    if not result["records"]:
        lines.extend(["No matching operational events were found.", ""])
        return "\n".join(lines)

    for story in result["records"]:
        record_id = clean_text(story.get("episode_id") or story.get("event_id"))
        record_type = _timeline_record_type(story)
        headline = clean_text(story.get("headline") or story.get("summary"))
        related_incidents = story.get("related_incident_ids", [])
        contextual_incidents = story.get("contextual_incident_ids", [])
        supporting_alerts = story.get("supporting_alert_pattern_ids", [])
        contextual_alerts = story.get("contextual_alert_pattern_ids", [])
        related_alerts = story.get("related_alert_pattern_ids", [])
        related_tickets = story.get("related_ticket_ids", [])
        lines.extend(
            [
                f"## {record_id} | {story['site_code']} | {story['started_at']} | {record_type}",
                "",
                f"- Headline: {_escape_markdown(headline)}",
                f"- Window: `{story['started_at']}` to `{story['ended_at']}`",
                f"- Signal assessment: `{story.get('signal_assessment', 'UNKNOWN')}`",
                f"- Zabbix correlation status: `{story.get('zabbix_correlation_status', story.get('signal_assessment', 'UNKNOWN'))}`",
                f"- Pattern family: `{story.get('pattern_family', 'UNKNOWN')}`",
                f"- Investigation priority: `{story.get('investigation_priority', 'UNKNOWN')}`",
                f"- Query alert count: `{story.get('query_alert_count', 'UNKNOWN')}`",
                f"- Query alert window: `{story.get('query_first_seen_at', 'UNKNOWN')}` to `{story.get('query_last_seen_at', 'UNKNOWN')}`",
                f"- User impact: `{story.get('user_impact_status', 'UNKNOWN')}`",
                f"- Evidence coverage: `{story.get('evidence_coverage', 'UNKNOWN')}`",
                f"- Why it matters: {_escape_markdown(story.get('why_it_matters', 'UNKNOWN'))}",
                f"- Related incidents: `{', '.join(related_incidents) or 'NONE'}`",
                f"- Contextual incidents: `{', '.join(contextual_incidents) or 'NONE'}`",
                f"- Supporting alert patterns: `{', '.join(supporting_alerts) or 'NONE'}`",
                f"- Contextual alert patterns: `{', '.join(contextual_alerts) or 'NONE'}`",
                f"- Related alert patterns: `{', '.join(related_alerts) or 'NONE'}`",
                f"- Related tickets: `{', '.join(related_tickets) or 'NONE'}`",
                "",
            ]
        )
    return "\n".join(lines)


def _render_incidents(incidents: list[dict[str, Any]]) -> str:
    lines = ["# Confirmed Incidents", ""]
    for record in incidents:
        lines.extend(
            [
                f"## {record['incident_id']} | {record['site_code']} | {record['incident_type']}",
                "",
                f"- Evidence label: `{record['evidence_label']}`",
                f"- Source: `{record['source_ref']}`",
                f"- Started: `{record['started_at']}`",
                f"- Resolved: `{record['resolved_at'] or 'UNKNOWN'}`",
                f"- ISP: `{record['isp']}`",
                f"- Severity: `{record['severity']}`",
                f"- Description: {_escape_markdown(record['description'])}",
                f"- Root cause: {_escape_markdown(record['root_cause'])}",
                f"- Troubleshooting: {_escape_markdown(record['troubleshooting_actions'])}",
                f"- Preventive action: {_escape_markdown(record['preventive_action'])}",
                "",
            ]
        )
    return "\n".join(lines)


def _render_recurrence(patterns: list[dict[str, Any]]) -> str:
    lines = ["# Confirmed Incident Recurrence Patterns", ""]
    for pattern in patterns:
        lines.extend(
            [
                f"## {pattern['recurrence_id']} | {pattern['site_code']} | {pattern['isp']} | {pattern['incident_type']}",
                "",
                f"- Evidence label: `{pattern['evidence_label']}`",
                f"- Confirmed incident count: `{pattern['incident_count']}`",
                f"- First seen: `{pattern['first_seen_at']}`",
                f"- Last seen: `{pattern['last_seen_at']}`",
                f"- Incident IDs: `{', '.join(pattern['incident_ids'])}`",
                f"- Root causes: {_escape_markdown(pattern['root_causes'])}",
                f"- Source references: `{'; '.join(pattern['source_refs'])}`",
                "",
            ]
        )
    return "\n".join(lines)


def _alert_site_family_summaries(patterns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for pattern in patterns:
        site_code = clean_text(pattern.get("site_code")) or "UNKNOWN"
        family = clean_text(pattern.get("pattern_family")) or "UNKNOWN"
        grouped[(site_code, family)].append(pattern)

    summaries: list[dict[str, Any]] = []
    for (site_code, family), records in grouped.items():
        priorities = sorted(
            {clean_text(record.get("investigation_priority")) or "UNKNOWN" for record in records},
            key=lambda priority: PRIORITY_RANK.get(priority, 99),
        )
        top_signatures = sorted(
            {
                clean_text(record.get("problem_signature"))
                for record in records
                if clean_text(record.get("problem_signature"))
            }
        )[:10]
        high_priority_records = [
            record
            for record in records
            if clean_text(record.get("investigation_priority")) == "HIGH"
        ]
        example_ids = [
            record["alert_pattern_id"]
            for record in sorted(
                high_priority_records or records,
                key=lambda record: (
                    PRIORITY_RANK.get(clean_text(record.get("investigation_priority")), 99),
                    -int(record.get("alert_count") or 0),
                    record["alert_pattern_id"],
                ),
            )[:8]
        ]
        summaries.append(
            {
                "site_code": site_code,
                "pattern_family": family,
                "pattern_count": len(records),
                "raw_alert_count": sum(int(record.get("alert_count") or 0) for record in records),
                "investigation_priorities": priorities,
                "sample_hosts": sorted(
                    {
                        clean_text(record.get("host"))
                        for record in records
                        if clean_text(record.get("host"))
                    }
                )[:12],
                "sample_signatures": top_signatures,
                "example_alert_pattern_ids": example_ids,
            }
        )
    return sorted(
        summaries,
        key=lambda item: (
            item["site_code"],
            min(
                (PRIORITY_RANK.get(priority, 99) for priority in item["investigation_priorities"]),
                default=99,
            ),
            -(item["raw_alert_count"] or 0),
            item["pattern_family"],
        ),
    )


def _render_alert_patterns(patterns: list[dict[str, Any]]) -> str:
    lines = ["# Raw Zabbix Alert Patterns", "", "> Alert pattern is not a confirmed incident.", ""]
    lines.extend(
        [
            "## Site Pattern Family Summary",
            "",
            "> Use this section first for Zabbix-alert questions. Individual alert patterns are evidence details, not the primary answer shape.",
            "",
        ]
    )
    for summary in _alert_site_family_summaries(patterns):
        lines.extend(
            [
                f"### {summary['site_code']} | {summary['pattern_family']}",
                "",
                f"- Pattern count: `{summary['pattern_count']}`",
                f"- Raw alert count across full export: `{summary['raw_alert_count']}`",
                f"- Investigation priorities: `{', '.join(summary['investigation_priorities']) or 'UNKNOWN'}`",
                f"- Sample signatures: `{', '.join(summary['sample_signatures']) or 'UNKNOWN'}`",
                f"- Sample hosts: `{', '.join(summary['sample_hosts']) or 'UNKNOWN'}`",
                f"- Example alert pattern IDs: `{', '.join(summary['example_alert_pattern_ids']) or 'UNKNOWN'}`",
                "",
            ]
        )
    lines.extend(["## Individual Alert Patterns", ""])
    for pattern in patterns:
        refs = pattern["source_refs"]
        lines.extend(
            [
                f"## {pattern['alert_pattern_id']} | {pattern['host']}",
                "",
                f"- Evidence label: `{pattern['evidence_label']}`",
                f"- Assessment: `{pattern['assessment']}`",
                f"- Problem signature: `{pattern['problem_signature']}`",
                f"- Pattern family: `{pattern.get('pattern_family', 'UNKNOWN')}`",
                f"- Investigation priority: `{pattern.get('investigation_priority', 'UNKNOWN')}`",
                f"- Why it matters: {_escape_markdown(pattern.get('why_it_matters', 'UNKNOWN'))}",
                f"- Raw alert count: `{pattern['alert_count']}`",
                f"- Resolved alerts: `{pattern['resolved_count']}`",
                f"- Open alerts: `{pattern['open_count']}`",
                f"- Site code: `{pattern['site_code']}`",
                f"- Domain: `{pattern['domain']}`",
                f"- Component: `{pattern['component']}`",
                f"- Source references: `{'; '.join(refs[:20])}`",
                f"- Source reference count: `{len(refs)}`",
                "",
            ]
        )
    return "\n".join(lines)


def _render_ticket_impact(
    tickets: list[dict[str, Any]],
    title: str = "Ticket Impact Evidence",
    note: str = "Tickets provide direct user evidence but are not automatically incidents.",
) -> str:
    lines = [f"# {title}", "", f"> {note}", ""]
    for ticket in tickets:
        lines.extend(
            [
                f"## TKT-{ticket['ticket_id']}",
                "",
                f"- Evidence label: `{ticket['evidence_label']}`",
                f"- Ticket ID: `{ticket['ticket_id']}`",
                f"- Ticket created at: `{ticket.get('ticket_created_at', 'UNKNOWN')}`",
                f"- Evidence started at: `{ticket.get('evidence_started_at', ticket.get('first_comment_at', 'UNKNOWN'))}`",
                f"- Evidence last seen at: `{ticket.get('evidence_last_seen_at', ticket.get('last_comment_at', 'UNKNOWN'))}`",
                f"- Evidence time basis: `{ticket.get('evidence_time_basis', 'UNKNOWN')}`",
                f"- First comment at: `{ticket.get('first_comment_at', 'UNKNOWN')}`",
                f"- Last comment at: `{ticket.get('last_comment_at', 'UNKNOWN')}`",
                f"- Last activity at: `{ticket.get('last_activity_at', ticket.get('last_comment_at', 'UNKNOWN'))}`",
                f"- Title: {_escape_markdown(ticket['title'])}",
                f"- Classification: `{ticket['classification']}`",
                f"- Subclassification: `{ticket.get('subclassification', 'UNKNOWN')}`",
                f"- Business unit: `{ticket['business_unit']}`",
                f"- Entity: `{ticket.get('entity_name', 'UNKNOWN')}`",
                f"- Location: `{ticket.get('location_full_name', 'UNKNOWN')}`",
                f"- Office display: `{ticket.get('office_display', 'UNKNOWN')}`",
                f"- Department: `{ticket.get('department', 'UNKNOWN')}`",
                f"- User: `{ticket.get('user_name', 'UNKNOWN')}`",
                f"- Email: `{ticket.get('user_email', 'UNKNOWN')}`",
                f"- Assignee: `{ticket.get('assignee', 'UNKNOWN')}`",
                f"- Comment count: `{ticket['comment_count']}`",
                f"- Partition months: `{', '.join(ticket.get('partition_months', [])) or 'UNKNOWN'}`",
                f"- Source files: `{', '.join(ticket.get('source_files', [])) or 'UNKNOWN'}`",
                f"- Comment author domains: `{', '.join(ticket.get('comment_author_domains', [])) or 'UNKNOWN'}`",
                f"- Comment summary: {_escape_markdown(ticket['comments_summary'])}",
                f"- Source references: `{'; '.join(ticket['source_refs'])}`",
                "",
            ]
        )
    return "\n".join(lines)


def _ticket_partition_filename(partition_month: str) -> str:
    return f"05_ticket_impact_{partition_month.replace('-', '_')}.md"


def _ticket_partition_months(tickets: list[dict[str, Any]]) -> list[str]:
    return sorted(
        {
            month
            for ticket in tickets
            for month in ticket.get("partition_months", [])
            if clean_text(month)
        }
    )


def _ticket_partition_records(
    tickets: list[dict[str, Any]], partition_month: str
) -> list[dict[str, Any]]:
    return [
        ticket
        for ticket in tickets
        if partition_month in ticket.get("partition_months", [])
    ]


def _render_ticket_partition_index(tickets: list[dict[str, Any]]) -> str:
    lines = [
        "# Ticket Impact Partition Index",
        "",
        "> Use this file to choose the monthly ticket partition before reading ticket details.",
        "",
        "```yaml",
        "record_type: TICKET_IMPACT_PARTITION_INDEX",
        "partition_basis: comment_activity_month",
        "dedupe_key: ticket_id",
        "primary_impact_time: ticket_created_at",
        "activity_time: comment_created_at/comment_updated_at",
        "```",
        "",
        "## Query Rules",
        "",
        "- Select partitions by requested activity window, then dedupe by `ticket_id`.",
        "- For user/business impact timing, prefer `ticket_created_at` / `evidence_started_at`.",
        "- For ongoing evidence, use `evidence_last_seen_at` and comment activity fields.",
        "- A ticket may appear in multiple month partitions when comments span months.",
        "",
        "## Available Partitions",
        "",
    ]
    for month in _ticket_partition_months(tickets):
        partition_tickets = _ticket_partition_records(tickets, month)
        comment_count = sum(
            int(ticket.get("partition_comment_counts", {}).get(month, 0))
            for ticket in partition_tickets
        )
        source_files = sorted(
            {
                source_file
                for ticket in partition_tickets
                for source_file in ticket.get("source_files", [])
            }
        )
        classifications = Counter(
            clean_text(ticket.get("classification")) or "UNKNOWN"
            for ticket in partition_tickets
        )
        locations = Counter(
            clean_text(ticket.get("location_full_name")) or "UNKNOWN"
            for ticket in partition_tickets
        )
        lines.extend(
            [
                f"### {month}",
                "",
                "```yaml",
                f"partition_month: {month}",
                f"knowledge_filename: {_ticket_partition_filename(month)}",
                f"ticket_count: {len(partition_tickets)}",
                f"comment_row_count: {comment_count}",
                f"source_files: {', '.join(source_files) or 'UNKNOWN'}",
                f"ticket_created_at_min: {min((_ticket_window_start(ticket) for ticket in partition_tickets), default='UNKNOWN')}",
                f"evidence_last_seen_at_max: {max((_ticket_window_end(ticket) for ticket in partition_tickets), default='UNKNOWN')}",
                "```",
                "",
                f"- Top classifications: {_counter_summary(classifications, 8)}",
                f"- Top locations: {_counter_summary(locations, 8)}",
                "",
            ]
        )
    return "\n".join(lines)


def _counter_summary(counter: Counter, limit: int) -> str:
    if not counter:
        return "`UNKNOWN`"
    return ", ".join(
        f"`{key}` ({value})" for key, value in counter.most_common(limit)
    )


def _render_data_quality(
    incidents: list[dict[str, Any]],
    master_data: dict[str, Any],
    rejected_rows: list[dict[str, Any]],
    warnings: list[str],
) -> str:
    missing_rca = sum(record["root_cause"] == "UNKNOWN" for record in incidents)
    missing_preventive = sum(record["preventive_action"] == "UNKNOWN" for record in incidents)
    lines = [
        "# Data Quality",
        "",
        "## Coverage",
        "",
        f"- Incidents missing RCA: `{missing_rca}`",
        f"- Incidents missing preventive action: `{missing_preventive}`",
        f"- Confirmed master-data sites: `{len(master_data['sites'])}`",
        f"- Confirmed master-data hosts: `{len(master_data['hosts'])}`",
        f"- Unconfirmed master-data rows ignored as facts: `{master_data['unconfirmed_rows']}`",
        f"- Rejected source rows: `{len(rejected_rows)}`",
        "",
        "## Limitations",
        "",
        "- User and dependency impact is `UNKNOWN` unless supported by direct ticket evidence or confirmed master data.",
        "- Raw alerts are not confirmed incidents.",
        "",
        "## Warnings",
        "",
    ]
    lines.extend(f"- {warning}" for warning in warnings)
    if not warnings:
        lines.append("- None.")
    lines.extend(["", "## Rejected Rows", ""])
    lines.extend(
        f"- `{rejected['source_ref']}`: {rejected['reason']}" for rejected in rejected_rows
    )
    if not rejected_rows:
        lines.append("- None.")
    lines.append("")
    return "\n".join(lines)


def _excel_scalar(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _append_sheet(workbook: Workbook, name: str, records: list[dict[str, Any]]):
    worksheet = workbook.create_sheet(name)
    if not records:
        worksheet.append(["status"])
        worksheet.append(["NO RECORDS"])
        return
    headers: list[str] = []
    for record in records:
        for key in record:
            if key not in headers:
                headers.append(key)
    worksheet.append(headers)
    for record in records:
        worksheet.append([_excel_scalar(record.get(header, "")) for header in headers])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _write_audit_workbook(
    path: Path,
    timeline: list[dict[str, Any]],
    incidents: list[dict[str, Any]],
    alerts: list[dict[str, Any]],
    tickets: list[dict[str, Any]],
    recurrence: list[dict[str, Any]],
    data_quality: list[dict[str, Any]],
    source_profile: dict[str, Any],
):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _append_sheet(workbook, "operational_timeline", timeline)
    _append_sheet(workbook, "confirmed_incidents", incidents)
    _append_sheet(workbook, "alert_patterns", alerts)
    _append_sheet(workbook, "ticket_evidence", tickets)
    _append_sheet(workbook, "recurrence_patterns", recurrence)
    _append_sheet(workbook, "data_quality", data_quality)
    _append_sheet(
        workbook,
        "source_profile",
        [{"metric": key, "value": value} for key, value in sorted(source_profile.items())],
    )
    workbook.save(path)


def build_knowledge_package(
    raw_dir: Path,
    master_data_path: Path | None,
    output_dir: Path,
    generated_at: str | None = None,
) -> dict[str, Any]:
    raw_dir = Path(raw_dir)
    output_dir = Path(output_dir)
    generated_at = generated_at or datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

    def resolve_raw_path(filename: str, *subdirs: str) -> Path:
        candidates = [raw_dir / filename]
        candidates.extend(raw_dir / subdir / filename for subdir in subdirs)
        for candidate in candidates:
            if candidate.exists():
                return candidate
        return candidates[0]

    issue_dir = raw_dir / ISSUE_REPORT_DIRNAME
    issue_path = issue_dir if issue_dir.exists() else resolve_raw_path(ISSUE_REPORT_FILENAME, ISSUE_REPORT_DIRNAME)
    incident_path = resolve_raw_path(INCIDENT_REPORT_FILENAME, INCIDENT_REPORT_DIRNAME)
    zabbix_file_path = raw_dir / ZABBIX_FILENAME
    zabbix_export_dir = raw_dir / ZABBIX_EXPORT_DIRNAME
    zabbix_path = zabbix_export_dir if zabbix_export_dir.exists() else zabbix_file_path
    inputs = [issue_path, incident_path, zabbix_path]
    for input_path in inputs:
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")
        if input_path.is_dir() and not any(input_path.glob("*.csv")):
            raise FileNotFoundError(f"No CSV files found in input directory: {input_path}")
    if master_data_path is not None:
        master_data_path = Path(master_data_path)
        if not master_data_path.exists():
            raise FileNotFoundError(
                f"Explicit master-data input file not found: {master_data_path}"
            )

    incidents = load_incidents(incident_path)
    alerts = load_zabbix_alerts(zabbix_path)
    tickets = load_issue_tickets(issue_path)
    master_data = load_confirmed_master_data(master_data_path)
    recurrence = group_incident_recurrence(incidents.records)
    alert_patterns = group_alert_patterns(alerts.records)
    operational_stories = build_operational_stories(
        incidents.records, alert_patterns, tickets.records, recurrence, alerts=alerts.records
    )
    incident_stories = [
        story
        for story in operational_stories
        if story["episode_type"] == "CONFIRMED_INCIDENT"
    ]
    monitoring_stories = [
        story
        for story in operational_stories
        if story["episode_type"] == "MONITORING_SIGNAL"
    ]

    rejected_rows = incidents.rejected_rows + alerts.rejected_rows + tickets.rejected_rows
    unmapped_incident_sites = sorted(
        {
            record["site_code"]
            for record in incidents.records
            if record["site_code"] not in master_data["sites"]
        }
    )
    unmapped_alert_hosts = sorted(
        {
            pattern["host"]
            for pattern in alert_patterns
            if pattern["host"] not in master_data["hosts"]
        }
    )
    warnings = [
        f"Rejected source rows: {len(rejected_rows)}" if rejected_rows else "",
        f"Duplicate Zabbix rows ignored after multi-export merge: {alerts.duplicate_rows}"
        if alerts.duplicate_rows
        else "",
        f"Unconfirmed master-data rows ignored as facts: {master_data['unconfirmed_rows']}"
        if master_data["unconfirmed_rows"]
        else "",
        f"Sites without confirmed master-data mapping: {len(unmapped_incident_sites)}"
        if unmapped_incident_sites
        else "",
        f"Alert hosts without confirmed master-data mapping: {len(unmapped_alert_hosts)}"
        if unmapped_alert_hosts
        else "",
        f"Incidents missing RCA: {sum(record['root_cause'] == 'UNKNOWN' for record in incidents.records)}",
        f"Incidents missing preventive action: {sum(record['preventive_action'] == 'UNKNOWN' for record in incidents.records)}",
        "No incident-to-Zabbix timeline correlations matched the current exports"
        if not any(story["related_alert_pattern_ids"] for story in incident_stories)
        else "",
        "No direct site-explicit ticket correlations matched confirmed incidents"
        if not any(story["related_ticket_ids"] for story in incident_stories)
        else "",
    ]
    warnings = [warning for warning in warnings if warning]
    source_profile = {
        "incident_form_rows": incidents.input_rows,
        "normalized_incidents": len(incidents.records),
        "rejected_incident_rows": len(incidents.rejected_rows),
        "zabbix_rows": alerts.input_rows,
        "normalized_alert_rows": len(alerts.records),
        "rejected_alert_rows": len(alerts.rejected_rows),
        "duplicate_alert_rows": alerts.duplicate_rows,
        "issue_comment_rows": tickets.input_rows,
        "aggregated_ticket_comment_count": sum(
            record["comment_count"] for record in tickets.records
        ),
        "rejected_comment_rows": len(tickets.rejected_rows),
        "normalized_tickets": len(tickets.records),
        "ticket_partition_count": len(_ticket_partition_months(tickets.records)),
        "alert_patterns": len(alert_patterns),
        "recurrence_patterns": len(recurrence),
        "operational_timeline_events": len(operational_stories),
        "timeline_confirmed_incident_events": sum(
            story["episode_type"] == "CONFIRMED_INCIDENT" for story in operational_stories
        ),
        "timeline_zabbix_pattern_events": sum(
            story["episode_type"] == "MONITORING_SIGNAL" for story in operational_stories
        ),
        "signal_assessment_related_to_incident": sum(
            story["signal_assessment"] == "RELATED_TO_CONFIRMED_INCIDENT"
            for story in monitoring_stories
        ),
        "signal_assessment_context_only": sum(
            story["signal_assessment"] == "TIME_ALIGNED_CONTEXT_ONLY"
            for story in monitoring_stories
        ),
        "signal_assessment_user_impact": sum(
            story["signal_assessment"] == "USER_IMPACT_SIGNAL"
            for story in monitoring_stories
        ),
        "signal_assessment_noise_review": sum(
            story["signal_assessment"] == "LIKELY_NOISE_OR_THRESHOLD_REVIEW"
            for story in monitoring_stories
        ),
        "signal_assessment_unconfirmed": sum(
            story["signal_assessment"] == "UNCONFIRMED_MONITORING_SIGNAL"
            for story in monitoring_stories
        ),
        "responsibility_incidents_inhouse": sum(
            story["responsibility_domain"] == "INHOUSE" for story in incident_stories
        ),
        "responsibility_incidents_isp": sum(
            story["responsibility_domain"] == "ISP" for story in incident_stories
        ),
        "responsibility_incidents_external": sum(
            story["responsibility_domain"] == "EXTERNAL" for story in incident_stories
        ),
        "responsibility_incidents_unknown": sum(
            story["responsibility_domain"] == "UNKNOWN" for story in incident_stories
        ),
        "user_impact_incidents_confirmed": sum(
            story["user_impact_status"] == "CONFIRMED" for story in incident_stories
        ),
        "user_impact_incidents_potential_impact": sum(
            story["user_impact_status"] == "POTENTIAL_IMPACT" for story in incident_stories
        ),
        "user_impact_incidents_no_direct_evidence": sum(
            story["user_impact_status"] == "NO_DIRECT_EVIDENCE" for story in incident_stories
        ),
        "incident_period_start": min(
            (record["started_at"] for record in incidents.records), default="UNKNOWN"
        ),
        "incident_period_end": max(
            (record["started_at"] for record in incidents.records), default="UNKNOWN"
        ),
        "alert_period_start": min(
            (record["seen_at"] for record in alerts.records), default="UNKNOWN"
        ),
        "alert_period_end": max(
            (record["seen_at"] for record in alerts.records), default="UNKNOWN"
        ),
        "ticket_period_start": min(
            (_ticket_window_start(record) for record in tickets.records), default="UNKNOWN"
        ),
        "ticket_period_end": max(
            (_ticket_window_end(record) for record in tickets.records), default="UNKNOWN"
        ),
        "ticket_activity_period_start": min(
            (record["first_comment_at"] for record in tickets.records), default="UNKNOWN"
        ),
        "ticket_activity_period_end": max(
            (
                clean_text(record.get("last_activity_at")) or record["last_comment_at"]
                for record in tickets.records
            ),
            default="UNKNOWN",
        ),
    }
    known_checks = _known_sample_checks(source_profile, recurrence)
    report = build_validation_report(source_profile, warnings, known_checks)
    data_quality_records = [
        {"record_type": "WARNING", "message": warning} for warning in warnings
    ] + [
        {"record_type": "REJECTED_ROW", **rejected_row}
        for rejected_row in rejected_rows
    ]

    output_dir.mkdir(parents=True, exist_ok=True)
    manifest_inputs = [
        {
            "filename": path.name,
            "input_type": "master_data"
            if master_data_path is not None and path == master_data_path
            else "raw_export",
            "sha256": sha256_file(path),
        }
        for path in sorted(
            [path for path in inputs if path.is_file()]
            + ([master_data_path] if master_data_path is not None else [])
        )
    ]
    if zabbix_path.is_dir():
        manifest_inputs.extend(
            {
                "filename": f"{zabbix_path.name}/{path.name}",
                "input_type": "raw_export",
                "sha256": sha256_file(path),
            }
            for path in sorted(zabbix_path.glob("*.csv"), key=lambda item: item.name.lower())
        )
    if issue_path.is_dir():
        manifest_inputs.extend(
            {
                "filename": f"{issue_path.name}/{path.name}",
                "input_type": "raw_export",
                "sha256": sha256_file(path),
            }
            for path in sorted(issue_path.glob("*.csv"), key=lambda item: item.name.lower())
        )
    manifest = {
        "package_version": PACKAGE_VERSION,
        "generated_at": generated_at,
        "upload_allowed": report["upload_allowed"],
        "validation_status": report["status"],
        "inputs": manifest_inputs,
        "source_profile": source_profile,
        "warnings": warnings,
    }
    _write_json(output_dir / "manifest.json", manifest)
    _write_json(output_dir / "validation_report.json", report)
    _write_json(
        output_dir / "operational_timeline.json",
        build_operational_timeline_payload(operational_stories, generated_at=generated_at),
    )
    (output_dir / "validation_report.md").write_text(
        _render_validation_markdown(report), encoding="utf-8"
    )
    (output_dir / "00_report_context.md").write_text(
        _render_context(source_profile, warnings), encoding="utf-8"
    )
    (output_dir / "01_executive_summary.md").write_text(
        _render_executive_summary(source_profile, recurrence, alert_patterns),
        encoding="utf-8",
    )
    (output_dir / "02_operational_timeline.md").write_text(
        _render_operational_timeline(operational_stories), encoding="utf-8"
    )
    (output_dir / "02_confirmed_incidents.md").write_text(
        _render_incidents(incidents.records), encoding="utf-8"
    )
    (output_dir / "03_recurrence_patterns.md").write_text(
        _render_recurrence(recurrence), encoding="utf-8"
    )
    (output_dir / "04_alert_patterns.md").write_text(
        _render_alert_patterns(alert_patterns), encoding="utf-8"
    )
    (output_dir / "05_ticket_impact.md").write_text(
        _render_ticket_impact(tickets.records), encoding="utf-8"
    )
    (output_dir / "05_ticket_impact_index.md").write_text(
        _render_ticket_partition_index(tickets.records), encoding="utf-8"
    )
    for partition_month in _ticket_partition_months(tickets.records):
        partition_tickets = _ticket_partition_records(tickets.records, partition_month)
        (output_dir / _ticket_partition_filename(partition_month)).write_text(
            _render_ticket_impact(
                partition_tickets,
                title=f"Ticket Impact Evidence - {partition_month}",
                note=(
                    "Monthly ticket partition. Select by activity month, then dedupe by "
                    "`ticket_id` in the query workflow."
                ),
            ),
            encoding="utf-8",
        )
    (output_dir / "06_data_quality.md").write_text(
        _render_data_quality(incidents.records, master_data, rejected_rows, warnings),
        encoding="utf-8",
    )
    _write_audit_workbook(
        output_dir / "normalized_data.xlsx",
        operational_stories,
        incidents.records,
        alert_patterns,
        tickets.records,
        recurrence,
        data_quality_records,
        source_profile,
    )
    return {
        "output_dir": str(output_dir),
        "upload_allowed": report["upload_allowed"],
        "validation_status": report["status"],
        "source_profile": source_profile,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build an Alpha Intelligence IT Operations knowledge package."
    )
    parser.add_argument("--raw-dir", type=Path, required=True, help="Directory containing the three raw Excel exports.")
    parser.add_argument("--master-data", type=Path, help="Optional reviewed master-data workbook.")
    parser.add_argument("--output-dir", type=Path, required=True, help="Directory for the generated Alpha-ready package.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    package = build_knowledge_package(args.raw_dir, args.master_data, args.output_dir)
    print(json.dumps(package, ensure_ascii=False, indent=2, sort_keys=True))
    return 0 if package["upload_allowed"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
