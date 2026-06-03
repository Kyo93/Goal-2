from pathlib import Path

from openpyxl import Workbook


FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"

ISSUE_HEADERS = [
    "itcenter.ticket.comment.created_at",
    "itcenter.ticket.id",
    "itcenter.ticket.category_en_name",
    "itcenter.ticket.classification_name",
    "itcenter.ticket.full_title",
    "itcenter.ticket.service_recipient.user.business_unit",
    "itcenter.ticket.service_recipient.user.department_name",
    "itcenter.ticket.service_recipient.user.name",
    "itcenter.ticket.service_recipient.user.email",
    "itcenter.ticket.assignee.user.name",
    "itcenter.ticket.comment.text",
]

INCIDENT_HEADERS = [
    "Timestamp",
    "ISP",
    "SITE / Location ",
    "Incident DATE",
    "Incident TIME",
    "Reporter \n",
    "Incindent Type",
    "Severity ",
    "Incident description",
    "Initial Cause",
    "Troubleshooting actions",
    "Root Cause",
    "Incident resolution date",
    "Incident resolution time",
    "Measures to prevent recurrence of incidents (if any) ",
]

ZABBIX_HEADERS = [
    "Severity",
    "Time",
    "Recovery time",
    "Status",
    "Host",
    "Problem",
    "Duration",
    "Ack",
    "Actions",
    "Tags",
]


def write_issue_report():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Issue Export"
    worksheet.append(ISSUE_HEADERS)
    worksheet.append(
        [
            "May 19, 2026 @ 09:00:00.000",
            "TKT-100",
            "Issue Report",
            "Computer",
            "Laptop cannot connect to Wi-Fi",
            "SPX",
            "Field Operations (Facility)",
            "User One",
            "user.one@example.com",
            "Assignee A",
            "Initial troubleshooting requested.",
        ]
    )
    worksheet.append(
        [
            "May 19, 2026 @ 10:00:00.000",
            "TKT-100",
            "Issue Report",
            "Computer",
            "Laptop cannot connect to Wi-Fi",
            "SPX",
            "Field Operations (Facility)",
            "User One",
            "user.one@example.com",
            "Assignee A",
            "Resolved after reconnecting to SSID.",
        ]
    )
    worksheet.append([""] * len(ISSUE_HEADERS))
    worksheet.append(
        [
            "May 19, 2026 @ 11:00:00.000",
            "TKT-200",
            "Issue Report",
            "System & Network Access",
            "VPN password reset required",
            "SPX",
            "Field Operations (Facility)",
            "User Two",
            "user.two@example.com",
            "Assignee B",
            "User cannot sign in to VPN.",
        ]
    )
    workbook.create_sheet("Sheet1")
    workbook.save(FIXTURE_DIR / "IssueReport.xlsx")


def write_incident_report():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Form Responses 1"
    worksheet.append(INCIDENT_HEADERS)
    worksheet.append([""] * len(INCIDENT_HEADERS))
    rows = [
        [
            "2026-01-01T08:30:00",
            "VNPT",
            "MS2 - Mega SOC - Cu Chi",
            "2026-01-01T00:00:00",
            "08:00:00",
            "Reporter A",
            "High latency",
            "Minor",
            "Latency increased.",
            "International routing fluctuations.",
            "Optimized routing.",
            "International routing fluctuations.",
            "2026-01-01T00:00:00",
            "08:20:00",
            "Continue monitoring.",
        ],
        [
            "2026-01-02T08:30:00",
            "VNPT",
            "MS2 - Mega SOC - Cu Chi",
            "2026-01-02T00:00:00",
            "08:00:00",
            "Reporter B",
            "High latency",
            "Minor",
            "Latency increased again.",
            "",
            "Optimized routing.",
            "",
            "2026-01-02T00:00:00",
            "08:10:00",
            "",
        ],
        [
            "2026-01-03T09:00:00",
            "CMC",
            "XAS - Warehouse - Xuyen A",
            "2026-01-03T00:00:00",
            "08:00:00",
            "Reporter C",
            "Fiber optic cable failure",
            "Major",
            "Cable damaged.",
            "Cable fire.",
            "Repaired cable.",
            "Cable fire.",
            "2026-01-03T00:00:00",
            "09:00:00",
            "Inspect cable route.",
        ],
        [
            "2026-01-04T09:00:00",
            "FPT",
            "SWS - SOC - SW",
            "not-a-date",
            "08:00:00",
            "Reporter D",
            "Total outage",
            "Critical",
            "Invalid timestamp fixture.",
            "",
            "Investigating.",
            "",
            "",
            "",
            "",
        ],
    ]
    for row in rows:
        worksheet.append(row)

    subset = workbook.create_sheet("MS2 -VNPT")
    subset.append(INCIDENT_HEADERS)
    subset.append(rows[0])
    subset.append(rows[1])
    workbook.save(FIXTURE_DIR / "SEA - Corp IT- ILL- Incident Report   (Responses).xlsx")


def write_zabbix_export():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "zbx_problems_export"
    worksheet.append(ZABBIX_HEADERS)
    rows = [
        [
            "Not classified",
            "2026-05-19T16:20:00",
            "2026-05-19T16:21:00",
            "RESOLVED",
            "VNMCCW-PNWSW01",
            "FPT Download Gi2/0/3>100M,Current:110.04 Mbps",
            "1m",
            "No",
            "Actions (2)",
            "class: network, component: network, description: fpt",
        ],
        [
            "Not classified",
            "2026-05-19T16:22:00",
            "2026-05-19T16:23:00",
            "RESOLVED",
            "VNMCCW-PNWSW01",
            "FPT Download Gi2/0/3>100M,Current:115.17 Mbps",
            "1m",
            "No",
            "Actions (2)",
            "class: network, component: network, description: fpt",
        ],
        [
            "High",
            "2026-05-19T16:24:00",
            "2026-05-19T16:33:00",
            "RESOLVED",
            "VNMSNT-UPSF2201",
            "APC UPS: Unacceptable input frequency (out of range 49.7-50.3Hz for 15m)",
            "9m",
            "No",
            "Actions (2)",
            "class: power, component: power, scope: notice",
        ],
        [
            "High",
            "2026-05-19T16:34:00",
            "2026-05-19T16:43:00",
            "RESOLVED",
            "VNMSNT-UPSF2201",
            "APC UPS: Unacceptable input frequency (out of range 49.8-50.2Hz for 15m)",
            "9m",
            "No",
            "Actions (2)",
            "class: power, component: power, scope: notice",
        ],
        [
            "Warning",
            "2026-05-19T16:40:00",
            "",
            "PROBLEM",
            "VNMCPL-PNCSW01",
            "Interface Vl100(to-pa1c): High bandwidth usage (>90%)",
            "23m",
            "No",
            "Actions (1)",
            "class: network, component: network, scope: performance",
        ],
        [
            "Warning",
            "not-a-timestamp",
            "",
            "PROBLEM",
            "VNMXXX-INVALID",
            "Invalid fixture row",
            "",
            "No",
            "",
            "class: network",
        ],
    ]
    for row in rows:
        worksheet.append(row)
    workbook.save(FIXTURE_DIR / "zbx_problems_export.xlsx")


def write_master_data():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "sites"
    worksheet.append(["site_code", "site_name", "review_status"])
    worksheet.append(["CCW", "CCW - Warehouse - Cu Chi", "CONFIRMED"])
    worksheet.append(["SNT", "SNT - Sonatus Office", "NEEDS_REVIEW"])

    hosts = workbook.create_sheet("hosts")
    hosts.append(["host_name", "site_code", "service_code", "review_status"])
    hosts.append(["VNMCCW-PNWSW01", "CCW", "site-connectivity", "CONFIRMED"])
    hosts.append(["VNMSNT-UPSF2201", "SNT", "site-power", "NEEDS_REVIEW"])
    workbook.save(FIXTURE_DIR / "master_data.xlsx")


def main():
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    write_issue_report()
    write_incident_report()
    write_zabbix_export()
    write_master_data()
    print(f"Created fixtures in {FIXTURE_DIR}")


if __name__ == "__main__":
    main()
