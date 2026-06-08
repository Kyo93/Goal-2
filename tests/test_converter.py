import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
FIXTURES = ROOT / "tests" / "fixtures"
sys.path.insert(0, str(SRC))

from itops_alpha.converter import (  # noqa: E402
    build_operational_timeline,
    build_operational_stories,
    build_knowledge_package,
    build_validation_report,
    classify_responsibility_domain,
    group_alert_patterns,
    group_incident_recurrence,
    load_confirmed_master_data,
    load_incidents,
    load_issue_tickets,
    load_zabbix_alerts,
    normalize_problem_signature,
    query_operational_timeline,
    source_ref,
)


class ConverterTest(unittest.TestCase):
    def test_source_ref_is_stable_and_human_readable(self):
        self.assertEqual(
            source_ref("IssueReport.xlsx", "Issue Export", 2),
            "IssueReport.xlsx#Issue Export:row-2",
        )

    def test_incident_adapter_reads_main_sheet_only_and_rejects_invalid_timestamp(self):
        result = load_incidents(
            FIXTURES / "SEA - Corp IT- ILL- Incident Report   (Responses).xlsx"
        )
        self.assertEqual(result.input_rows, 4)
        self.assertEqual(len(result.records), 3)
        self.assertEqual(len(result.rejected_rows), 1)
        self.assertIn("not-a-date", result.rejected_rows[0]["reason"])
        self.assertEqual(result.records[0]["incident_id"], "INC-3")
        self.assertIn("#Form Responses 1:row-3", result.records[0]["source_ref"])

    def test_issue_adapter_aggregates_comments_by_ticket_without_losing_count(self):
        result = load_issue_tickets(FIXTURES / "IssueReport.xlsx")
        self.assertEqual(result.input_rows, 3)
        self.assertEqual(len(result.records), 2)
        self.assertEqual(sum(record["comment_count"] for record in result.records), 3)
        ticket = next(record for record in result.records if record["ticket_id"] == "TKT-100")
        self.assertEqual(ticket["comment_count"], 2)
        self.assertEqual(len(ticket["source_refs"]), 2)

    def test_issue_adapter_reads_monthly_csv_partitions_and_preserves_ticket_created_at(self):
        with tempfile.TemporaryDirectory() as temporary:
            ticket_dir = Path(temporary) / "Ticket"
            ticket_dir.mkdir()
            header = ",".join(
                [
                    "itcenter.ticket.id",
                    "itcenter.ticket.created_at",
                    "itcenter.ticket.full_title",
                    "itcenter.ticket.category_en_name",
                    "itcenter.ticket.classification_name",
                    "itcenter.ticket.service_recipient.user.business_unit",
                    "itcenter.ticket.service_recipient.user.entity_name",
                    "itcenter.ticket.location_full_name",
                    "itcenter.ticket.comment.created_by.user.email",
                    "itcenter.ticket.comment.text",
                    "itcenter.ticket.comment.updated_at",
                    "itcenter.ticket.comment.created_at",
                    "itcenter.ticket.service_recipient.user.office_display",
                    "itcenter.ticket.subclassification_name",
                ]
            )
            (ticket_dir / "tickets_2026_05.csv").write_text(
                "\n".join(
                    [
                        header,
                        '"TKT-100","May 1, 2026 @ 08:00:00.000","CPL VPN issue","Issue Report","General IT Inquiry","SPX","SPX Express","SOC - Hanoi","user@example.com","User reported CPL access issue","May 1, 2026 @ 08:25:00.000","May 1, 2026 @ 08:20:00.000","VN > Hanoi > Capital Place","VPN"',
                        '"TKT-100","May 1, 2026 @ 08:00:00.000","CPL VPN issue","Issue Report","General IT Inquiry","SPX","SPX Express","SOC - Hanoi","agent@itcenter.sea.com","IT replied","May 1, 2026 @ 09:00:00.000","May 1, 2026 @ 09:00:00.000","VN > Hanoi > Capital Place","VPN"',
                        '"TKT-200","May 2, 2026 @ 10:00:00.000","Printer issue","Issue Report","Equipment","SPX","SPX Express","SPX - HCM HUB","user@example.com","Printer broken","May 2, 2026 @ 10:10:00.000","May 2, 2026 @ 10:05:00.000","HCM Mega SOC","Printer"',
                    ]
                ),
                encoding="utf-8",
            )
            (ticket_dir / "tickets_2026_06.csv").write_text(
                "\n".join(
                    [
                        header,
                        '"TKT-100","May 1, 2026 @ 08:00:00.000","CPL VPN issue","Issue Report","General IT Inquiry","SPX","SPX Express","SOC - Hanoi","agent@itcenter.sea.com","Follow-up in June","Jun 1, 2026 @ 09:10:00.000","Jun 1, 2026 @ 09:05:00.000","VN > Hanoi > Capital Place","VPN"',
                    ]
                ),
                encoding="utf-8",
            )

            result = load_issue_tickets(ticket_dir)

        self.assertEqual(result.input_rows, 4)
        self.assertEqual(len(result.records), 2)
        self.assertEqual(sum(record["comment_count"] for record in result.records), 4)
        ticket = next(record for record in result.records if record["ticket_id"] == "TKT-100")
        self.assertEqual(ticket["ticket_created_at"], "2026-05-01T08:00:00")
        self.assertEqual(ticket["evidence_started_at"], "2026-05-01T08:00:00")
        self.assertEqual(ticket["evidence_time_basis"], "ticket_created_at")
        self.assertEqual(ticket["first_comment_at"], "2026-05-01T08:20:00")
        self.assertEqual(ticket["last_comment_at"], "2026-06-01T09:05:00")
        self.assertEqual(ticket["last_activity_at"], "2026-06-01T09:10:00")
        self.assertEqual(ticket["partition_months"], ["2026-05", "2026-06"])
        self.assertEqual(
            ticket["source_files"], ["tickets_2026_05.csv", "tickets_2026_06.csv"]
        )
        self.assertEqual(ticket["location_full_name"], "SOC - Hanoi")
        self.assertEqual(ticket["office_display"], "VN > Hanoi > Capital Place")
        self.assertEqual(ticket["subclassification"], "VPN")

    def test_issue_adapter_builds_lifecycle_comment_intelligence(self):
        with tempfile.TemporaryDirectory() as temporary:
            ticket_dir = Path(temporary) / "Ticket"
            ticket_dir.mkdir()
            header = ",".join(
                [
                    "itcenter.ticket.id",
                    "itcenter.ticket.created_at",
                    "itcenter.ticket.full_title",
                    "itcenter.ticket.category_en_name",
                    "itcenter.ticket.classification_name",
                    "itcenter.ticket.service_recipient.user.business_unit",
                    "itcenter.ticket.service_recipient.user.entity_name",
                    "itcenter.ticket.location_full_name",
                    "itcenter.ticket.comment.created_by.user.email",
                    "itcenter.ticket.comment.text",
                    "itcenter.ticket.comment.updated_at",
                    "itcenter.ticket.comment.created_at",
                    "itcenter.ticket.service_recipient.user.office_display",
                    "itcenter.ticket.subclassification_name",
                ]
            )
            rows = [
                header,
                '"TKT-900","May 3, 2026 @ 08:00:00.000","CPL VPN cannot access","Issue Report","General IT Inquiry","SPX","SPX Express","SOC - Hanoi","user@example.com","User cannot access VPN from CPL","May 3, 2026 @ 08:05:00.000","May 3, 2026 @ 08:05:00.000","VN > Hanoi > Capital Place","VPN"',
                '"TKT-900","May 3, 2026 @ 08:00:00.000","CPL VPN cannot access","Issue Report","General IT Inquiry","SPX","SPX Express","SOC - Hanoi","agent@example.com","Helpdesk asked user to retry login","May 3, 2026 @ 08:10:00.000","May 3, 2026 @ 08:10:00.000","VN > Hanoi > Capital Place","VPN"',
                '"TKT-900","May 3, 2026 @ 08:00:00.000","CPL VPN cannot access","Issue Report","General IT Inquiry","SPX","SPX Express","SOC - Hanoi","agent@example.com","Escalated to network team for ISP routing check","May 3, 2026 @ 08:30:00.000","May 3, 2026 @ 08:30:00.000","VN > Hanoi > Capital Place","VPN"',
                '"TKT-900","May 3, 2026 @ 08:00:00.000","CPL VPN cannot access","Issue Report","General IT Inquiry","SPX","SPX Express","SOC - Hanoi","network@example.com","Network team found latency on path","May 3, 2026 @ 09:00:00.000","May 3, 2026 @ 09:00:00.000","VN > Hanoi > Capital Place","VPN"',
                '"TKT-900","May 3, 2026 @ 08:00:00.000","CPL VPN cannot access","Issue Report","General IT Inquiry","SPX","SPX Express","SOC - Hanoi","network@example.com","VPN restored after routing fix","May 3, 2026 @ 09:40:00.000","May 3, 2026 @ 09:40:00.000","VN > Hanoi > Capital Place","VPN"',
                '"TKT-900","May 3, 2026 @ 08:00:00.000","CPL VPN cannot access","Issue Report","General IT Inquiry","SPX","SPX Express","SOC - Hanoi","user@example.com","User confirmed working now","May 3, 2026 @ 10:00:00.000","May 3, 2026 @ 10:00:00.000","VN > Hanoi > Capital Place","VPN"',
            ]
            (ticket_dir / "tickets_2026_05.csv").write_text(
                "\n".join(rows), encoding="utf-8"
            )

            result = load_issue_tickets(ticket_dir)

        ticket = result.records[0]
        self.assertEqual(
            ticket["initial_comments_sample"],
            ["User cannot access VPN from CPL", "Helpdesk asked user to retry login"],
        )
        self.assertEqual(
            ticket["key_comments_sample"],
            ["Escalated to network team for ISP routing check"],
        )
        self.assertEqual(
            ticket["final_comments_sample"],
            [
                "Network team found latency on path",
                "VPN restored after routing fix",
                "User confirmed working now",
            ],
        )
        self.assertIn("Final comment sample:", ticket["comments_summary"])
        self.assertIn("User confirmed working now", ticket["comments_summary"])
        self.assertEqual(ticket["comment_summary_strategy"], "lifecycle_key_sample")
        self.assertEqual(ticket["symptom_family"], "vpn_or_remote_access")
        self.assertEqual(ticket["resolution_signal"], "RESOLVED")
        self.assertIn("ESCALATION_SIGNAL", ticket["detected_comment_signals"])
        self.assertIn("USER_CONFIRMATION_SIGNAL", ticket["detected_comment_signals"])

    def test_issue_adapter_truncates_long_comment_samples_but_keeps_source_refs(self):
        with tempfile.TemporaryDirectory() as temporary:
            ticket_dir = Path(temporary) / "Ticket"
            ticket_dir.mkdir()
            header = ",".join(
                [
                    "itcenter.ticket.id",
                    "itcenter.ticket.created_at",
                    "itcenter.ticket.full_title",
                    "itcenter.ticket.category_en_name",
                    "itcenter.ticket.classification_name",
                    "itcenter.ticket.service_recipient.user.business_unit",
                    "itcenter.ticket.service_recipient.user.entity_name",
                    "itcenter.ticket.location_full_name",
                    "itcenter.ticket.comment.created_by.user.email",
                    "itcenter.ticket.comment.text",
                    "itcenter.ticket.comment.updated_at",
                    "itcenter.ticket.comment.created_at",
                    "itcenter.ticket.service_recipient.user.office_display",
                    "itcenter.ticket.subclassification_name",
                ]
            )
            long_comment = "VPN restored and user confirmed. " + ("diagnostic detail " * 200)
            rows = [
                header,
                f'"TKT-901","May 4, 2026 @ 08:00:00.000","VPN access fixed","Issue Report","General IT Inquiry","SPX","SPX Express","SOC - Hanoi","user@example.com","Initial VPN issue","May 4, 2026 @ 08:05:00.000","May 4, 2026 @ 08:05:00.000","VN > Hanoi > Capital Place","VPN"',
                f'"TKT-901","May 4, 2026 @ 08:00:00.000","VPN access fixed","Issue Report","General IT Inquiry","SPX","SPX Express","SOC - Hanoi","agent@example.com","{long_comment}","May 4, 2026 @ 09:00:00.000","May 4, 2026 @ 09:00:00.000","VN > Hanoi > Capital Place","VPN"',
            ]
            (ticket_dir / "tickets_2026_05.csv").write_text(
                "\n".join(rows), encoding="utf-8"
            )

            result = load_issue_tickets(ticket_dir)

        ticket = result.records[0]
        self.assertIn("[truncated; see source_refs]", ticket["final_comments_sample"][-1])
        self.assertLess(len(ticket["final_comments_sample"][-1]), len(long_comment))
        self.assertEqual(len(ticket["source_refs"]), 2)
        self.assertIn("tickets_2026_05.csv#csv:row-3", ticket["source_refs"])

    def test_zabbix_adapter_groups_metric_variants_but_preserves_raw_count(self):
        result = load_zabbix_alerts(FIXTURES / "zbx_problems_export.xlsx")
        self.assertEqual(result.input_rows, 6)
        self.assertEqual(len(result.records), 5)
        self.assertEqual(len(result.rejected_rows), 1)
        patterns = group_alert_patterns(result.records)
        self.assertEqual(len(patterns), 3)
        fpt = next(pattern for pattern in patterns if pattern["host"] == "VNMCCW-PNWSW01")
        self.assertEqual(fpt["alert_count"], 2)
        self.assertIn("<value>", fpt["problem_signature"])

    def test_zabbix_adapter_reads_multi_csv_exports_and_deduplicates_overlap(self):
        with tempfile.TemporaryDirectory() as temporary:
            export_dir = Path(temporary) / "Export zabbix"
            export_dir.mkdir()
            header = '"Severity","Time","Recovery time","Status","Host","Problem","Duration","Ack","Actions","Tags"\n'
            duplicate = '"High","2026-06-03 03:11:37 PM","2026-06-03 03:12:07 PM","RESOLVED","VNMMS2-FPT-FMS-45.119.218.130","Unavailable by ICMP ping","30s","No","Actions (2)","Application: Status"\n'
            unique = '"Warning","2026-06-03 03:10:37 PM","2026-06-03 03:10:52 PM","RESOLVED","VNMMSB-FPT-FMS-45.119.218.130","High ICMP ping loss","15s","No","Actions (2)","Application: Status"\n'
            (export_dir / "part-a.csv").write_text(header + duplicate, encoding="utf-8")
            (export_dir / "part-b.csv").write_text(header + duplicate + unique, encoding="utf-8")

            result = load_zabbix_alerts(export_dir)

        self.assertEqual(result.input_rows, 3)
        self.assertEqual(len(result.records), 2)
        self.assertEqual(result.duplicate_rows, 1)
        self.assertEqual(len(result.rejected_rows), 0)
        self.assertEqual(result.records[0]["seen_at"], "2026-06-03T15:11:37")
        self.assertIn("part-a.csv#csv:row-2", result.records[0]["source_ref"])

    def test_problem_signature_replaces_dynamic_values(self):
        signature = normalize_problem_signature(
            "FPT Download Gi2/0/3>100M,Current:115.17 Mbps"
        )
        self.assertEqual(
            signature,
            "fpt download gi2/0/3><value>,current:<value> mbps",
        )

    def test_master_data_uses_confirmed_rows_only(self):
        data = load_confirmed_master_data(FIXTURES / "master_data.xlsx")
        self.assertEqual(set(data["sites"]), {"CCW"})
        self.assertEqual(set(data["hosts"]), {"VNMCCW-PNWSW01"})
        self.assertEqual(data["unconfirmed_rows"], 2)

    def test_incident_recurrence_is_deterministic(self):
        incidents = load_incidents(
            FIXTURES / "SEA - Corp IT- ILL- Incident Report   (Responses).xlsx"
        ).records
        recurrence = group_incident_recurrence(incidents)
        ms2 = next(pattern for pattern in recurrence if pattern["site_code"] == "MS2")
        self.assertEqual(ms2["incident_count"], 2)
        self.assertTrue(ms2["recurrence_id"].startswith("REC-"))

    def test_responsibility_domain_classification_is_conservative_and_auditable(self):
        self.assertEqual(
            classify_responsibility_domain(
                "The issue was caused by international routing fluctuations."
            ),
            ("ISP", "RCA keyword rule: routing"),
        )
        self.assertEqual(
            classify_responsibility_domain("The issue belongs to Shopee's local segment."),
            ("INHOUSE", "RCA keyword rule: local segment"),
        )
        self.assertEqual(
            classify_responsibility_domain("UNKNOWN"),
            ("UNKNOWN", "RCA does not support responsibility classification"),
        )

    def test_operational_timeline_links_only_matching_site_and_time_windows(self):
        incidents = [
            {
                "incident_id": "INC-3",
                "source_ref": "incident.xlsx#main:row-3",
                "started_at": "2026-05-19T16:10:00",
                "resolved_at": "2026-05-19T16:40:00",
                "site_code": "CCW",
                "incident_type": "High latency",
                "description": "Users are experiencing slow access.",
                "root_cause": "FPT uplink routing issue.",
                "resolution_status": "RESOLVED",
            }
        ]
        alert_patterns = [
            {
                "alert_pattern_id": "ALP-MATCH",
                "source_refs": ["zabbix.xlsx#main:row-2"],
                "first_seen_at": "2026-05-19T16:20:00",
                "last_seen_at": "2026-05-19T16:30:00",
                "site_code": "CCW",
                "host": "VNMCCW-PNWSW01",
                "problem_signature": "high latency",
                "alert_count": 2,
            },
            {
                "alert_pattern_id": "ALP-WRONG-SITE",
                "source_refs": ["zabbix.xlsx#main:row-3"],
                "first_seen_at": "2026-05-19T16:20:00",
                "last_seen_at": "2026-05-19T16:30:00",
                "site_code": "SNT",
                "host": "VNMSNT-UPSF2201",
                "problem_signature": "power",
                "alert_count": 1,
            },
        ]
        tickets = [
            {
                "ticket_id": "TKT-100",
                "source_refs": ["IssueReport.xlsx#main:row-2"],
                "first_comment_at": "2026-05-19T16:15:00",
                "last_comment_at": "2026-05-19T16:25:00",
                "title": "CCW users cannot access service",
                "comments_summary": "CCW users report slow access.",
            },
            {
                "ticket_id": "TKT-200",
                "source_refs": ["IssueReport.xlsx#main:row-3"],
                "first_comment_at": "2026-05-19T16:15:00",
                "last_comment_at": "2026-05-19T16:25:00",
                "title": "SNT issue",
                "comments_summary": "SNT power alert.",
            },
        ]

        timeline = build_operational_timeline(incidents, alert_patterns, tickets)
        incident = next(event for event in timeline if event["event_id"] == "EVT-INC-3")
        signal = next(event for event in timeline if event["event_id"] == "EVT-ALP-MATCH")

        self.assertEqual(incident["event_type"], "CONFIRMED_INCIDENT")
        self.assertEqual(incident["related_alert_pattern_ids"], ["ALP-MATCH"])
        self.assertEqual(incident["related_ticket_ids"], ["TKT-100"])
        self.assertEqual(incident["user_impact_status"], "CONFIRMED")
        self.assertEqual(incident["responsibility_domain"], "ISP")
        self.assertEqual(signal["event_type"], "ZABBIX_ALERT_PATTERN")
        self.assertEqual(signal["related_incident_ids"], ["INC-3"])

        query = query_operational_timeline(
            timeline,
            site_code="ccw",
            from_at="2026-05-19T16:00:00",
            to_at="2026-05-19T16:50:00",
        )
        self.assertEqual(query["matched_event_count"], 2)
        self.assertEqual(
            query["counts_by_episode_type"],
            {"CONFIRMED_INCIDENT": 1, "ZABBIX_ALERT_PATTERN": 1},
        )
        self.assertEqual(
            [record["event_id"] for record in query["records"]],
            ["EVT-INC-3", "EVT-ALP-MATCH"],
        )

    def test_operational_timeline_query_excludes_non_overlapping_ranges(self):
        timeline = [
            {
                "episode_id": "OPS-INC-1",
                "episode_type": "CONFIRMED_INCIDENT",
                "started_at": "2026-05-19T10:00:00",
                "ended_at": "2026-05-19T11:00:00",
                "site_code": "MS2",
            },
            {
                "episode_id": "OPS-INC-2",
                "episode_type": "CONFIRMED_INCIDENT",
                "started_at": "2026-05-19T10:30:00",
                "ended_at": "UNKNOWN",
                "site_code": "SNT",
            },
        ]

        same_site = query_operational_timeline(
            timeline,
            site_code="MS2",
            from_at="2026-05-19T10:30:00",
            to_at="2026-05-19T12:00:00",
        )
        wrong_site = query_operational_timeline(
            timeline,
            site_code="SNT",
            from_at="2026-05-19T11:30:00",
            to_at="2026-05-19T12:00:00",
        )

        self.assertEqual(same_site["matched_event_count"], 1)
        self.assertEqual(wrong_site["matched_event_count"], 0)

    def test_operational_timeline_query_uses_monitoring_milestones_not_long_pattern_span(self):
        timeline = [
            {
                "episode_id": "OPS-ALP-OLD",
                "episode_type": "MONITORING_SIGNAL",
                "started_at": "2026-04-01T00:00:00",
                "ended_at": "2026-05-27T04:30:00",
                "site_code": "SWS",
                "milestones": [
                    {
                        "at": "2026-04-01T00:00:00",
                        "type": "ZABBIX_DETECTED",
                        "message": "Old alert.",
                    },
                    {
                        "at": "2026-05-27T04:30:00",
                        "type": "SIGNAL_RECOVERED",
                        "message": "Recovered outside query.",
                    },
                ],
            },
            {
                "episode_id": "OPS-ALP-MATCH",
                "episode_type": "MONITORING_SIGNAL",
                "started_at": "2026-04-29T19:33:48",
                "ended_at": "2026-05-27T00:54:48",
                "site_code": "SWS",
                "milestones": [
                    {
                        "at": "2026-05-27T00:48:18",
                        "type": "ALERTS_GROUPED",
                        "message": "Grouped raw alerts.",
                    }
                ],
            },
        ]

        query = query_operational_timeline(
            timeline,
            site_code="SWS",
            from_at="2026-05-27T00:00:00",
            to_at="2026-05-27T02:00:00",
        )

        self.assertEqual(query["matched_event_count"], 1)
        self.assertEqual(query["records"][0]["episode_id"], "OPS-ALP-MATCH")

    def test_operational_timeline_query_summarizes_zabbix_patterns_by_family(self):
        raw_alerts = [
            {
                "alert_id": "ZBX-NET-OLD",
                "source_ref": "zabbix.csv#row-1",
                "seen_at": "2026-04-30T23:50:00",
                "recovered_at": "2026-04-30T23:55:00",
                "status": "RESOLVED",
                "site_code": "CPL",
                "host": "VNMCPL-PNWFW01",
                "severity": "High",
                "problem": "VNMCPL-PNWFW01 is down!",
                "problem_signature": "vnmcpl-pnwfw01 is down!",
                "domain": "network",
                "component": "firewall",
                "scope": "connectivity",
            },
            {
                "alert_id": "ZBX-NET-1",
                "source_ref": "zabbix.csv#row-2",
                "seen_at": "2026-05-17T16:38:17",
                "recovered_at": "2026-05-17T16:41:19",
                "status": "RESOLVED",
                "site_code": "CPL",
                "host": "VNMCPL-PNWFW01",
                "severity": "High",
                "problem": "VNMCPL-PNWFW01 is down!",
                "problem_signature": "vnmcpl-pnwfw01 is down!",
                "domain": "network",
                "component": "firewall",
                "scope": "connectivity",
            },
            {
                "alert_id": "ZBX-NET-2",
                "source_ref": "zabbix.csv#row-3",
                "seen_at": "2026-05-17T16:39:05",
                "recovered_at": "2026-05-17T16:41:04",
                "status": "RESOLVED",
                "site_code": "CPL",
                "host": "VNMCPL-PNWFW01",
                "severity": "High",
                "problem": "VNMCPL-PNWFW01 is down!",
                "problem_signature": "vnmcpl-pnwfw01 is down!",
                "domain": "network",
                "component": "firewall",
                "scope": "connectivity",
            },
            {
                "alert_id": "ZBX-HTTP-1",
                "source_ref": "zabbix.csv#row-4",
                "seen_at": "2026-05-31T06:01:40",
                "recovered_at": "2026-05-31T06:07:40",
                "status": "RESOLVED",
                "site_code": "CPL",
                "host": "VNMCPL-VSSPM01",
                "severity": "Warning",
                "problem": "Jira HTTP monitoring",
                "problem_signature": "jira http monitoring",
                "domain": "application",
                "component": "http",
                "scope": "service",
            },
        ]
        stories = build_operational_stories(
            incidents=[],
            alert_patterns=group_alert_patterns(raw_alerts),
            tickets=[],
            recurrence=[],
            alerts=raw_alerts,
        )

        query = query_operational_timeline(
            stories,
            site_code="CPL",
            from_at="2026-05-01T00:00:00",
            to_at="2026-06-01T00:00:00",
        )

        self.assertEqual(query["matched_event_count"], 2)
        self.assertEqual(query["monitoring_family_summary"][0]["pattern_family"], "Network reachability")
        self.assertEqual(query["monitoring_family_summary"][0]["query_alert_count"], 2)
        self.assertEqual(query["monitoring_family_summary"][0]["investigation_priorities"], ["HIGH"])
        self.assertEqual(query["monitoring_family_summary"][1]["pattern_family"], "HTTP monitoring")
        network_record = next(
            record
            for record in query["records"]
            if record["problem_signature"] == "vnmcpl-pnwfw01 is down!"
        )
        self.assertEqual(network_record["query_alert_count"], 2)
        self.assertEqual(network_record["query_first_seen_at"], "2026-05-17T16:38:17")
        self.assertEqual(network_record["query_last_seen_at"], "2026-05-17T16:39:05")

    def test_operational_timeline_uses_event_level_alert_overlap(self):
        incidents = [
            {
                "incident_id": "INC-MDN",
                "source_ref": "incident.xlsx#main:row-10",
                "started_at": "2026-05-10T13:20:00",
                "resolved_at": "2026-05-10T17:50:00",
                "site_code": "MDN",
                "incident_type": "Others",
                "description": "Local segment issue.",
                "root_cause": "The issue belongs to Shopee's local segment.",
                "resolution_status": "RESOLVED",
            },
            {
                "incident_id": "INC-MBD",
                "source_ref": "incident.xlsx#main:row-11",
                "started_at": "2026-05-29T12:20:00",
                "resolved_at": "2026-05-30T15:48:00",
                "site_code": "MBD",
                "incident_type": "Fiber optic cable failure",
                "description": "Fiber attenuation near customer site.",
                "root_cause": "Fiber attenuation on the cable section.",
                "resolution_status": "RESOLVED",
            },
        ]
        alerts = [
            {
                "alert_id": "ALT-MDN-1",
                "source_ref": "zabbix.csv#csv:row-2",
                "seen_at": "2026-05-05T09:39:20",
                "recovered_at": "2026-05-05T09:40:20",
                "status": "RESOLVED",
                "severity": "High",
                "host": "VNMMDN-VSSPM01",
                "site_code": "MDN",
                "problem": "HTTP Monitoring",
                "problem_signature": "http monitoring",
                "duration": "1m",
                "ack": "No",
                "actions": "UNKNOWN",
                "domain": "UNKNOWN",
                "component": "UNKNOWN",
                "scope": "UNKNOWN",
                "tags": "{}",
                "evidence_label": "SOURCE FACT",
            },
            {
                "alert_id": "ALT-MDN-2",
                "source_ref": "zabbix.csv#csv:row-3",
                "seen_at": "2026-05-30T15:31:20",
                "recovered_at": "2026-05-30T15:32:20",
                "status": "RESOLVED",
                "severity": "High",
                "host": "VNMMDN-VSSPM01",
                "site_code": "MDN",
                "problem": "HTTP Monitoring",
                "problem_signature": "http monitoring",
                "duration": "1m",
                "ack": "No",
                "actions": "UNKNOWN",
                "domain": "UNKNOWN",
                "component": "UNKNOWN",
                "scope": "UNKNOWN",
                "tags": "{}",
                "evidence_label": "SOURCE FACT",
            },
            {
                "alert_id": "ALT-MBD-1",
                "source_ref": "zabbix.csv#csv:row-4",
                "seen_at": "2026-05-29T11:05:04",
                "recovered_at": "2026-05-30T15:42:34",
                "status": "RESOLVED",
                "severity": "High",
                "host": "VNMMBD-PE-FPT-42.116.48.97",
                "site_code": "MBD",
                "problem": "Unavailable by ICMP ping",
                "problem_signature": "unavailable by icmp ping",
                "duration": "1d 4h 37m 30s",
                "ack": "No",
                "actions": "UNKNOWN",
                "domain": "UNKNOWN",
                "component": "UNKNOWN",
                "scope": "UNKNOWN",
                "tags": "{}",
                "evidence_label": "SOURCE FACT",
            },
        ]
        patterns = group_alert_patterns(alerts)

        timeline = build_operational_timeline(incidents, patterns, [], alerts=alerts)

        mdn = next(event for event in timeline if event["event_id"] == "EVT-INC-MDN")
        mbd = next(event for event in timeline if event["event_id"] == "EVT-INC-MBD")

        self.assertEqual(mdn["related_alert_pattern_ids"], [])
        self.assertEqual(mbd["related_alert_pattern_ids"], ["ALP-CDCB1323E9CB"])

    def test_operational_story_uses_matching_alert_event_not_pattern_first_seen(self):
        incidents = [
            {
                "incident_id": "INC-53",
                "source_ref": "incident.xlsx#main:row-53",
                "submitted_at": "2026-05-27T08:25:02",
                "started_at": "2026-05-27T00:04:00",
                "resolved_at": "2026-05-27T01:00:00",
                "site_code": "SWS",
                "isp": "FPT",
                "incident_type": "Fiber optic cable failure",
                "description": "Fiber attenuation.",
                "root_cause": "Suy hao cap ha tang",
                "resolution_status": "RESOLVED",
            }
        ]
        alerts = [
            {
                "alert_id": "ALT-OLD",
                "source_ref": "zabbix.csv#csv:row-10",
                "seen_at": "2026-04-29T19:33:48",
                "recovered_at": "2026-04-29T19:40:48",
                "status": "RESOLVED",
                "severity": "High",
                "host": "VNMSWS-PSSJP1",
                "site_code": "SWS",
                "problem": "Windows: Host has been restarted (uptime < 10m)",
                "problem_signature": "windows: host has been restarted (uptime < <value>)",
                "duration": "7m",
                "ack": "No",
                "actions": "UNKNOWN",
                "domain": "os",
                "component": "system",
                "scope": "notice",
                "tags": "{}",
                "evidence_label": "SOURCE FACT",
            },
            {
                "alert_id": "ALT-MATCH",
                "source_ref": "zabbix.csv#csv:row-1001",
                "seen_at": "2026-05-27T00:48:18",
                "recovered_at": "2026-05-27T00:54:48",
                "status": "RESOLVED",
                "severity": "High",
                "host": "VNMSWS-PSSJP1",
                "site_code": "SWS",
                "problem": "Windows: Host has been restarted (uptime < 10m)",
                "problem_signature": "windows: host has been restarted (uptime < <value>)",
                "duration": "6m 30s",
                "ack": "No",
                "actions": "UNKNOWN",
                "domain": "os",
                "component": "system",
                "scope": "notice",
                "tags": "{}",
                "evidence_label": "SOURCE FACT",
            },
        ]
        patterns = group_alert_patterns(alerts)
        stories = build_operational_stories(
            incidents,
            patterns,
            tickets=[],
            recurrence=[],
            alerts=alerts,
        )

        incident_story = next(story for story in stories if story["episode_id"] == "OPS-INC-53")
        zabbix_milestones = [
            milestone
            for milestone in incident_story["milestones"]
            if milestone["type"] == "ZABBIX_CONTEXT"
        ]

        self.assertEqual(incident_story["signal_assessment"], "TIME_ALIGNED_CONTEXT_ONLY")
        self.assertEqual(incident_story["related_alert_pattern_ids"], [])
        self.assertEqual(
            incident_story["contextual_alert_pattern_ids"],
            [patterns[0]["alert_pattern_id"]],
        )
        self.assertEqual(len(zabbix_milestones), 1)
        self.assertEqual(zabbix_milestones[0]["at"], "2026-05-27T00:48:18")
        self.assertNotIn("2026-04-29", zabbix_milestones[0]["message"])

    def test_operational_stories_render_sequence_conclusion_and_explicit_gaps(self):
        incidents = [
            {
                "incident_id": "INC-3",
                "source_ref": "incident.xlsx#main:row-3",
                "submitted_at": "2026-05-19T16:22:00",
                "started_at": "2026-05-19T16:10:00",
                "resolved_at": "2026-05-19T16:40:00",
                "site_code": "CCW",
                "isp": "FPT",
                "incident_type": "High latency",
                "description": "CCW users are experiencing slow access.",
                "root_cause": "FPT uplink routing issue.",
                "resolution_status": "RESOLVED",
            }
        ]
        alert_patterns = [
            {
                "alert_pattern_id": "ALP-MATCH",
                "source_refs": ["zabbix.xlsx#main:row-2", "zabbix.xlsx#main:row-3"],
                "first_seen_at": "2026-05-19T16:02:00",
                "last_seen_at": "2026-05-19T16:08:00",
                "last_recovered_at": "2026-05-19T16:09:00",
                "site_code": "CCW",
                "host": "VNMCCW-PNWSW01",
                "problem_signature": "high latency",
                "alert_count": 2,
                "resolved_count": 2,
                "open_count": 0,
            }
        ]
        tickets = [
            {
                "ticket_id": "TKT-100",
                "source_refs": ["IssueReport.xlsx#main:row-2"],
                "first_comment_at": "2026-05-19T16:15:00",
                "last_comment_at": "2026-05-19T16:25:00",
                "title": "CCW users cannot access service",
                "comments_summary": "CCW users report slow access.",
            }
        ]
        recurrence = [
            {
                "recurrence_id": "REC-MATCH",
                "incident_ids": ["INC-3", "INC-4"],
                "incident_count": 2,
            }
        ]

        stories = build_operational_stories(incidents, alert_patterns, tickets, recurrence)
        incident = next(story for story in stories if story["episode_id"] == "OPS-INC-3")
        signal = next(story for story in stories if story["episode_id"] == "OPS-ALP-MATCH")

        self.assertEqual(incident["evidence_coverage"], "ZABBIX + INCIDENT FORM + TICKET")
        self.assertEqual(incident["recurrence_summary"], "2 confirmed incidents (REC-MATCH)")
        self.assertIn("Confirmed RCA", " ".join(item["message"] for item in incident["milestones"]))
        self.assertIn("Service recovered", " ".join(item["message"] for item in incident["milestones"]))
        self.assertEqual(signal["signal_assessment"], "RELATED_TO_CONFIRMED_INCIDENT")
        self.assertIn("Zabbix detected", signal["milestones"][0]["message"])
        self.assertIn("Grouped 2 raw alerts", " ".join(item["message"] for item in signal["milestones"]))

    def test_fiber_incident_keeps_unrelated_same_window_zabbix_as_context_only(self):
        incidents = [
            {
                "incident_id": "INC-53",
                "source_ref": "incident.xlsx#main:row-53",
                "submitted_at": "2026-05-27T08:25:02",
                "started_at": "2026-05-27T00:04:00",
                "resolved_at": "2026-05-27T01:00:00",
                "site_code": "SWS",
                "isp": "VNPT",
                "incident_type": "Fiber optic cable failure",
                "description": "Service degradation caused by fiber infrastructure attenuation.",
                "root_cause": "Suy hao cáp hạ tầng.",
                "resolution_status": "RESOLVED",
            }
        ]
        raw_alerts = [
            {
                "alert_id": "ZBX-1",
                "source_ref": "zabbix.csv#row-2",
                "seen_at": "2026-05-27T00:48:18",
                "recovered_at": "2026-05-27T00:54:48",
                "status": "RESOLVED",
                "site_code": "SWS",
                "host": "VNMSWS-PSSJP1",
                "severity": "Warning",
                "problem": "Windows: Host has been restarted (uptime < 10m)",
                "problem_signature": "windows: host has been restarted (uptime < <value>)",
                "domain": "server",
                "component": "windows",
                "scope": "host",
            }
        ]
        alert_patterns = group_alert_patterns(raw_alerts)
        pattern_id = alert_patterns[0]["alert_pattern_id"]

        stories = build_operational_stories(
            incidents, alert_patterns, tickets=[], recurrence=[], alerts=raw_alerts
        )
        incident = next(story for story in stories if story["episode_id"] == "OPS-INC-53")
        signal = next(story for story in stories if story["episode_id"] == f"OPS-{pattern_id}")

        self.assertEqual(incident["signal_assessment"], "TIME_ALIGNED_CONTEXT_ONLY")
        self.assertEqual(incident["evidence_coverage"], "INCIDENT FORM + ZABBIX CONTEXT")
        self.assertEqual(incident["related_alert_pattern_ids"], [])
        self.assertEqual(incident["supporting_alert_pattern_ids"], [])
        self.assertEqual(incident["contextual_alert_pattern_ids"], [pattern_id])
        self.assertIn("does not directly support", incident["zabbix_relevance_basis"])
        context_milestone = next(
            milestone
            for milestone in incident["milestones"]
            if milestone["type"] == "ZABBIX_CONTEXT"
        )
        self.assertEqual(context_milestone["at"], "2026-05-27T00:48:18")
        self.assertEqual(signal["signal_assessment"], "TIME_ALIGNED_CONTEXT_ONLY")
        self.assertEqual(signal["related_incident_ids"], [])
        self.assertEqual(signal["contextual_incident_ids"], ["INC-53"])
        self.assertEqual(signal["supporting_alert_pattern_ids"], [])
        self.assertEqual(signal["contextual_alert_pattern_ids"], [pattern_id])

    def test_fiber_incident_keeps_network_relevant_zabbix_as_supporting_evidence(self):
        incidents = [
            {
                "incident_id": "INC-54",
                "source_ref": "incident.xlsx#main:row-54",
                "submitted_at": "2026-05-28T08:00:00",
                "started_at": "2026-05-28T00:10:00",
                "resolved_at": "2026-05-28T01:00:00",
                "site_code": "MBD",
                "isp": "VNPT",
                "incident_type": "Fiber optic cable failure",
                "description": "MBD users lost connectivity due to fiber cable failure.",
                "root_cause": "Fiber optic cable failure.",
                "resolution_status": "RESOLVED",
            }
        ]
        raw_alerts = [
            {
                "alert_id": "ZBX-2",
                "source_ref": "zabbix.csv#row-3",
                "seen_at": "2026-05-28T00:12:00",
                "recovered_at": "2026-05-28T00:50:00",
                "status": "RESOLVED",
                "site_code": "MBD",
                "host": "VNMMBD-PNWRT01",
                "severity": "High",
                "problem": "Unavailable by ICMP ping",
                "problem_signature": "unavailable by icmp ping",
                "domain": "network",
                "component": "wan",
                "scope": "connectivity",
            }
        ]
        alert_patterns = group_alert_patterns(raw_alerts)
        pattern_id = alert_patterns[0]["alert_pattern_id"]

        stories = build_operational_stories(
            incidents, alert_patterns, tickets=[], recurrence=[], alerts=raw_alerts
        )
        incident = next(story for story in stories if story["episode_id"] == "OPS-INC-54")

        self.assertEqual(incident["signal_assessment"], "RELATED_TO_CONFIRMED_INCIDENT")
        self.assertEqual(incident["evidence_coverage"], "ZABBIX + INCIDENT FORM")
        self.assertEqual(incident["related_alert_pattern_ids"], [pattern_id])
        self.assertEqual(incident["supporting_alert_pattern_ids"], [pattern_id])
        self.assertEqual(incident["contextual_alert_pattern_ids"], [])

    def test_monitoring_story_does_not_claim_user_impact_from_unrelated_same_site_ticket(self):
        stories = build_operational_stories(
            incidents=[],
            alert_patterns=[
                {
                    "alert_pattern_id": "ALP-UPS",
                    "source_refs": ["zabbix.xlsx#main:row-2"],
                    "first_seen_at": "2026-05-19T16:02:00",
                    "last_seen_at": "2026-05-19T16:08:00",
                    "last_recovered_at": "2026-05-19T16:09:00",
                    "site_code": "SNT",
                    "host": "VNMSNT-UPSF2201",
                    "problem_signature": "ups input frequency out of range",
                    "domain": "power",
                    "component": "power",
                    "alert_count": 2,
                    "resolved_count": 2,
                    "open_count": 0,
                }
            ],
            tickets=[
                {
                    "ticket_id": "TKT-NETWORK",
                    "source_refs": ["IssueReport.xlsx#main:row-2"],
                    "first_comment_at": "2026-05-19T16:03:00",
                    "last_comment_at": "2026-05-19T16:04:00",
                    "title": "SNT network segmentation review",
                    "comments_summary": "Review wireless security policy.",
                }
            ],
            recurrence=[],
        )

        signal = stories[0]
        self.assertEqual(signal["user_impact_status"], "UNKNOWN")
        self.assertEqual(signal["related_ticket_ids"], [])
        self.assertEqual(signal["signal_assessment"], "LIKELY_NOISE_OR_THRESHOLD_REVIEW")

    def test_validation_report_fails_when_reconciliation_is_wrong(self):
        report = build_validation_report(
            source_profile={
                "incident_form_rows": 4,
                "normalized_incidents": 2,
                "rejected_incident_rows": 1,
                "zabbix_rows": 0,
                "normalized_alert_rows": 0,
                "rejected_alert_rows": 0,
                "issue_comment_rows": 0,
                "aggregated_ticket_comment_count": 0,
                "rejected_comment_rows": 0,
            },
            warnings=[],
            known_checks=[],
        )
        self.assertEqual(report["status"], "FAIL")
        self.assertFalse(report["invariants"][0]["passed"])

    def test_package_contains_alpha_documents_audit_workbook_and_pass_gate(self):
        with tempfile.TemporaryDirectory() as temporary:
            output_dir = Path(temporary) / "package"
            package = build_knowledge_package(
                raw_dir=FIXTURES,
                master_data_path=FIXTURES / "master_data.xlsx",
                output_dir=output_dir,
                generated_at="2026-06-02T00:00:00Z",
            )
            required = {
                "manifest.json",
                "operational_timeline.json",
                "validation_report.json",
                "validation_report.md",
                "00_report_context.md",
                "01_executive_summary.md",
                "02_operational_timeline.md",
                "02_confirmed_incidents.md",
                "03_recurrence_patterns.md",
                "04_alert_patterns.md",
                "05_ticket_impact.md",
                "05_ticket_impact_2026_05.md",
                "05_ticket_impact_index.md",
                "06_data_quality.md",
                "normalized_data.xlsx",
            }
            self.assertEqual({path.name for path in output_dir.iterdir()}, required)
            report = json.loads((output_dir / "validation_report.json").read_text("utf-8"))
            self.assertEqual(report["status"], "PASS")
            self.assertTrue(package["upload_allowed"])
            manifest = json.loads((output_dir / "manifest.json").read_text("utf-8"))
            self.assertIn(
                "master_data.xlsx",
                {item["filename"] for item in manifest["inputs"]},
            )
            timeline_payload = json.loads(
                (output_dir / "operational_timeline.json").read_text("utf-8")
            )
            self.assertEqual(timeline_payload["record_type"], "OPERATIONAL_TIMELINE_COLLECTION")
            self.assertEqual(
                timeline_payload["record_count"],
                package["source_profile"]["operational_timeline_events"],
            )
            workbook = load_workbook(output_dir / "normalized_data.xlsx", read_only=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "operational_timeline",
                    "confirmed_incidents",
                    "alert_patterns",
                    "ticket_evidence",
                    "recurrence_patterns",
                    "data_quality",
                    "source_profile",
                ],
            )
            quality_rows = list(workbook["data_quality"].iter_rows(values_only=True))
            quality_text = "\n".join(
                " | ".join(str(value) for value in row if value is not None)
                for row in quality_rows
            )
            workbook.close()
            timeline_text = (output_dir / "02_operational_timeline.md").read_text("utf-8")
            self.assertIn("# Operational Timeline", timeline_text)
            self.assertIn("record_type: OPERATIONAL_TIMELINE_EVENT", timeline_text)
            self.assertIn("site_code:", timeline_text)
            self.assertIn("CONFIRMED_INCIDENT", timeline_text)
            self.assertIn("MONITORING_SIGNAL", timeline_text)
            self.assertIn("### Timeline", timeline_text)
            self.assertIn("### Conclusion", timeline_text)
            self.assertIn("Evidence coverage", timeline_text)
            self.assertIn("Investigation gaps", timeline_text)
            alert_patterns_text = (output_dir / "04_alert_patterns.md").read_text("utf-8")
            self.assertIn("Site Pattern Family Summary", alert_patterns_text)
            self.assertIn("Individual Alert Patterns", alert_patterns_text)
            ticket_index_text = (output_dir / "05_ticket_impact_index.md").read_text("utf-8")
            self.assertIn("Ticket Impact Partition Index", ticket_index_text)
            self.assertIn("partition_basis: comment_activity_month", ticket_index_text)
            self.assertIn("05_ticket_impact_2026_05.md", ticket_index_text)
            self.assertIn("Top symptom families", ticket_index_text)
            self.assertIn("Top resolution signals", ticket_index_text)
            monthly_ticket_text = (output_dir / "05_ticket_impact_2026_05.md").read_text("utf-8")
            self.assertIn("Ticket Impact Evidence - 2026-05", monthly_ticket_text)
            self.assertIn("Evidence time basis", monthly_ticket_text)
            self.assertIn("Symptom family", monthly_ticket_text)
            self.assertIn("Resolution signal", monthly_ticket_text)
            self.assertIn("Initial comment sample", monthly_ticket_text)
            self.assertIn("Final comment sample", monthly_ticket_text)
            self.assertIn("Comment summary strategy", monthly_ticket_text)
            self.assertIn("Comment summary: See lifecycle samples below.", monthly_ticket_text)
            self.assertIn("Incidents missing RCA", quality_text)
            self.assertIn("Sites without confirmed master-data mapping", quality_text)

    def test_explicit_missing_master_data_fails_clearly(self):
        with tempfile.TemporaryDirectory() as temporary:
            with self.assertRaisesRegex(FileNotFoundError, "master-data"):
                build_knowledge_package(
                    raw_dir=FIXTURES,
                    master_data_path=Path(temporary) / "missing-master-data.xlsx",
                    output_dir=Path(temporary) / "package",
                    generated_at="2026-06-02T00:00:00Z",
                )

    def test_package_is_deterministic_for_same_inputs_and_generated_at(self):
        with tempfile.TemporaryDirectory() as temporary:
            first = Path(temporary) / "first"
            second = Path(temporary) / "second"
            for output_dir in [first, second]:
                build_knowledge_package(
                    raw_dir=FIXTURES,
                    master_data_path=FIXTURES / "master_data.xlsx",
                    output_dir=output_dir,
                    generated_at="2026-06-02T00:00:00Z",
                )
            for filename in [
                "manifest.json",
                "operational_timeline.json",
                "validation_report.json",
                "validation_report.md",
                "00_report_context.md",
                "01_executive_summary.md",
                "02_operational_timeline.md",
                "02_confirmed_incidents.md",
                "03_recurrence_patterns.md",
                "04_alert_patterns.md",
                "05_ticket_impact.md",
                "05_ticket_impact_2026_05.md",
                "05_ticket_impact_index.md",
                "06_data_quality.md",
            ]:
                self.assertEqual(
                    (first / filename).read_bytes(),
                    (second / filename).read_bytes(),
                    filename,
                )


class SuppliedRawDataIntegrationTest(unittest.TestCase):
    def test_supplied_raw_data_matches_known_counts_and_patterns(self):
        raw_dir = ROOT / "01-RawData"
        incidents = load_incidents(
            raw_dir
            / "ISP Incident Report"
            / "SEA - Corp IT- ILL- Incident Report   (Responses).xlsx"
        )
        alerts = load_zabbix_alerts(raw_dir / "Export zabbix")
        tickets = load_issue_tickets(raw_dir / "Ticket")
        recurrence = group_incident_recurrence(incidents.records)
        alert_patterns = group_alert_patterns(alerts.records)
        timeline = build_operational_timeline(
            incidents.records,
            alert_patterns,
            tickets.records,
            alerts=alerts.records,
        )

        self.assertEqual(len(incidents.records), 52)
        self.assertEqual(len(alerts.records), 5199)
        self.assertEqual(len(tickets.records), 2681)
        self.assertEqual(sum(ticket["comment_count"] for ticket in tickets.records), 9525)
        self.assertEqual(
            sorted({month for ticket in tickets.records for month in ticket["partition_months"]}),
            ["2026-04", "2026-05", "2026-06"],
        )

        ms2 = next(
            pattern
            for pattern in recurrence
            if pattern["site_code"] == "MS2"
            and pattern["isp"] == "VNPT"
            and pattern["incident_type"] == "High latency"
        )
        xas = next(
            pattern
            for pattern in recurrence
            if pattern["site_code"] == "XAS"
            and pattern["isp"] == "CMC"
            and pattern["incident_type"] == "Fiber optic cable failure"
        )
        self.assertEqual(ms2["incident_count"], 16)
        self.assertEqual(xas["incident_count"], 4)
        self.assertEqual(len(alert_patterns), 365)
        self.assertEqual(len(timeline), 417)
        self.assertEqual(
            sum(event["event_type"] == "CONFIRMED_INCIDENT" for event in timeline),
            52,
        )
        self.assertEqual(
            sum(event["event_type"] == "ZABBIX_ALERT_PATTERN" for event in timeline),
            365,
        )
        may_incidents = {
            event["event_id"]: event["related_alert_pattern_ids"]
            for event in timeline
            if event["event_type"] == "CONFIRMED_INCIDENT"
            and event["started_at"].startswith("2026-05")
        }
        self.assertEqual(
            may_incidents["EVT-INC-51"],
            [],
            "MDN long-running alert patterns should not match without event overlap.",
        )
        self.assertEqual(may_incidents["EVT-INC-53"], ["ALP-21A98CCEA4A6"])
        self.assertEqual(
            may_incidents["EVT-INC-54"],
            [
                "ALP-031BE4B5C2DB",
                "ALP-6E01E457377C",
                "ALP-C031A95EC809",
                "ALP-CDCB1323E9CB",
            ],
        )


if __name__ == "__main__":
    unittest.main()
