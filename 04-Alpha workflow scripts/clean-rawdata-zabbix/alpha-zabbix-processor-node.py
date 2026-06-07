import csv
import hashlib
import io
import json
import re
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from dateutil import parser as dtparser
from openpyxl import load_workbook


ZABBIX_REQUIRED_COLUMNS = [
    "Severity",
    "Time",
    "Recovery time",
    "Status",
    "Host",
    "Problem",
    "Duration",
    "Ack",
    "Actions",
    "Tags",
]

ZABBIX_NORMALIZED_REQUIRED_COLUMNS = [
    "event_id",
    "started_at_local",
    "recovered_at_local",
    "status",
    "severity",
    "host",
    "site_code",
    "problem_raw",
    "problem_signature",
    "duration_text",
    "acknowledged",
    "action_count",
    "domain",
    "component",
    "scope",
    "tags_json",
]

MAX_SOURCE_REFS_PER_PATTERN = 20
DEFAULT_MAX_OUTPUT_RECORDS = 0
DEFAULT_MAX_OUTPUT_PATTERNS = 500


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


RAW_HEADER_ALIASES = {
    _normalize_header("Severity"): "Severity",
    _normalize_header("Time"): "Time",
    _normalize_header("Recovery time"): "Recovery time",
    _normalize_header("Status"): "Status",
    _normalize_header("Host"): "Host",
    _normalize_header("Problem"): "Problem",
    _normalize_header("Duration"): "Duration",
    _normalize_header("Ack"): "Ack",
    _normalize_header("Actions"): "Actions",
    _normalize_header("Tags"): "Tags",
}

NORMALIZED_HEADER_ALIASES = {
    _normalize_header(name): name for name in ZABBIX_NORMALIZED_REQUIRED_COLUMNS
}


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return str(value)
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()


def _unknown_if_blank(value: Any) -> str:
    text = _clean_text(value)
    return text if text else "UNKNOWN"


def _stable_hash(*parts: Any, length: int = 12) -> str:
    payload = "\x1f".join(_clean_text(part).lower() for part in parts)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:length].upper()


def _parse_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = _clean_text(value)
    if not text:
        raise ValueError("empty datetime value")
    try:
        return dtparser.parse(text, dayfirst=False, fuzzy=True).replace(tzinfo=None)
    except Exception as exc:
        raise ValueError(f"invalid datetime value: {text}") from exc


def _iso_datetime(value: Optional[datetime]) -> str:
    return "" if value is None else value.isoformat(timespec="seconds")


def _derive_site_code_from_host(host: Any) -> str:
    match = re.match(r"^VNM([A-Z0-9]{3})", _clean_text(host).upper())
    return match.group(1) if match else "UNKNOWN"


def _looks_like_html(s: str) -> bool:
    text = s.lstrip().lower()
    return (
        text.startswith("<!doctype html")
        or text.startswith("<html")
        or ("<head" in text and "<body" in text)
        or "accounts.google.com" in text
    )


def _is_blank_row(values: Iterable[Any]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        return False
    return True


def _resolve_file_path(file_obj: Any) -> str:
    if isinstance(file_obj, (str, Path)):
        return str(file_obj)
    if isinstance(file_obj, dict):
        for key in ("path", "file_path", "filepath", "file", "filename", "file_name", "name"):
            if key in file_obj and file_obj[key]:
                return str(file_obj[key])
        if file_obj.get("id"):
            return str(file_obj["id"])
    return str(file_obj)


def _parse_tags(value: Any) -> Dict[str, List[str]]:
    tags: Dict[str, List[str]] = defaultdict(list)
    for token in _clean_text(value).split(", "):
        if ": " not in token:
            continue
        key, tag_value = token.split(": ", 1)
        if tag_value not in tags[key]:
            tags[key].append(tag_value)
    return dict(tags)


def _parse_tags_json(value: Any) -> Dict[str, List[str]]:
    text = _clean_text(value)
    if not text:
        return {}
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        return _parse_tags(text)
    if not isinstance(payload, dict):
        return {}
    tags: Dict[str, List[str]] = {}
    for key, raw_values in payload.items():
        if isinstance(raw_values, list):
            values = [_clean_text(item) for item in raw_values if _clean_text(item)]
        else:
            values = [_clean_text(raw_values)] if _clean_text(raw_values) else []
        if values:
            tags[_clean_text(key)] = values
    return tags


def _normalize_problem_signature(problem: Any) -> str:
    text = _clean_text(problem).lower()
    text = re.sub(
        r"current:\s*\d+(?:\.\d+)?\s*mbps",
        "current:<value> mbps",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(
        r"\d+(?:\.\d+)?(?:-\d+(?:\.\d+)?)?\s*(?:mbps|gbps|%|hz|m|s|c|°c|℃)\b",
        "<value>",
        text,
        flags=re.IGNORECASE,
    )
    return text


def _join_unique(values: Iterable[str]) -> str:
    unique = sorted({value for value in values if value and value != "UNKNOWN"})
    return ",".join(unique) if unique else "UNKNOWN"


def _classify_alert_pattern_family(pattern: Dict[str, Any]) -> str:
    signature = _clean_text(pattern.get("problem_signature")).lower()
    host = _clean_text(pattern.get("host")).lower()
    component = _clean_text(pattern.get("component")).lower()
    domain = _clean_text(pattern.get("domain")).lower()
    joined = " ".join([signature, host, component, domain])
    if "http monitoring" in signature:
        return "HTTP monitoring"
    if any(
        term in joined
        for term in [
            "icmp",
            "unavailable",
            "unreachable",
            "interface down",
            "link down",
            "port down",
            "packet loss",
            "latency",
            " is down",
            "down!",
            "gateway",
            "tunnel",
            "vpn",
            "bgp",
        ]
    ):
        return "Network reachability"
    if any(term in joined for term in ["active checks", "zabbix agent", "nodata"]):
        return "Zabbix agent/active check"
    if any(term in joined for term in ["restarted", "uptime"]):
        return "Host restart/uptime"
    if any(term in joined for term in ["ups", "battery", "frequency", "power"]):
        return "Power/UPS"
    if any(term in joined for term in ["cpu", "memory", "disk", "filesystem", "space"]):
        return "Resource/capacity"
    return "Other"


def _alert_investigation_priority(pattern: Dict[str, Any], family: str) -> str:
    signature = _clean_text(pattern.get("problem_signature")).lower()
    host = _clean_text(pattern.get("host")).lower()
    alert_count = int(pattern.get("alert_count") or 0)
    if family == "Network reachability":
        if alert_count > 1 or "fw" in host or "sw" in host or "down" in signature:
            return "HIGH"
        return "MEDIUM"
    if family == "HTTP monitoring":
        return "MEDIUM" if alert_count >= 10 else "LOW"
    if family == "Power/UPS":
        return "HIGH"
    if family in {"Zabbix agent/active check", "Host restart/uptime", "Resource/capacity"}:
        return "LOW"
    return "LOW"


def _alert_why_it_matters(pattern: Dict[str, Any], family: str) -> str:
    host = _clean_text(pattern.get("host"))
    signature = _clean_text(pattern.get("problem_signature"))
    if family == "Network reachability":
        return (
            f"{host} reported `{signature}`. This can indicate connectivity loss, "
            "device/path flap, ISP path issue, or monitoring-path instability."
        )
    if family == "HTTP monitoring":
        return (
            f"{host} reported `{signature}`. This usually means service probe timeout "
            "or threshold sensitivity unless supported by ticket/incident evidence."
        )
    if family == "Zabbix agent/active check":
        return (
            f"{host} reported `{signature}`. This points to monitoring-agent health, "
            "not confirmed user impact by itself."
        )
    if family == "Host restart/uptime":
        return (
            f"{host} reported `{signature}`. This is host-health context and should "
            "not be treated as service impact without another source."
        )
    if family == "Power/UPS":
        return (
            f"{host} reported `{signature}`. This may indicate power/UPS instability "
            "and deserves device/facility verification."
        )
    if family == "Resource/capacity":
        return (
            f"{host} reported `{signature}`. This is resource/capacity context; "
            "treat as potential technical signal unless corroborated."
        )
    return f"{host} reported `{signature}`. Treat as monitoring context until corroborated."


def _signal_assessment(pattern: Dict[str, Any], family: str) -> str:
    if family == "HTTP monitoring" and int(pattern.get("alert_count") or 0) > 1:
        return "LIKELY_NOISE_OR_THRESHOLD_REVIEW"
    return "UNCONFIRMED_MONITORING_SIGNAL"


def _build_error(reason: str, source_type: str, rejected: Optional[List[Dict[str, Any]]] = None):
    json_result = {
        "ok": False,
        "source_type": source_type,
        "input_rows": 0,
        "record_count": 0,
        "rejected_row_count": 0,
        "duplicate_rows": 0,
        "skipped_blank_rows": 0,
        "records": [],
        "rejected_rows": rejected
        if rejected is not None
        else [{"row_number": None, "source_ref": "input", "reason": reason, "row_data": {}}],
    }
    summary = {
        "ok": False,
        "source_type": source_type,
        "input_rows": 0,
        "record_count": 0,
        "alert_pattern_count": 0,
        "rejected_row_count": len(json_result["rejected_rows"]),
        "duplicate_rows": 0,
        "skipped_blank_rows": 0,
        "reconciliation_passed": False,
        "upload_ready": False,
    }
    return {
        "json_text": json.dumps(json_result, ensure_ascii=False, indent=2),
        "alert_patterns_text": json.dumps(
            {"ok": False, "records": [], "record_count": 0}, ensure_ascii=False, indent=2
        ),
        "json_result": json_result,
        "alert_patterns_result": {"ok": False, "records": [], "record_count": 0},
        "markdown": "",
        "summary": summary,
    }


def _canonical_headers(header_row: Sequence[Any]) -> Tuple[Dict[str, int], bool, List[str]]:
    raw_map: Dict[str, int] = {}
    normalized_map: Dict[str, int] = {}
    headers = tuple(header_row)

    for idx, header in enumerate(headers):
        norm = _normalize_header(header)
        if norm in RAW_HEADER_ALIASES:
            raw_map[RAW_HEADER_ALIASES[norm]] = idx
        if norm in NORMALIZED_HEADER_ALIASES:
            normalized_map[NORMALIZED_HEADER_ALIASES[norm]] = idx

    is_normalized = all(name in normalized_map for name in ZABBIX_NORMALIZED_REQUIRED_COLUMNS)
    required = ZABBIX_NORMALIZED_REQUIRED_COLUMNS if is_normalized else ZABBIX_REQUIRED_COLUMNS
    mapping = normalized_map if is_normalized else raw_map
    missing = [name for name in required if name not in mapping]
    return mapping, is_normalized, missing


def _cell(row: Sequence[Any], mapping: Dict[str, int], name: str) -> Any:
    idx = mapping[name]
    return row[idx] if idx < len(row) else None


def _process_rows(
    *,
    rows: Iterable[Sequence[Any]],
    source_type: str,
    source_label: str,
) -> Dict[str, Any]:
    records: List[Dict[str, Any]] = []
    rejected_rows: List[Dict[str, Any]] = []
    input_rows = 0
    duplicate_rows = 0
    skipped_blank_rows = 0
    seen_rows = set()

    rows_iter = iter(rows)
    header_row_idx = None
    header_row = None
    for idx, row in enumerate(rows_iter, start=1):
        if _is_blank_row(row):
            continue
        header_row_idx = idx
        header_row = tuple(row)
        break

    if header_row_idx is None or header_row is None:
        return _build_error("No header row found", source_type)

    mapping, is_normalized, missing = _canonical_headers(header_row)
    if missing:
        return _build_error(
            "Missing required columns",
            source_type,
            rejected=[
                {
                    "row_number": header_row_idx,
                    "source_ref": f"{source_label}:row-{header_row_idx}",
                    "reason": "Missing required columns: " + ", ".join(missing),
                    "row_data": {"headers": [str(item) if item is not None else "" for item in header_row]},
                }
            ],
        )

    for row_number, row in enumerate(rows_iter, start=header_row_idx + 1):
        row_tuple = tuple(row)
        if _is_blank_row(row_tuple):
            skipped_blank_rows += 1
            continue

        input_rows += 1
        source_ref = f"{source_label}:row-{row_number}"

        if is_normalized:
            event_id = _clean_text(_cell(row_tuple, mapping, "event_id"))
            dedupe_key = ("normalized", event_id)
        else:
            dedupe_key = tuple(_clean_text(_cell(row_tuple, mapping, col)) for col in ZABBIX_REQUIRED_COLUMNS)

        if dedupe_key in seen_rows:
            duplicate_rows += 1
            continue
        seen_rows.add(dedupe_key)

        try:
            if is_normalized:
                seen_at = _parse_datetime(_cell(row_tuple, mapping, "started_at_local"))
                recovered_raw = _cell(row_tuple, mapping, "recovered_at_local")
                recovered_at = _parse_datetime(recovered_raw) if _clean_text(recovered_raw) else None
                host = _unknown_if_blank(_cell(row_tuple, mapping, "host"))
                problem = _unknown_if_blank(_cell(row_tuple, mapping, "problem_raw"))
                tags = _parse_tags_json(_cell(row_tuple, mapping, "tags_json"))
                action_count = _clean_text(_cell(row_tuple, mapping, "action_count"))
                alert_id = f"ALT-ZBX-{event_id}" if event_id else f"ALT-{_stable_hash(source_label, row_number)}"
                record = {
                    "alert_id": alert_id,
                    "source_ref": source_ref,
                    "seen_at": _iso_datetime(seen_at),
                    "recovered_at": _iso_datetime(recovered_at),
                    "status": _unknown_if_blank(_cell(row_tuple, mapping, "status")).upper(),
                    "severity": _unknown_if_blank(_cell(row_tuple, mapping, "severity")),
                    "host": host,
                    "site_code": _clean_text(_cell(row_tuple, mapping, "site_code")) or _derive_site_code_from_host(host),
                    "problem": problem,
                    "problem_signature": _clean_text(_cell(row_tuple, mapping, "problem_signature"))
                    or _normalize_problem_signature(problem),
                    "duration": _unknown_if_blank(_cell(row_tuple, mapping, "duration_text")),
                    "ack": "Yes"
                    if _clean_text(_cell(row_tuple, mapping, "acknowledged")).lower() == "true"
                    else "No",
                    "actions": f"Actions ({action_count})" if action_count and action_count != "0" else "UNKNOWN",
                    "domain": _clean_text(_cell(row_tuple, mapping, "domain")) or ",".join(tags.get("class", [])) or "UNKNOWN",
                    "component": _clean_text(_cell(row_tuple, mapping, "component"))
                    or ",".join(tags.get("component", []))
                    or "UNKNOWN",
                    "scope": _clean_text(_cell(row_tuple, mapping, "scope")) or ",".join(tags.get("scope", [])) or "UNKNOWN",
                    "tags": json.dumps(tags, ensure_ascii=False, sort_keys=True),
                    "evidence_label": "SOURCE FACT",
                }
            else:
                seen_at = _parse_datetime(_cell(row_tuple, mapping, "Time"))
                recovered_raw = _cell(row_tuple, mapping, "Recovery time")
                recovered_at = _parse_datetime(recovered_raw) if _clean_text(recovered_raw) else None
                host = _unknown_if_blank(_cell(row_tuple, mapping, "Host"))
                problem = _unknown_if_blank(_cell(row_tuple, mapping, "Problem"))
                tags = _parse_tags(_cell(row_tuple, mapping, "Tags"))
                record = {
                    "alert_id": f"ALT-{_stable_hash(source_label, row_number)}",
                    "source_ref": source_ref,
                    "seen_at": _iso_datetime(seen_at),
                    "recovered_at": _iso_datetime(recovered_at),
                    "status": _unknown_if_blank(_cell(row_tuple, mapping, "Status")).upper(),
                    "severity": _unknown_if_blank(_cell(row_tuple, mapping, "Severity")),
                    "host": host,
                    "site_code": _derive_site_code_from_host(host),
                    "problem": problem,
                    "problem_signature": _normalize_problem_signature(problem),
                    "duration": _unknown_if_blank(_cell(row_tuple, mapping, "Duration")),
                    "ack": _unknown_if_blank(_cell(row_tuple, mapping, "Ack")),
                    "actions": _unknown_if_blank(_cell(row_tuple, mapping, "Actions")),
                    "domain": ",".join(tags.get("class", [])) or "UNKNOWN",
                    "component": ",".join(tags.get("component", [])) or "UNKNOWN",
                    "scope": ",".join(tags.get("scope", [])) or "UNKNOWN",
                    "tags": json.dumps(tags, ensure_ascii=False, sort_keys=True),
                    "evidence_label": "SOURCE FACT",
                }
        except Exception as exc:
            rejected_rows.append(
                {
                    "row_number": row_number,
                    "source_ref": source_ref,
                    "reason": str(exc),
                    "row_data": [str(item) if item is not None else "" for item in row_tuple[:20]],
                }
            )
            continue

        records.append(record)

    patterns = _group_alert_patterns(records)

    json_result = {
        "ok": True,
        "source_type": source_type,
        "input_rows": input_rows,
        "record_count": len(records),
        "rejected_row_count": len(rejected_rows),
        "duplicate_rows": duplicate_rows,
        "skipped_blank_rows": skipped_blank_rows,
        "is_normalized_input": is_normalized,
        "records": records,
        "rejected_rows": rejected_rows,
    }
    patterns_result = {
        "ok": True,
        "source_type": "zabbix_alert_patterns",
        "record_type": "ALERT_PATTERN_COLLECTION",
        "record_count": len(patterns),
        "records": patterns,
    }
    summary = {
        "ok": True,
        "source_type": source_type,
        "input_rows": input_rows,
        "record_count": len(records),
        "alert_pattern_count": len(patterns),
        "rejected_row_count": len(rejected_rows),
        "duplicate_rows": duplicate_rows,
        "skipped_blank_rows": skipped_blank_rows,
        "reconciliation_passed": len(records) + len(rejected_rows) + duplicate_rows == input_rows,
        "upload_ready": len(rejected_rows) == 0 and len(records) > 0 and len(patterns) > 0,
        "is_normalized_input": is_normalized,
    }
    return {
        "json_text": json.dumps(json_result, ensure_ascii=False, indent=2),
        "alert_patterns_text": json.dumps(patterns_result, ensure_ascii=False, indent=2),
        "json_result": json_result,
        "alert_patterns_result": patterns_result,
        "markdown": _render_patterns_markdown(patterns, summary),
        "summary": summary,
    }


def _group_alert_patterns(alerts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for alert in alerts:
        grouped[(alert["host"], alert["problem_signature"])].append(alert)

    patterns: List[Dict[str, Any]] = []
    for (host, signature), records in grouped.items():
        records = sorted(records, key=lambda record: (record["seen_at"], record["source_ref"]))
        source_refs = [record["source_ref"] for record in records]
        pattern = {
            "record_type": "ZABBIX_ALERT_PATTERN",
            "alert_pattern_id": f"ALP-{_stable_hash(host, signature)}",
            "source_refs": source_refs[:MAX_SOURCE_REFS_PER_PATTERN],
            "first_seen_at": records[0]["seen_at"],
            "last_seen_at": records[-1]["seen_at"],
            "last_recovered_at": max(
                (record["recovered_at"] for record in records if record["recovered_at"]),
                default="",
            ),
            "alert_count": len(records),
            "resolved_count": sum(record["status"] == "RESOLVED" for record in records),
            "open_count": sum(record["status"] != "RESOLVED" for record in records),
            "host": host,
            "site_code": records[0]["site_code"],
            "severity": _join_unique(record["severity"] for record in records),
            "problem_signature": signature,
            "domain": _join_unique(record["domain"] for record in records),
            "component": _join_unique(record["component"] for record in records),
            "scope": _join_unique(record["scope"] for record in records),
            "assessment": "RAW ALERT PATTERN - NOT A CONFIRMED INCIDENT",
            "evidence_label": "COMPUTED FACT",
            "source_reference_count": len(source_refs),
        }
        pattern["pattern_family"] = _classify_alert_pattern_family(pattern)
        pattern["investigation_priority"] = _alert_investigation_priority(
            pattern, pattern["pattern_family"]
        )
        pattern["why_it_matters"] = _alert_why_it_matters(pattern, pattern["pattern_family"])
        pattern["signal_assessment"] = _signal_assessment(pattern, pattern["pattern_family"])
        patterns.append(pattern)
    return sorted(patterns, key=lambda pattern: (-pattern["alert_count"], pattern["alert_pattern_id"]))


def _render_patterns_markdown(patterns: List[Dict[str, Any]], summary: Dict[str, Any]) -> str:
    lines = [
        "# Zabbix Alert Patterns",
        "",
        "```yaml",
        "record_type: ZABBIX_ALERT_PATTERN_SOURCE",
        f"source_type: {summary['source_type']}",
        f"input_rows: {summary['input_rows']}",
        f"record_count: {summary['record_count']}",
        f"alert_pattern_count: {summary['alert_pattern_count']}",
        f"rejected_row_count: {summary['rejected_row_count']}",
        f"duplicate_rows: {summary['duplicate_rows']}",
        f"skipped_blank_rows: {summary['skipped_blank_rows']}",
        f"reconciliation_passed: {str(summary['reconciliation_passed']).lower()}",
        f"upload_ready: {str(summary['upload_ready']).lower()}",
        "validation_status: PASS" if summary["upload_ready"] else "validation_status: FAIL",
        "note: Raw Zabbix alert patterns are monitoring signals, not confirmed incidents.",
        "```",
        "",
    ]
    for pattern in patterns:
        lines.extend(
            [
                f"## Alert Pattern {pattern['alert_pattern_id']}",
                "",
                "```yaml",
                f"record_type: {pattern['record_type']}",
                f"alert_pattern_id: {pattern['alert_pattern_id']}",
                f"host: {pattern['host']}",
                f"site_code: {pattern['site_code']}",
                f"first_seen_at: {pattern['first_seen_at']}",
                f"last_seen_at: {pattern['last_seen_at']}",
                f"last_recovered_at: {pattern['last_recovered_at'] or 'UNKNOWN'}",
                f"alert_count: {pattern['alert_count']}",
                f"resolved_count: {pattern['resolved_count']}",
                f"open_count: {pattern['open_count']}",
                f"severity: {pattern['severity']}",
                f"pattern_family: {pattern['pattern_family']}",
                f"investigation_priority: {pattern['investigation_priority']}",
                f"signal_assessment: {pattern['signal_assessment']}",
                f"domain: {pattern['domain']}",
                f"component: {pattern['component']}",
                f"scope: {pattern['scope']}",
                f"assessment: {pattern['assessment']}",
                f"evidence_label: {pattern['evidence_label']}",
                f"source_reference_count: {pattern['source_reference_count']}",
                "problem_signature: |",
                "  " + pattern["problem_signature"].replace("\n", "\n  "),
                "why_it_matters: |",
                "  " + pattern["why_it_matters"].replace("\n", "\n  "),
                "source_refs_sample: |",
                "  " + "; ".join(pattern["source_refs"][:20]),
                "```",
                "",
            ]
        )
    return "\n".join(lines).strip() + "\n"


def _rows_from_xlsx(path: Path):
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        ws = wb.active
        return list(ws.iter_rows(values_only=True)), f"{path.name}:{ws.title}"
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _suffix_from_hint(path: Path, source_label: Optional[str]) -> str:
    if source_label:
        suffix = Path(_clean_text(source_label)).suffix.lower()
        if suffix:
            return suffix
    return path.suffix.lower()


def _path_looks_like_csv(path: Path) -> bool:
    try:
        raw = path.read_bytes()[:4096]
    except Exception:
        return False
    if not raw:
        return False
    # XLSX files are zip archives. Do not misread them as text CSV.
    if raw.startswith(b"PK\x03\x04"):
        return False

    sample = raw.decode("utf-8-sig", errors="replace")
    if _looks_like_html(sample):
        return True
    first_line = sample.splitlines()[0] if sample.splitlines() else sample
    if "," not in first_line:
        return False

    try:
        header_cells = next(csv.reader(io.StringIO(first_line)))
    except Exception:
        return False

    normalized_headers = {_normalize_header(cell) for cell in header_cells}
    known_headers = set(RAW_HEADER_ALIASES.keys()) | set(NORMALIZED_HEADER_ALIASES.keys())
    return len(normalized_headers & known_headers) >= 3


def _as_list(value: Any) -> List[Any]:
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return list(value)
    return [value]


def _label_at(labels: List[Any], index: int, default: str) -> str:
    if index < len(labels):
        label = _clean_text(labels[index])
        if label:
            return label
    return default


def _record_dedupe_key(record: Dict[str, Any]) -> Tuple[str, ...]:
    alert_id = _clean_text(record.get("alert_id"))
    if alert_id:
        return ("alert_id", alert_id)
    return (
        "fallback",
        _clean_text(record.get("source_ref")),
        _clean_text(record.get("seen_at")),
        _clean_text(record.get("host")),
        _clean_text(record.get("problem_signature")),
    )


def _merge_processed_results(processed_results: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not processed_results:
        return _build_error("No Zabbix sources were provided", "zabbix_multi_source")
    if len(processed_results) == 1:
        return processed_results[0]

    records: List[Dict[str, Any]] = []
    rejected_rows: List[Dict[str, Any]] = []
    seen_record_keys = set()
    input_rows = 0
    duplicate_rows = 0
    skipped_blank_rows = 0
    normalized_flags: List[bool] = []

    for result in processed_results:
        json_result = result.get("json_result", {})
        summary = result.get("summary", {})
        input_rows += int(json_result.get("input_rows") or summary.get("input_rows") or 0)
        duplicate_rows += int(
            json_result.get("duplicate_rows") or summary.get("duplicate_rows") or 0
        )
        skipped_blank_rows += int(
            json_result.get("skipped_blank_rows") or summary.get("skipped_blank_rows") or 0
        )
        if json_result.get("ok") is True:
            normalized_flags.append(bool(json_result.get("is_normalized_input")))
        rejected_rows.extend(json_result.get("rejected_rows") or [])

        for record in json_result.get("records") or []:
            dedupe_key = _record_dedupe_key(record)
            if dedupe_key in seen_record_keys:
                duplicate_rows += 1
                continue
            seen_record_keys.add(dedupe_key)
            records.append(record)

    patterns = _group_alert_patterns(records)
    source_type = "zabbix_multi_source"
    is_normalized_input = all(normalized_flags) if normalized_flags else False
    reconciliation_passed = len(records) + len(rejected_rows) + duplicate_rows == input_rows

    json_result = {
        "ok": True,
        "source_type": source_type,
        "input_rows": input_rows,
        "record_count": len(records),
        "rejected_row_count": len(rejected_rows),
        "duplicate_rows": duplicate_rows,
        "skipped_blank_rows": skipped_blank_rows,
        "is_normalized_input": is_normalized_input,
        "source_file_count": len(processed_results),
        "records": records,
        "rejected_rows": rejected_rows,
    }
    patterns_result = {
        "ok": True,
        "source_type": "zabbix_alert_patterns",
        "record_type": "ALERT_PATTERN_COLLECTION",
        "record_count": len(patterns),
        "records": patterns,
    }
    summary = {
        "ok": True,
        "source_type": source_type,
        "input_rows": input_rows,
        "record_count": len(records),
        "alert_pattern_count": len(patterns),
        "rejected_row_count": len(rejected_rows),
        "duplicate_rows": duplicate_rows,
        "skipped_blank_rows": skipped_blank_rows,
        "source_file_count": len(processed_results),
        "reconciliation_passed": reconciliation_passed,
        "upload_ready": len(rejected_rows) == 0 and len(records) > 0 and len(patterns) > 0,
        "is_normalized_input": is_normalized_input,
    }
    return {
        "json_text": json.dumps(json_result, ensure_ascii=False, indent=2),
        "alert_patterns_text": json.dumps(patterns_result, ensure_ascii=False, indent=2),
        "json_result": json_result,
        "alert_patterns_result": patterns_result,
        "markdown": _render_patterns_markdown(patterns, summary),
        "summary": summary,
    }


def _to_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = _clean_text(value).lower()
    if text in ("1", "true", "yes", "y", "on"):
        return True
    if text in ("0", "false", "no", "n", "off"):
        return False
    return default


def _to_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except Exception:
        return default


def _compact_output(
    payload: Dict[str, Any],
    *,
    include_full_records: bool = False,
    max_output_records: int = DEFAULT_MAX_OUTPUT_RECORDS,
    max_output_patterns: int = DEFAULT_MAX_OUTPUT_PATTERNS,
) -> Dict[str, Any]:
    """Keep Alpha outputs small enough to avoid Code Executor kill/timeouts.

    The Markdown contains the knowledge-ready output. Full normalized rows are
    usually not needed by Alpha Knowledge upload and can make Code Executor
    output too large, so they are omitted by default.
    """

    json_result = dict(payload.get("json_result") or {})
    patterns_result = dict(payload.get("alert_patterns_result") or {})

    records = list(json_result.get("records") or [])
    patterns = list(patterns_result.get("records") or [])

    if include_full_records:
        output_records = records
    else:
        output_records = records[: max(0, max_output_records)]
    json_result["records"] = output_records
    json_result["records_omitted"] = max(0, len(records) - len(output_records))

    output_patterns = patterns[: max(0, max_output_patterns)]
    patterns_result["records"] = output_patterns
    patterns_result["records_omitted"] = max(0, len(patterns) - len(output_patterns))

    compact_payload = dict(payload)
    compact_payload["json_result"] = json_result
    compact_payload["alert_patterns_result"] = patterns_result
    compact_payload["json_text"] = json.dumps(json_result, ensure_ascii=False, indent=2)
    compact_payload["alert_patterns_text"] = json.dumps(
        patterns_result, ensure_ascii=False, indent=2
    )
    return compact_payload


def _process_body_input(body: Any, source_label: str) -> Dict[str, Any]:
    if isinstance(body, dict):
        for key in ("source_label", "file_name", "filename", "name", "id"):
            if body.get(key):
                source_label = _clean_text(body.get(key))
                break
        for key in ("body", "content", "text", "data", "file_body"):
            if key in body and body[key] is not None:
                body = body[key]
                break

    if body is None or _clean_text(body) == "":
        return _build_error(
            "Empty CSV body",
            "zabbix_csv",
            rejected=[
                {
                    "row_number": None,
                    "source_ref": source_label,
                    "reason": "Empty CSV body",
                    "row_data": {},
                }
            ],
        )

    if isinstance(body, (bytes, bytearray)):
        body_text = bytes(body).decode("utf-8-sig", errors="replace")
    else:
        body_text = str(body)
        if body_text.startswith("\ufeff"):
            body_text = body_text.lstrip("\ufeff")

    if _looks_like_html(body_text):
        return _build_error(
            "HTTP body looks like HTML/login page; expected CSV export",
            "zabbix_csv",
            rejected=[
                {
                    "row_number": None,
                    "source_ref": source_label,
                    "reason": "HTTP body looks like HTML/login page; expected CSV export",
                    "row_data": {"body_prefix": body_text[:500]},
                }
            ],
        )

    reader = csv.reader(io.StringIO(body_text), delimiter=",", quotechar='"')
    return _process_rows(rows=reader, source_type="zabbix_csv", source_label=source_label)


def _process_file_input(file_obj: Any, source_label: Optional[str] = None) -> Dict[str, Any]:
    try:
        if isinstance(file_obj, (bytes, bytearray)):
            return _process_body_input(file_obj, source_label or "zabbix_export_file")

        if isinstance(file_obj, dict) and not source_label:
            for key in ("source_label", "file_name", "filename", "name", "id"):
                if file_obj.get(key):
                    source_label = _clean_text(file_obj.get(key))
                    break

        if isinstance(file_obj, dict):
            for key in ("body", "content", "text", "data", "file_body"):
                if key in file_obj and file_obj[key] is not None:
                    return _process_body_input(
                        file_obj[key], source_label or "zabbix_export_file"
                    )

        path = Path(_resolve_file_path(file_obj))
        if not path.exists():
            alt = Path.cwd() / str(path)
            if alt.exists():
                path = alt

        if not path.exists():
            raise FileNotFoundError(str(path))

        suffix = _suffix_from_hint(path, source_label)
        is_csv = suffix == ".csv" or (suffix == "" and _path_looks_like_csv(path))
        if is_csv:
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                return _process_rows(
                    rows=csv.reader(handle),
                    source_type="zabbix_csv",
                    source_label=source_label or path.name,
                )

        if suffix and suffix not in (".xlsx", ".xls", ".xlsm", ".xltx", ".xltm"):
            raise ValueError(
                f"Unsupported Zabbix export file suffix '{suffix}'. "
                "Expected .csv or .xlsx/.xlsm/.xltx/.xltm."
            )

        rows, detected_source_label = _rows_from_xlsx(path)
        return _process_rows(
            rows=rows,
            source_type="zabbix_xlsx",
            source_label=source_label or detected_source_label,
        )
    except Exception as exc:
        return _build_error(
            f"Failed to read Zabbix export: {exc}",
            "zabbix_export",
            rejected=[
                {
                    "row_number": None,
                    "source_ref": source_label or "zabbix_export_file",
                    "reason": f"Failed to read Zabbix export: {exc}",
                    "row_data": {"zabbix_export_file": str(file_obj)},
                }
            ],
        )


def _process_source_object(source: Any, index: int) -> Dict[str, Any]:
    if not isinstance(source, dict):
        return _process_body_input(source, f"zabbix_source_{index + 1}")

    label = ""
    for key in ("source_label", "file_name", "filename", "name", "id"):
        if source.get(key):
            label = _clean_text(source.get(key))
            break
    if not label:
        label = f"zabbix_source_{index + 1}"

    for key in ("body", "content", "text", "data", "file_body"):
        if key in source and source[key] is not None:
            return _process_body_input(source[key], label)

    for key in ("path", "file_path", "filepath", "file"):
        if key in source and source[key]:
            return _process_file_input(source[key], label)

    return _build_error(
        "Source object does not contain CSV body or file path",
        "zabbix_source",
        rejected=[
            {
                "row_number": None,
                "source_ref": label,
                "reason": "Source object does not contain CSV body or file path",
                "row_data": {"keys": sorted(source.keys())},
            }
        ],
    )


def main(
    zabbix_export_file=None,
    zabbix_export_body=None,
    zabbix_export_files=None,
    zabbix_export_bodies=None,
    source_label=None,
    source_labels=None,
    zabbix_sources=None,
    include_full_records=False,
    max_output_records=DEFAULT_MAX_OUTPUT_RECORDS,
    max_output_patterns=DEFAULT_MAX_OUTPUT_PATTERNS,
):
    def finalize(payload: Dict[str, Any]) -> Dict[str, Any]:
        return _compact_output(
            payload,
            include_full_records=_to_bool(include_full_records, default=False),
            max_output_records=_to_int(max_output_records, DEFAULT_MAX_OUTPUT_RECORDS),
            max_output_patterns=_to_int(max_output_patterns, DEFAULT_MAX_OUTPUT_PATTERNS),
        )

    processed_results: List[Dict[str, Any]] = []
    labels = _as_list(source_labels)

    for index, source in enumerate(_as_list(zabbix_sources)):
        processed_results.append(_process_source_object(source, index))

    for index, body in enumerate(_as_list(zabbix_export_bodies)):
        label = _label_at(labels, index, f"zabbix_csv_{index + 1}")
        processed_results.append(_process_body_input(body, label))

    file_label_offset = len(processed_results)
    for index, file_obj in enumerate(_as_list(zabbix_export_files)):
        label = _label_at(labels, file_label_offset + index, "")
        processed_results.append(_process_file_input(file_obj, label or None))

    if processed_results:
        return finalize(_merge_processed_results(processed_results))

    if zabbix_export_body is not None and _clean_text(zabbix_export_body) != "":
        return finalize(
            _process_body_input(
                zabbix_export_body, _clean_text(source_label) or "zabbix_csv"
            )
        )

    if zabbix_export_file:
        return finalize(
            _process_file_input(zabbix_export_file, _clean_text(source_label) or None)
        )

    return finalize(
        _build_error(
            "No valid CSV body or Zabbix file found",
            "zabbix_export",
            rejected=[
                {
                    "row_number": None,
                    "source_ref": "input",
                    "reason": "No valid CSV body or Zabbix file found",
                    "row_data": {
                        "zabbix_export_file": str(zabbix_export_file)
                        if zabbix_export_file is not None
                        else None,
                        "zabbix_export_body_type": type(zabbix_export_body).__name__
                        if zabbix_export_body is not None
                        else None,
                    },
                },
            ],
        )
    )
