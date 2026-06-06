from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "01-RawData"
OUTPUT_DIR = ROOT / "input"
OUTPUT_FILE = OUTPUT_DIR / "it-operations-master-data-template.xlsx"

INCIDENT_FILE = (
    RAW_DIR / "ISP Incident Report" / "SEA - Corp IT- ILL- Incident Report   (Responses).xlsx"
)
ZABBIX_FILE = RAW_DIR / "zbx_problems_export.xlsx"
ZABBIX_EXPORT_DIR = RAW_DIR / "Export zabbix"
ISSUE_FILE = RAW_DIR / "Ticket" / "IssueReport.xlsx"

STATUS_VALUES = ["CONFIRMED", "NEEDS_REVIEW", "DRAFT", "REJECTED"]
CRITICALITY_VALUES = ["critical", "high", "normal", "low"]
ACTIVE_VALUES = ["yes", "no"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")
CONFIRMED_FILL = PatternFill("solid", fgColor="E2F0D9")
REJECTED_FILL = PatternFill("solid", fgColor="F4CCCC")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


def normalized_rows(worksheet):
    rows = []
    for raw_row in worksheet.iter_rows(values_only=True):
        values = ["" if value is None else str(value).strip() for value in raw_row]
        if any(values):
            rows.append(values)
    width = max((len(row) for row in rows), default=0)
    return [row + [""] * (width - len(row)) for row in rows]


def add_table(worksheet, name):
    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return
    table = Table(displayName=name, ref=worksheet.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    worksheet.add_table(table)


def style_data_sheet(worksheet, widths):
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    worksheet.row_dimensions[1].height = 34
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width


def add_status_formatting(worksheet, status_column_letter):
    row_range = f"{status_column_letter}2:{status_column_letter}5000"
    worksheet.conditional_formatting.add(
        row_range,
        FormulaRule(
            formula=[f'${status_column_letter}2="NEEDS_REVIEW"'],
            fill=REVIEW_FILL,
        ),
    )
    worksheet.conditional_formatting.add(
        row_range,
        FormulaRule(
            formula=[f'${status_column_letter}2="CONFIRMED"'],
            fill=CONFIRMED_FILL,
        ),
    )
    worksheet.conditional_formatting.add(
        row_range,
        FormulaRule(
            formula=[f'${status_column_letter}2="REJECTED"'],
            fill=REJECTED_FILL,
        ),
    )


def add_list_validation(worksheet, column_letter, values_formula):
    validation = DataValidation(type="list", formula1=values_formula, allow_blank=True)
    validation.error = "Choose one of the allowed values from the dropdown."
    validation.errorTitle = "Invalid value"
    validation.prompt = "Select a value from the dropdown."
    validation.promptTitle = "Allowed values"
    worksheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}5000")


def parse_site_code(site_name):
    return site_name.split(" - ", 1)[0].strip()


def infer_site_type(site_name):
    lower = site_name.lower()
    if "warehouse" in lower or "warehosue" in lower:
        return "warehouse"
    if "soc" in lower:
        return "soc"
    if "office" in lower:
        return "office"
    return "unknown"


def infer_device_type(host_name):
    upper = host_name.upper()
    if "UPS" in upper:
        return "ups"
    if "WLC" in upper:
        return "wireless_controller"
    if "FW" in upper:
        return "firewall"
    if "SW" in upper:
        return "switch"
    if "GOOGLE" in upper:
        return "synthetic_probe"
    if "VSSJC" in upper or "VSEMS" in upper or "VSSPM" in upper:
        return "server"
    return "unknown"


def extract_incident_seed():
    workbook = load_workbook(INCIDENT_FILE, read_only=True, data_only=True)
    rows = normalized_rows(workbook["Form Responses 1"])
    headers = rows[0]
    data = rows[1:]
    index = {header: position for position, header in enumerate(headers)}

    site_names = {}
    site_counts = Counter()
    site_isp_counts = Counter()
    for row in data:
        site_name = row[index["SITE / Location"]]
        site_code = parse_site_code(site_name)
        isp = row[index["ISP"]]
        site_names.setdefault(site_code, site_name)
        site_counts[site_code] += 1
        site_isp_counts[(site_code, isp)] += 1
    return site_names, site_counts, site_isp_counts


def extract_zabbix_seed():
    if ZABBIX_EXPORT_DIR.exists():
        rows = []
        for path in sorted(ZABBIX_EXPORT_DIR.glob("*.csv"), key=lambda item: item.name.lower()):
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                file_rows = [
                    ["" if value is None else str(value).strip() for value in row]
                    for row in reader
                ]
            if not file_rows:
                continue
            if not rows:
                rows.extend(file_rows)
            else:
                rows.extend(file_rows[1:])
    else:
        workbook = load_workbook(ZABBIX_FILE, read_only=True, data_only=True)
        rows = normalized_rows(workbook.worksheets[0])
        workbook.close()
    headers = rows[0]
    data = rows[1:]
    index = {header: position for position, header in enumerate(headers)}

    hosts = defaultdict(
        lambda: {
            "site_code": "",
            "alert_count": 0,
            "classes": set(),
            "components": set(),
            "descriptions": set(),
        }
    )
    seen_rows = set()
    for row in data:
        dedupe_key = tuple(row[index[header]] for header in headers)
        if dedupe_key in seen_rows:
            continue
        seen_rows.add(dedupe_key)
        host_name = row[index["Host"]]
        match = re.match(r"^VNM([A-Z0-9]{3})", host_name)
        host = hosts[host_name]
        host["site_code"] = match.group(1) if match else ""
        host["alert_count"] += 1
        for token in row[index["Tags"]].split(", "):
            if ": " not in token:
                continue
            key, value = token.split(": ", 1)
            if key == "class":
                host["classes"].add(value)
            elif key == "component":
                host["components"].add(value)
            elif key == "description":
                host["descriptions"].add(value)
    return hosts


def extract_user_group_seed():
    workbook = load_workbook(ISSUE_FILE, read_only=True, data_only=True)
    rows = normalized_rows(workbook.worksheets[0])
    headers = rows[0]
    data = rows[1:]
    index = {header: position for position, header in enumerate(headers)}

    unique_tickets = {}
    for row in data:
        unique_tickets.setdefault(row[index["itcenter.ticket.id"]], row)

    counts = Counter()
    for row in unique_tickets.values():
        counts[
            (
                row[index["itcenter.ticket.service_recipient.user.business_unit"]],
                row[index["itcenter.ticket.service_recipient.user.department_name"]],
            )
        ] += 1
    return counts


def create_readme(workbook):
    worksheet = workbook.active
    worksheet.title = "README"
    rows = [
        ["IT Operations Intelligence - Master Data Template"],
        [""],
        ["Mục đích"],
        ["Workbook này bổ sung context còn thiếu trong 3 raw source để PoC xác định impact, recurrence và confidence có thể kiểm chứng."],
        [""],
        ["Cách sử dụng"],
        ["1. Ưu tiên review sheet sites và hosts. Chuyển review_status từ NEEDS_REVIEW thành CONFIRMED khi dữ liệu đúng."],
        ["2. Điền network_links và user_groups ở mức bạn hiện biết. Các ô chưa rõ có thể để trống."],
        ["3. Điền services và service_dependencies dần để cải thiện impact analysis."],
        ["4. Không xóa cột source_reference hoặc inference_reason: chúng giúp audit nguồn gốc dữ liệu."],
        ["5. Khi hoàn tất vòng đầu, gửi lại workbook để pipeline PoC ingest."],
        [""],
        ["Quy ước review_status"],
        ["CONFIRMED", "Đã được IT Ops xác nhận."],
        ["NEEDS_REVIEW", "Seed từ raw data hoặc naming convention, cần kiểm tra."],
        ["DRAFT", "Khai báo mới nhưng chưa xác nhận."],
        ["REJECTED", "Không dùng cho pipeline ingest."],
        [""],
        ["Thứ tự ưu tiên điền"],
        ["1", "sites", "Bắt buộc"],
        ["2", "hosts", "Bắt buộc"],
        ["3", "network_links", "Nên có cho network impact"],
        ["4", "user_groups", "Nên có cho user impact"],
        ["5", "services", "Có thể bổ sung dần"],
        ["6", "service_dependencies", "Có thể bổ sung dần"],
    ]
    for row in rows:
        worksheet.append(row)
    worksheet["A1"].font = Font(size=16, bold=True, color="1F4E78")
    for row_number in [3, 6, 13, 19]:
        worksheet[f"A{row_number}"].font = Font(bold=True, color="1F4E78")
        worksheet[f"A{row_number}"].fill = SUBHEADER_FILL
    worksheet.column_dimensions["A"].width = 110
    worksheet.column_dimensions["B"].width = 34
    worksheet.column_dimensions["C"].width = 20
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def create_sites_sheet(workbook, site_names, site_counts, hosts):
    worksheet = workbook.create_sheet("sites")
    worksheet.append(
        [
            "site_code",
            "site_name",
            "site_type",
            "location",
            "business_unit",
            "criticality",
            "active",
            "review_status",
            "source_reference",
            "notes",
        ]
    )
    all_codes = set(site_names)
    all_codes.update(host["site_code"] for host in hosts.values() if host["site_code"])
    for site_code in sorted(all_codes):
        site_name = site_names.get(site_code, "")
        source = "incident_form" if site_code in site_names else "zabbix_host_prefix_inference"
        notes = f"{site_counts[site_code]} confirmed incident rows" if site_code in site_counts else "Site code inferred from Zabbix hostname; fill site_name and location."
        worksheet.append(
            [
                site_code,
                site_name,
                infer_site_type(site_name),
                "",
                "",
                "normal",
                "yes",
                "NEEDS_REVIEW",
                source,
                notes,
            ]
        )
    style_data_sheet(
        worksheet,
        {"A": 16, "B": 40, "C": 18, "D": 24, "E": 20, "F": 14, "G": 10, "H": 16, "I": 30, "J": 54},
    )
    add_list_validation(worksheet, "F", "=validation_lists!$B$2:$B$5")
    add_list_validation(worksheet, "G", "=validation_lists!$C$2:$C$3")
    add_list_validation(worksheet, "H", "=validation_lists!$A$2:$A$5")
    add_status_formatting(worksheet, "H")
    add_table(worksheet, "SitesTable")


def create_hosts_sheet(workbook, hosts):
    worksheet = workbook.create_sheet("hosts")
    worksheet.append(
        [
            "host_name",
            "site_code",
            "device_type",
            "domain",
            "components",
            "isp_hint",
            "service_code",
            "criticality",
            "active",
            "review_status",
            "inference_reason",
            "source_reference",
            "notes",
        ]
    )
    for host_name, host in sorted(hosts.items(), key=lambda item: (item[1]["site_code"], item[0])):
        worksheet.append(
            [
                host_name,
                host["site_code"],
                infer_device_type(host_name),
                ",".join(sorted(host["classes"])),
                ",".join(sorted(host["components"])),
                ",".join(sorted(host["descriptions"])),
                "",
                "normal",
                "yes",
                "NEEDS_REVIEW",
                "site_code inferred from VNM<site> hostname prefix; device_type inferred from hostname; domain/components from Zabbix tags",
                "zbx_problems_export.xlsx",
                f'{host["alert_count"]} alert rows in provided export',
            ]
        )
    style_data_sheet(
        worksheet,
        {"A": 28, "B": 14, "C": 22, "D": 18, "E": 24, "F": 20, "G": 22, "H": 14, "I": 10, "J": 16, "K": 74, "L": 28, "M": 32},
    )
    add_list_validation(worksheet, "H", "=validation_lists!$B$2:$B$5")
    add_list_validation(worksheet, "I", "=validation_lists!$C$2:$C$3")
    add_list_validation(worksheet, "J", "=validation_lists!$A$2:$A$5")
    add_status_formatting(worksheet, "J")
    add_table(worksheet, "HostsTable")


def create_network_links_sheet(workbook, site_isp_counts):
    worksheet = workbook.create_sheet("network_links")
    worksheet.append(
        [
            "link_code",
            "site_code",
            "isp",
            "link_role",
            "service_type",
            "criticality",
            "active",
            "review_status",
            "source_reference",
            "notes",
        ]
    )
    for (site_code, isp), count in sorted(site_isp_counts.items()):
        worksheet.append(
            [
                f"{site_code}-{isp}-TBD",
                site_code,
                isp,
                "unknown",
                "site_connectivity",
                "normal",
                "yes",
                "NEEDS_REVIEW",
                "incident_form",
                f"{count} incident rows mention this site/ISP pair. Replace TBD with actual circuit or link identifier if available.",
            ]
        )
    style_data_sheet(
        worksheet,
        {"A": 24, "B": 14, "C": 14, "D": 18, "E": 22, "F": 14, "G": 10, "H": 16, "I": 22, "J": 72},
    )
    add_list_validation(worksheet, "F", "=validation_lists!$B$2:$B$5")
    add_list_validation(worksheet, "G", "=validation_lists!$C$2:$C$3")
    add_list_validation(worksheet, "H", "=validation_lists!$A$2:$A$5")
    add_status_formatting(worksheet, "H")
    add_table(worksheet, "NetworkLinksTable")


def create_services_sheet(workbook):
    worksheet = workbook.create_sheet("services")
    worksheet.append(
        [
            "service_code",
            "service_name",
            "domain",
            "criticality",
            "active",
            "review_status",
            "description",
            "notes",
        ]
    )
    rows = [
        ["site-connectivity", "Site network connectivity", "network", "high", "yes", "NEEDS_REVIEW", "WAN or ISP connectivity serving a site.", "Seed service; confirm naming and criticality."],
        ["site-power", "Site power infrastructure", "power", "high", "yes", "NEEDS_REVIEW", "UPS and power-related infrastructure.", "Seed service; confirm scope."],
        ["site-lan", "Site LAN", "network", "normal", "yes", "NEEDS_REVIEW", "Switching and internal site network.", "Seed service; confirm scope."],
        ["site-wifi", "Site Wi-Fi", "network", "normal", "yes", "NEEDS_REVIEW", "Wireless connectivity including PDA access.", "Seed service; confirm scope."],
        ["remote-support", "Remote support agent", "application", "normal", "yes", "NEEDS_REVIEW", "Remote support client availability.", "Seed service; confirm scope."],
        ["site-server", "Site server workload", "server", "normal", "yes", "NEEDS_REVIEW", "Site-local server or operating-system workload.", "Seed service; confirm scope."],
    ]
    for row in rows:
        worksheet.append(row)
    style_data_sheet(
        worksheet,
        {"A": 22, "B": 34, "C": 18, "D": 14, "E": 10, "F": 16, "G": 64, "H": 42},
    )
    add_list_validation(worksheet, "D", "=validation_lists!$B$2:$B$5")
    add_list_validation(worksheet, "E", "=validation_lists!$C$2:$C$3")
    add_list_validation(worksheet, "F", "=validation_lists!$A$2:$A$5")
    add_status_formatting(worksheet, "F")
    add_table(worksheet, "ServicesTable")


def create_user_groups_sheet(workbook, group_counts):
    worksheet = workbook.create_sheet("user_groups")
    worksheet.append(
        [
            "group_code",
            "business_unit",
            "department",
            "site_code",
            "service_code",
            "estimated_user_count",
            "criticality",
            "active",
            "review_status",
            "source_reference",
            "notes",
        ]
    )
    for (business_unit, department), ticket_count in group_counts.most_common():
        group_code = re.sub(r"[^a-z0-9]+", "-", f"{business_unit}-{department}".lower()).strip("-")
        worksheet.append(
            [
                group_code,
                business_unit,
                department,
                "",
                "",
                "",
                "normal",
                "yes",
                "NEEDS_REVIEW",
                "IssueReport.xlsx",
                f"{ticket_count} unique tickets in provided export. Fill site_code only if this group belongs to a specific site.",
            ]
        )
    style_data_sheet(
        worksheet,
        {"A": 40, "B": 22, "C": 38, "D": 14, "E": 22, "F": 22, "G": 14, "H": 10, "I": 16, "J": 22, "K": 76},
    )
    add_list_validation(worksheet, "G", "=validation_lists!$B$2:$B$5")
    add_list_validation(worksheet, "H", "=validation_lists!$C$2:$C$3")
    add_list_validation(worksheet, "I", "=validation_lists!$A$2:$A$5")
    add_status_formatting(worksheet, "I")
    add_table(worksheet, "UserGroupsTable")


def create_dependencies_sheet(workbook):
    worksheet = workbook.create_sheet("service_dependencies")
    worksheet.append(
        [
            "dependency_id",
            "upstream_type",
            "upstream_id",
            "downstream_type",
            "downstream_id",
            "impact_description",
            "active",
            "review_status",
            "notes",
        ]
    )
    worksheet.append(
        [
            "EXAMPLE-DELETE-ME",
            "network_link",
            "MS2-VNPT-TBD",
            "service",
            "site-connectivity",
            "If the WAN link fails, site connectivity may be degraded or unavailable.",
            "no",
            "DRAFT",
            "Example only. Replace with confirmed dependencies and set active=yes.",
        ]
    )
    style_data_sheet(
        worksheet,
        {"A": 24, "B": 20, "C": 26, "D": 20, "E": 28, "F": 74, "G": 10, "H": 16, "I": 70},
    )
    add_list_validation(worksheet, "G", "=validation_lists!$C$2:$C$3")
    add_list_validation(worksheet, "H", "=validation_lists!$A$2:$A$5")
    add_status_formatting(worksheet, "H")
    add_table(worksheet, "ServiceDependenciesTable")


def create_dictionary_sheet(workbook):
    worksheet = workbook.create_sheet("field_dictionary")
    worksheet.append(["sheet_name", "field_name", "required_for_poc", "description", "allowed_values_or_example"])
    rows = [
        ["sites", "site_code", "yes", "Stable short code used to join incidents, hosts and links.", "MS2, CCW, SNT"],
        ["sites", "criticality", "yes", "Operational importance of the site.", ", ".join(CRITICALITY_VALUES)],
        ["hosts", "host_name", "yes", "Exact monitored hostname from Zabbix.", "VNMCCW-PNWSW01"],
        ["hosts", "site_code", "yes", "Site served by the monitored host.", "CCW"],
        ["hosts", "service_code", "recommended", "Service impacted when the host has a confirmed issue.", "site-lan"],
        ["network_links", "link_code", "recommended", "Stable circuit or logical-link identifier.", "MS2-VNPT-primary"],
        ["network_links", "link_role", "recommended", "Purpose of the circuit within the site.", "primary, backup, unknown"],
        ["user_groups", "site_code", "recommended", "Populate only when a group maps to a specific site.", "MS2"],
        ["user_groups", "estimated_user_count", "optional", "Used for estimated impact when exact user evidence is unavailable.", "250"],
        ["service_dependencies", "upstream_id", "optional", "Entity whose failure may affect another entity.", "MS2-VNPT-primary"],
        ["service_dependencies", "downstream_id", "optional", "Entity potentially impacted by the upstream entity.", "site-connectivity"],
        ["*", "review_status", "yes", "Human validation state. Only CONFIRMED rows should be treated as authoritative.", ", ".join(STATUS_VALUES)],
    ]
    for row in rows:
        worksheet.append(row)
    style_data_sheet(worksheet, {"A": 24, "B": 26, "C": 18, "D": 82, "E": 42})
    add_table(worksheet, "FieldDictionaryTable")


def create_validation_lists(workbook):
    worksheet = workbook.create_sheet("validation_lists")
    worksheet.append(["review_status", "criticality", "active"])
    max_rows = max(len(STATUS_VALUES), len(CRITICALITY_VALUES), len(ACTIVE_VALUES))
    for index in range(max_rows):
        worksheet.append(
            [
                STATUS_VALUES[index] if index < len(STATUS_VALUES) else "",
                CRITICALITY_VALUES[index] if index < len(CRITICALITY_VALUES) else "",
                ACTIVE_VALUES[index] if index < len(ACTIVE_VALUES) else "",
            ]
        )
    style_data_sheet(worksheet, {"A": 18, "B": 16, "C": 12})
    worksheet.sheet_state = "hidden"


def main():
    site_names, site_counts, site_isp_counts = extract_incident_seed()
    hosts = extract_zabbix_seed()
    user_group_counts = extract_user_group_seed()

    workbook = Workbook()
    create_readme(workbook)
    create_validation_lists(workbook)
    create_sites_sheet(workbook, site_names, site_counts, hosts)
    create_hosts_sheet(workbook, hosts)
    create_network_links_sheet(workbook, site_isp_counts)
    create_services_sheet(workbook)
    create_user_groups_sheet(workbook, user_group_counts)
    create_dependencies_sheet(workbook)
    create_dictionary_sheet(workbook)

    workbook._sheets = [
        workbook["README"],
        workbook["sites"],
        workbook["hosts"],
        workbook["network_links"],
        workbook["services"],
        workbook["user_groups"],
        workbook["service_dependencies"],
        workbook["field_dictionary"],
        workbook["validation_lists"],
    ]
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_FILE)
    print(f"Created: {OUTPUT_FILE}")
    print(f"Sites: {workbook['sites'].max_row - 1}")
    print(f"Hosts: {workbook['hosts'].max_row - 1}")
    print(f"Network links: {workbook['network_links'].max_row - 1}")
    print(f"User groups: {workbook['user_groups'].max_row - 1}")


if __name__ == "__main__":
    main()
